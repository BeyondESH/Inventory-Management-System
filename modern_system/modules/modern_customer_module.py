#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
现代化客户管理模块
"""

import tkinter as tk
from tkinter import ttk, messagebox
from typing import Dict, List, Any, Optional
import datetime
import json

# 导入数据管理中心
try:
    from .data_manager import data_manager
except ImportError:
    try:
        from data_manager import data_manager
    except ImportError:
        # 模拟数据管理器
        class MockDataManager:
            def load_data(self, data_type):
                return []
            def register_module(self, module_type, instance):
                pass
        data_manager = MockDataManager()

class ModernCustomerModule:
    def __init__(self, parent_frame, title_frame):
        self.parent_frame = parent_frame
        self.title_frame = title_frame
        
        # 注册到数据管理中心
        data_manager.register_module('customer', self)
        
        # 现代化配色方案
        self.colors = {
            'primary': '#FF6B35',      # 主色调
            'secondary': '#F7931E',    # 次色调
            'accent': '#FFD23F',       # 强调色
            'background': '#F8F9FA',   # 背景色
            'surface': '#FFFFFF',      # 卡片背景
            'text_primary': '#2D3436', # 主文字
            'text_secondary': '#636E72', # 次文字
            'border': '#E0E0E0',       # 边框
            'success': '#00B894',      # 成功色
            'warning': '#FDCB6E',      # 警告色
            'error': '#E84393',        # 错误色
            'card_shadow': '#F0F0F0',  # 卡片阴影
            'white': '#FFFFFF',        # 白色
            'info': '#3498DB',         # 信息色
            'danger': '#E74C3C'        # 危险色
        }
        
        # 字体配置
        self.fonts = {
            'title': ('Microsoft YaHei UI', 16, 'bold'),
            'heading': ('Microsoft YaHei UI', 14, 'bold'),
            'body': ('Microsoft YaHei UI', 12),
            'small': ('Microsoft YaHei UI', 10)
        }
        
        self.main_frame = None
        self.customer_data = self.load_customer_data()
        
    def load_customer_data(self):
        """从数据库加载客户数据"""
        try:
            # 数据已在data_manager中处理好字段名和聚合信息
            customers = data_manager.get_customers()
            if not customers:
                print("⚠️ 未找到客户数据，将使用默认示例。")
                return [
                    {"id": "CUST001", "name": "张三 (示例)", "phone": "13800000000", "address": "示例地址1", "total_orders": 0, "total_amount": 0.0},
                    {"id": "CUST002", "name": "李四 (示例)", "phone": "13900000000", "address": "示例地址2", "total_orders": 0, "total_amount": 0.0},
                ]
            return customers
        except Exception as e:
            messagebox.showerror("错误", f"加载客户数据失败: {e}")
            return []
    
    def refresh_customer_data(self):
        """刷新客户数据"""
        self.customer_data = self.load_customer_data()
        self.refresh_customer_list()
    
    def refresh_customer_list(self):
        """刷新客户列表显示"""
        if hasattr(self, 'customer_tree'):
            # 清空现有数据
            for item in self.customer_tree.get_children():
                self.customer_tree.delete(item)
            
            # 重新插入数据
            for customer in self.customer_data:
                values = (
                    customer.get('id', ''),
                    customer.get('name', ''),
                    customer.get('phone', ''),
                    customer.get('address', ''),
                    customer.get('total_orders', 0),
                    f"￥{customer.get('total_amount', 0):.0f}"
                )
                self.customer_tree.insert("", "end", values=values)
        
    def create_customer_buttons(self, parent):
        """创建客户操作按钮"""
        button_frame = tk.Frame(parent, bg=self.colors['surface'])
        button_frame.pack(fill="x", padx=10, pady=(0, 10))
        
        # 添加客户按钮
        add_btn = tk.Button(button_frame, text="➕ 添加客户", 
                           font=self.fonts['body'],
                           bg=self.colors['success'], fg='white',
                           bd=0, pady=8, padx=15, cursor="hand2",
                           command=self.add_customer)
        add_btn.pack(side="left", padx=(0, 10))
        
        # 编辑客户按钮
        edit_btn = tk.Button(button_frame, text="✏️ 编辑客户", 
                            font=self.fonts['body'],
                            bg=self.colors['info'], fg='white',
                            bd=0, pady=8, padx=15, cursor="hand2",
                            command=self.edit_customer)
        edit_btn.pack(side="left", padx=(0, 10))
        
        # 删除客户按钮
        delete_btn = tk.Button(button_frame, text="🗑️ 删除客户", 
                              font=self.fonts['body'],
                              bg=self.colors['danger'], fg='white',
                              bd=0, pady=8, padx=15, cursor="hand2",
                              command=self.delete_customer)
        delete_btn.pack(side="left", padx=(0, 10))
        
        # 刷新按钮
        refresh_btn = tk.Button(button_frame, text="🔄 刷新", 
                               font=self.fonts['body'],
                               bg=self.colors['secondary'], fg='white',
                               bd=0, pady=8, padx=15, cursor="hand2",
                               command=self.refresh_customer_data)
        refresh_btn.pack(side="right")
        
        # 导出按钮
        export_btn = tk.Button(button_frame, text="📊 导出", 
                              font=self.fonts['body'],
                              bg=self.colors['success'], fg='white',
                              bd=0, pady=8, padx=15, cursor="hand2",
                              command=self.export_customers)
        export_btn.pack(side="right")
    
    def add_customer(self):
        """添加客户"""
        try:
            # 创建输入对话框
            dialog = tk.Toplevel(self.main_frame)
            dialog.title("添加客户")
            dialog.geometry("400x350")
            dialog.configure(bg=self.colors['background'])
            dialog.transient(self.main_frame)
            dialog.grab_set()
            
            # 居中显示
            dialog.update_idletasks()
            x = (dialog.winfo_screenwidth() // 2) - (200)
            y = (dialog.winfo_screenheight() // 2) - (175)
            dialog.geometry(f"400x350+{x}+{y}")
            
            # 输入字段
            tk.Label(dialog, text="添加客户", font=self.fonts['heading'],
                    bg=self.colors['background'], fg=self.colors['text_primary']).pack(pady=10)
            
            # 姓名
            tk.Label(dialog, text="姓名:", bg=self.colors['background'], 
                    fg=self.colors['text_primary']).pack(anchor="w", padx=20)
            name_var = tk.StringVar(dialog)
            name_entry = tk.Entry(dialog, textvariable=name_var, font=self.fonts['body'])
            name_entry.pack(fill="x", padx=20, pady=5)
            
            # 电话
            tk.Label(dialog, text="电话:", bg=self.colors['background'], 
                    fg=self.colors['text_primary']).pack(anchor="w", padx=20)
            phone_var = tk.StringVar(dialog)
            phone_entry = tk.Entry(dialog, textvariable=phone_var, font=self.fonts['body'])
            phone_entry.pack(fill="x", padx=20, pady=5)
            
            # 地址
            tk.Label(dialog, text="地址:", bg=self.colors['background'], 
                    fg=self.colors['text_primary']).pack(anchor="w", padx=20)
            address_var = tk.StringVar(dialog)
            address_entry = tk.Entry(dialog, textvariable=address_var, font=self.fonts['body'])
            address_entry.pack(fill="x", padx=20, pady=5)
            
            # 按钮
            btn_frame = tk.Frame(dialog, bg=self.colors['background'])
            btn_frame.pack(fill="x", padx=20, pady=20)
            
            def save_customer():
                try:
                    name = name_var.get().strip()
                    phone = phone_var.get().strip()
                    address = address_var.get().strip()
                    
                    if not name or not phone:
                        messagebox.showerror("错误", "请填写姓名和电话", parent=dialog)
                        return
                    
                    # 准备数据
                    first_name = name.split(' ')[0] if ' ' in name else name
                    last_name = name.split(' ')[1] if ' ' in name else ''
                    
                    customer_data = {
                        'first_name': first_name,
                        'last_name': last_name,
                        'phone': phone,
                        'address': address,
                    }
                    
                    data_manager.add_customer(customer_data)
                    messagebox.showinfo("成功", "客户添加成功！", parent=dialog)
                    dialog.destroy()
                    self.refresh_customer_data()

                except Exception as e:
                    messagebox.showerror("错误", f"添加客户失败: {e}", parent=dialog)
            
            tk.Button(btn_frame, text="保存", command=save_customer,
                     bg=self.colors['success'], fg='white', bd=0, pady=8, padx=20).pack(side="left")
            tk.Button(btn_frame, text="取消", command=dialog.destroy,
                     bg=self.colors['text_secondary'], fg='white', bd=0, pady=8, padx=20).pack(side="right")
                     
        except Exception as e:
            messagebox.showerror("错误", f"打开添加客户对话框失败：{e}")
    
    def edit_customer(self):
        """编辑客户"""
        try:
            # 获取选中的客户
            selected_item = self.customer_tree.selection()
            if not selected_item:
                messagebox.showwarning("提示", "请先选择一个客户")
                return
            
            # 获取客户数据
            item_values = self.customer_tree.item(selected_item[0])['values']
            customer_id = item_values[0]
            
            # 查找客户数据
            customer = None
            for c in self.customer_data:
                if str(c.get('id', '')) == str(customer_id):
                    customer = c
                    break
            
            if not customer:
                messagebox.showerror("错误", "未找到客户数据")
                return
            
            # 创建编辑对话框
            dialog = tk.Toplevel(self.main_frame)
            dialog.title("编辑客户")
            dialog.geometry("400x350")
            dialog.configure(bg=self.colors['background'])
            dialog.transient(self.main_frame)
            dialog.grab_set()
            
            # 居中显示
            dialog.update_idletasks()
            x = (dialog.winfo_screenwidth() // 2) - (200)
            y = (dialog.winfo_screenheight() // 2) - (175)
            dialog.geometry(f"400x350+{x}+{y}")
            
            # 输入字段
            tk.Label(dialog, text="编辑客户", font=self.fonts['heading'],
                    bg=self.colors['background'], fg=self.colors['text_primary']).pack(pady=10)
            
            # 姓名
            tk.Label(dialog, text="姓名:", bg=self.colors['background'], 
                    fg=self.colors['text_primary']).pack(anchor="w", padx=20)
            name_var = tk.StringVar(dialog, value=customer.get('name', ''))
            name_entry = tk.Entry(dialog, textvariable=name_var, font=self.fonts['body'])
            name_entry.pack(fill="x", padx=20, pady=5)
            
            # 电话
            tk.Label(dialog, text="电话:", bg=self.colors['background'], 
                    fg=self.colors['text_primary']).pack(anchor="w", padx=20)
            phone_var = tk.StringVar(dialog, value=customer.get('phone', ''))
            phone_entry = tk.Entry(dialog, textvariable=phone_var, font=self.fonts['body'])
            phone_entry.pack(fill="x", padx=20, pady=5)
            
            # 地址
            tk.Label(dialog, text="地址:", bg=self.colors['background'], 
                    fg=self.colors['text_primary']).pack(anchor="w", padx=20)
            address_var = tk.StringVar(dialog, value=customer.get('address', ''))
            address_entry = tk.Entry(dialog, textvariable=address_var, font=self.fonts['body'])
            address_entry.pack(fill="x", padx=20, pady=5)
            
            # 按钮
            btn_frame = tk.Frame(dialog, bg=self.colors['background'])
            btn_frame.pack(fill="x", padx=20, pady=20)
            
            def update_customer():
                try:
                    name = name_var.get().strip()
                    phone = phone_var.get().strip()
                    address = address_var.get().strip()

                    if not name or not phone:
                        messagebox.showerror("错误", "姓名和电话不能为空", parent=dialog)
                        return
                    
                    first_name = name.split(' ')[0] if ' ' in name else name
                    last_name = name.split(' ')[1] if ' ' in name else ''

                    update_data = {
                        'first_name': first_name,
                        'last_name': last_name,
                        'phone': phone,
                        'address': address,
                    }
                    
                    data_manager.update_customer(customer_id, update_data)
                    messagebox.showinfo("成功", "客户信息更新成功！", parent=dialog)
                    dialog.destroy()
                    self.refresh_customer_data()

                except Exception as e:
                    messagebox.showerror("错误", f"更新客户失败: {e}", parent=dialog)
            
            tk.Button(btn_frame, text="更新", command=update_customer,
                     bg=self.colors['success'], fg='white', bd=0, pady=8, padx=20).pack(side="left")
            tk.Button(btn_frame, text="取消", command=dialog.destroy,
                     bg=self.colors['text_secondary'], fg='white', bd=0, pady=8, padx=20).pack(side="right")
                     
        except Exception as e:
            messagebox.showerror("错误", f"编辑客户失败：{e}")
    
    def delete_customer(self):
        """删除客户"""
        selected_item = self.customer_tree.selection()
        if not selected_item:
            messagebox.showwarning("提示", "请选择要删除的客户")
            return
        
        customer_id_str = self.customer_tree.item(selected_item[0])['values'][0]
        customer_name = self.customer_tree.item(selected_item[0])['values'][1]
        
        if messagebox.askyesno("确认删除", f"确定要删除客户 '{customer_name}' 吗？此操作不可逆！"):
            try:
                customer_id = int(customer_id_str)
                data_manager.delete_customer(customer_id)
                messagebox.showinfo("成功", "客户已删除。")
                self.refresh_customer_data()
            except ValueError as ve:
                 messagebox.showerror("删除失败", str(ve))
            except Exception as e:
                messagebox.showerror("删除失败", f"删除客户时发生错误: {e}")
    
    def show(self):
        """显示客户管理界面"""
        if self.main_frame:
            self.main_frame.destroy()
        
        self.main_frame = tk.Frame(self.parent_frame, bg=self.colors['background'])
        self.main_frame.pack(fill="both", expand=True)
        
        # 标题
        title_label = tk.Label(self.main_frame, text="👥 客户管理", 
                              font=self.fonts['title'],
                              bg=self.colors['background'], 
                              fg=self.colors['text_primary'])
        title_label.pack(pady=(0, 20))
        
        # 客户统计
        self.create_customer_stats()
        
        # 客户列表
        self.create_customer_list()
        
    def create_customer_stats(self):
        """创建客户统计"""
        stats_frame = tk.Frame(self.main_frame, bg=self.colors['background'])
        stats_frame.pack(fill="x", pady=(0, 20))
        
        # 统计数据
        total_customers = len(self.customer_data)
        total_orders = sum(customer.get('total_orders', 0) for customer in self.customer_data)
        total_amount = sum(customer.get('total_amount', 0) for customer in self.customer_data)
        avg_amount = total_amount / total_customers if total_customers > 0 else 0
        
        stats = [
            {"title": "客户总数", "value": str(total_customers), "icon": "👥", "color": self.colors['primary']},
            {"title": "总订单数", "value": str(total_orders), "icon": "📋", "color": self.colors['info']},
            {"title": "总消费额", "value": f"￥{total_amount:.0f}", "icon": "💰", "color": self.colors['success']},
            {"title": "平均消费", "value": f"￥{avg_amount:.0f}", "icon": "📊", "color": self.colors['secondary']}
        ]
        
        for stat in stats:
            card = tk.Frame(stats_frame, bg=self.colors['surface'], relief="flat", bd=1)
            card.pack(side="left", fill="both", expand=True, padx=(0, 10))
            
            # 图标
            icon_label = tk.Label(card, text=stat['icon'], font=('Segoe UI Emoji', 24),
                                bg=self.colors['surface'], fg=stat['color'])
            icon_label.pack(pady=(10, 5))
            
            # 数值
            value_label = tk.Label(card, text=stat['value'], font=self.fonts['heading'],
                                 bg=self.colors['surface'], fg=self.colors['text_primary'])
            value_label.pack()
            
            # 标题
            title_label = tk.Label(card, text=stat['title'], font=self.fonts['body'],
                                 bg=self.colors['surface'], fg=self.colors['text_secondary'])
            title_label.pack(pady=(5, 10))
            
    def create_customer_list(self):
        """创建客户列表"""
        list_frame = tk.Frame(self.main_frame, bg=self.colors['surface'])
        list_frame.pack(fill="both", expand=True)
        
        # 表格标题
        list_title = tk.Label(list_frame, text="📋 客户列表", 
                             font=self.fonts['heading'],
                             bg=self.colors['surface'], 
                             fg=self.colors['text_primary'])
        list_title.pack(pady=10)
        
        # 创建Treeview表格
        columns = ("客户ID", "姓名", "电话", "地址", "订单数", "消费金额")
        self.customer_tree = ttk.Treeview(list_frame, columns=columns, show='headings', height=15)
        
        # 设置列标题
        for col in columns:
            self.customer_tree.heading(col, text=col)
            if col == "地址":
                self.customer_tree.column(col, width=200, anchor="w")
            else:
                self.customer_tree.column(col, width=120, anchor="center")
        
        # 添加客户数据
        for customer in self.customer_data:
            values = (
                customer.get('id', ''),
                customer.get('name', ''),
                customer.get('phone', ''),
                customer.get('address', ''),
                customer.get('total_orders', 0),
                f"￥{customer.get('total_amount', 0):.0f}"
            )
            self.customer_tree.insert("", "end", values=values)
        
        # 添加滚动条
        scrollbar = ttk.Scrollbar(list_frame, orient="vertical", command=self.customer_tree.yview)
        self.customer_tree.configure(yscrollcommand=scrollbar.set)
        
        # 布局
        self.customer_tree.pack(side="left", fill="both", expand=True, padx=10, pady=10)
        scrollbar.pack(side="right", fill="y", pady=10)
        
        # 添加操作按钮
        self.create_customer_buttons(list_frame)

    def export_customers(self):
        """导出客户数据"""
        try:
            from tkinter import filedialog
            import datetime
            
            # 创建导出选择对话框
            dialog = tk.Toplevel(self.parent_frame)
            dialog.title("导出客户数据")
            dialog.geometry("400x300")
            dialog.configure(bg=self.colors['background'])
            dialog.transient(self.parent_frame)
            dialog.grab_set()
            
            # 居中显示
            dialog.update_idletasks()
            x = (dialog.winfo_screenwidth() // 2) - (200)
            y = (dialog.winfo_screenheight() // 2) - (150)
            dialog.geometry(f"400x300+{x}+{y}")
            
            # 标题
            tk.Label(dialog, text="导出客户数据", font=('Microsoft YaHei UI', 14, 'bold'),
                    bg=self.colors['background'], fg=self.colors['text']).pack(pady=15)
            
            # 导出选项框架
            options_frame = tk.Frame(dialog, bg=self.colors['background'])
            options_frame.pack(fill="both", expand=True, padx=20, pady=10)
            
            # 导出格式选择
            tk.Label(options_frame, text="选择导出格式:", font=('Microsoft YaHei UI', 12),
                    bg=self.colors['background'], fg=self.colors['text']).pack(anchor="w", pady=(0, 10))
            
            format_var = tk.StringVar(dialog, value="Excel")
            format_options = ["Excel", "CSV", "PDF"]
            
            format_frame = tk.Frame(options_frame, bg=self.colors['background'])
            format_frame.pack(anchor="w")
            
            for i, fmt in enumerate(format_options):
                rb = tk.Radiobutton(format_frame, text=fmt, variable=format_var, value=fmt,
                                  font=('Microsoft YaHei UI', 10), bg=self.colors['background'], 
                                  fg=self.colors['text'], selectcolor=self.colors['surface'])
                rb.grid(row=0, column=i, sticky="w", padx=(0, 20))
            
            # 客户类型筛选
            tk.Label(options_frame, text="客户类型:", font=('Microsoft YaHei UI', 12),
                    bg=self.colors['background'], fg=self.colors['text']).pack(anchor="w", pady=(20, 10))
            
            type_var = tk.StringVar(dialog, value="全部")
            type_options = ["全部", "个人", "企业"]
            
            type_combo = ttk.Combobox(options_frame, textvariable=type_var, 
                                    values=type_options, state="readonly", width=20)
            type_combo.pack(anchor="w")
            
            # 按钮框架
            btn_frame = tk.Frame(dialog, bg=self.colors['background'])
            btn_frame.pack(fill="x", padx=20, pady=20)
            
            def do_export():
                try:
                    file_format = format_var.get()
                    customer_type = type_var.get()
                    
                    # 获取当前时间戳
                    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                    filename = f"客户数据_{customer_type}_{timestamp}"
                    
                    # 选择保存路径
                    if file_format == "Excel":
                        file_path = filedialog.asksaveasfilename(
                            defaultextension=".xlsx",
                            filetypes=[("Excel文件", "*.xlsx")],
                            initialname=filename
                        )
                        if file_path:
                            success = self.export_customers_to_excel(file_path, customer_type)
                    elif file_format == "CSV":
                        file_path = filedialog.asksaveasfilename(
                            defaultextension=".csv",
                            filetypes=[("CSV文件", "*.csv")],
                            initialname=filename
                        )
                        if file_path:
                            success = self.export_customers_to_csv(file_path, customer_type)
                    elif file_format == "PDF":
                        file_path = filedialog.asksaveasfilename(
                            defaultextension=".pdf",
                            filetypes=[("PDF文件", "*.pdf")],
                            initialname=filename
                        )
                        if file_path:
                            success = self.export_customers_to_pdf(file_path, customer_type)
                    
                    if success:
                        messagebox.showinfo("导出成功", f"客户数据已成功导出为 {file_format} 格式", parent=dialog)
                        dialog.destroy()
                    else:
                        messagebox.showerror("导出失败", "导出过程中发生错误", parent=dialog)
                        
                except Exception as e:
                    messagebox.showerror("错误", f"导出失败：{e}", parent=dialog)
            
            tk.Button(btn_frame, text="📊 开始导出", command=do_export,
                     bg=self.colors['primary'], fg='white', bd=0, pady=8, padx=20,
                     font=('Microsoft YaHei UI', 10)).pack(side="left")
            tk.Button(btn_frame, text="取消", command=dialog.destroy,
                     bg=self.colors['text_light'], fg='white', bd=0, pady=8, padx=20,
                     font=('Microsoft YaHei UI', 10)).pack(side="right")
                     
        except Exception as e:
            messagebox.showerror("错误", f"打开导出对话框失败：{e}")
    
    def export_customers_to_excel(self, file_path: str, customer_type: str) -> bool:
        """导出客户为Excel格式"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "客户数据"
            
            # 设置标题
            title = f"智慧餐饮管理系统 - 客户数据 ({customer_type})"
            ws['A1'] = title
            ws['A1'].font = Font(size=16, bold=True)
            ws.merge_cells('A1:F1')
            
            # 设置表头样式
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # 表头
            headers = ["客户姓名", "联系电话", "客户类型", "地址", "备注", "注册时间"]
            ws.append(headers)
            
            # 设置表头样式
            for cell in ws[2]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # 获取客户数据
            customers = self.get_filtered_customers(customer_type)
            
            # 添加数据
            for customer in customers:
                row = [
                    customer.get('name', ''),
                    customer.get('phone', ''),
                    customer.get('type', ''),
                    customer.get('address', ''),
                    customer.get('note', ''),
                    customer.get('created_at', '')
                ]
                ws.append(row)
            
            # 调整列宽
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(file_path)
            return True
            
        except ImportError:
            messagebox.showerror("错误", "请安装openpyxl库：pip install openpyxl")
            return False
        except Exception as e:
            print(f"导出Excel失败: {e}")
            return False
    
    def export_customers_to_csv(self, file_path: str, customer_type: str) -> bool:
        """导出客户为CSV格式"""
        try:
            import csv
            
            with open(file_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
                fieldnames = ["客户姓名", "联系电话", "客户类型", "地址", "备注", "注册时间"]
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                writer.writeheader()
                
                # 获取客户数据
                customers = self.get_filtered_customers(customer_type)
                
                for customer in customers:
                    writer.writerow({
                        "客户姓名": customer.get('name', ''),
                        "联系电话": customer.get('phone', ''),
                        "客户类型": customer.get('type', ''),
                        "地址": customer.get('address', ''),
                        "备注": customer.get('note', ''),
                        "注册时间": customer.get('created_at', '')
                    })
            
            return True
            
        except Exception as e:
            print(f"导出CSV失败: {e}")
            return False
    
    def export_customers_to_pdf(self, file_path: str, customer_type: str) -> bool:
        """导出客户为PDF格式"""
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            
            doc = SimpleDocTemplate(file_path, pagesize=A4)
            story = []
            
            # 标题样式
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                spaceAfter=30,
                alignment=1  # 居中
            )
            
            # 添加标题
            title = Paragraph(f"智慧餐饮管理系统 - 客户数据 ({customer_type})", title_style)
            story.append(title)
            story.append(Spacer(1, 20))
            
            # 获取客户数据
            customers = self.get_filtered_customers(customer_type)
            
            # 创建表格数据
            table_data = [["客户姓名", "联系电话", "客户类型", "地址", "备注", "注册时间"]]
            
            for customer in customers:
                row = [
                    customer.get('name', ''),
                    customer.get('phone', ''),
                    customer.get('type', ''),
                    customer.get('address', ''),
                    customer.get('note', ''),
                    customer.get('created_at', '')
                ]
                table_data.append(row)
            
            # 创建表格
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.beige, colors.white])
            ]))
            story.append(table)
            
            doc.build(story)
            return True
            
        except ImportError:
            messagebox.showerror("错误", "请安装reportlab库：pip install reportlab")
            return False
        except Exception as e:
            print(f"导出PDF失败: {e}")
            return False
    
    def get_filtered_customers(self, customer_type: str) -> List[Dict]:
        """获取筛选后的客户数据"""
        if customer_type == "全部":
            return self.customer_data
        else:
            return [customer for customer in self.customer_data if customer.get('type') == customer_type]
