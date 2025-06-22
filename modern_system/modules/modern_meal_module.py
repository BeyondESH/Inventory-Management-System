#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Modern Meal Configuration Module
Adopts modern design style meal management interface
"""

import tkinter as tk
from tkinter import ttk, messagebox, simpledialog
from typing import Dict, List, Any
import datetime
import json
import os

# Import data manager
try:
    from .data_manager import data_manager
except ImportError:
    try:
        from data_manager import data_manager
    except ImportError:
        # Create mock data manager
        class MockDataManager:
            def load_data(self, data_type):
                return []
            def save_data(self, data_type, data):
                return True
        data_manager = MockDataManager()

class ModernMealModule:
    def __init__(self, parent_frame, title_frame):
        self.parent_frame = parent_frame
        self.title_frame = title_frame
        
        # Modern color theme
        self.colors = {
            'primary': '#FF6B35',      # Primary
            'secondary': '#F7931E',    # Secondary
            'accent': '#FFD23F',       # Accent
            'background': '#F8F9FA',   # Background
            'surface': '#FFFFFF',      # Card background
            'text': '#2D3436',         # Default text color
            'text_primary': '#2D3436', # Main text
            'text_secondary': '#636E72', # Secondary text
            'text_light': '#7F8C8D',   # Light text for dialogs/buttons
            'border': '#E0E0E0',       # Border
            'success': '#00B894',      # Success
            'warning': '#FDCB6E',      # Warning
            'error': '#E84393',        # Error
            'card_shadow': '#F0F0F0',   # Card shadow
            'info': '#366092',         # Info
            'white': '#FFFFFF',        # White
            'danger': '#E74C3C'        # Danger
        }
        
        # Font configuration
        self.fonts = {
            'title': ('Segoe UI', 20, 'bold'),
            'heading': ('Segoe UI', 16, 'bold'),
            'subheading': ('Segoe UI', 14, 'bold'),
            'body': ('Segoe UI', 12),
            'small': ('Segoe UI', 10),
            'button': ('Segoe UI', 11, 'bold'),
            'price': ('Segoe UI', 14, 'bold')
        }
        
        # Meal data
        self.meal_data = self.load_meal_data()        # Interface variables (lazy initialization)
        
        # Stats label references
        self.stats_labels = {}
    
    def load_meal_data(self):
        """Load meal data from data management center"""
        try:
            # Get meal data from data manager
            meals_data = data_manager.load_data('meals')
            
            # Convert data format to adapt to existing interface
            formatted_data = []
            for meal in meals_data:
                # 处理可用状态，兼容不同字段名
                # Determine availability: prefer 'is_available', fallback to 'isActive'
                available = meal.get('is_available', None)
                if available is None:
                    available = meal.get('isActive', True)
                
                # Process fields
                formatted_meal = {
                    "id": meal.get('id', ''),
                    "name": meal.get('name', ''),
                    "category": meal.get('category', 'Other'),
                    "price": meal.get('price', 0.0),
                    "cost": meal.get('cost', 0.0),
                    "description": meal.get('description', 'No description'),
                    "ingredients": meal.get('ingredients', []),
                    "cooking_time": meal.get('cooking_time', 15),
                    "calories": meal.get('calories', 200),
                    "is_spicy": meal.get('is_spicy', False),
                    "is_vegetarian": meal.get('is_vegetarian', False),
                    "is_available": available,
                    "image": meal.get('image', '🍽️'),
                    "created_date": meal.get('created_date', datetime.datetime.now().strftime('%Y-%m-%d'))
                }
                formatted_data.append(formatted_meal)
            
            return formatted_data
        except Exception as e:
            print(f"Failed to load meal data: {e}")
            # Return default meal data
            return [
                {
                    "id": "MEAL001", "name": "Tomato Beef Noodles", "category": "Noodles", "price": 25.0,
                    "cost": 15.0, "description": "Classic tomato beef noodles, fresh and delicious",
                    "ingredients": ["Tomato", "Beef", "Noodles"], "cooking_time": 15,
                    "calories": 450, "is_spicy": False, "is_vegetarian": False,
                    "is_available": True, "image": "🍽️", "created_date": "2025-06-21"
                },
                {
                    "id": "MEAL002", "name": "Egg Fried Rice", "category": "Fried Rice", "price": 18.0,
                    "cost": 10.0, "description": "Delicious egg fried rice",
                    "ingredients": ["Egg", "Rice"], "cooking_time": 10,
                    "calories": 350, "is_spicy": False, "is_vegetarian": False,
                    "is_available": True, "image": "🍚", "created_date": "2025-06-21"
                },
                {
                    "id": "MEAL003", "name": "Beef Burger", "category": "Western", "price": 32.0,
                    "cost": 20.0, "description": "Delicious beef burger meal",
                    "ingredients": ["Beef", "Bread", "Lettuce"], "cooking_time": 12,
                    "calories": 520, "is_spicy": False, "is_vegetarian": False,
                    "is_available": True, "image": "🍔", "created_date": "2025-06-21"
                },
                {
                    "id": "MEAL004", "name": "French Fries", "category": "Snacks", "price": 12.0,
                    "cost": 6.0, "description": "Crispy golden french fries",
                    "ingredients": ["Potato"], "cooking_time": 8,
                    "calories": 280, "is_spicy": False, "is_vegetarian": True,
                    "is_available": True, "image": "🍽️", "created_date": "2025-06-21"                }            ]
    
    def save_meal_data(self):
        """Save meal data to data management center"""
        try:
            # Convert internal format data to standard format
            standard_data = []
            for meal in self.meal_data:
                standard_meal = {
                    'id': meal.get('id', ''),
                    'name': meal.get('name', ''),
                    'category': meal.get('category', 'Other'),
                    'price': meal.get('price', 0.0),
                    'cost': meal.get('cost', 0.0),
                    'description': meal.get('description', ''),
                    'ingredients': meal.get('ingredients', []),
                    'cooking_time': meal.get('cooking_time', 15),
                    'calories': meal.get('calories', 200),
                    'is_spicy': meal.get('is_spicy', False),
                    'is_vegetarian': meal.get('is_vegetarian', False),
                    'is_available': meal.get('is_available', True),
                    'image': meal.get('image', '🍽️'),
                    'created_date': meal.get('created_date', datetime.datetime.now().strftime('%Y-%m-%d'))
                }
                standard_data.append(standard_meal)
            
            # Save to data manager
            data_manager.save_data('meals', standard_data)
            return True
        except Exception as e:
            print(f"Failed to save meal data: {e}")
            return False
    
    def notify_data_update(self):
        """Notify other modules that data has been updated"""
        try:
            # Notify sales management module to refresh meal data
            if hasattr(data_manager, 'notify_modules'):
                data_manager.notify_modules('meals_updated')
            else:
                # Directly notify registered modules
                if hasattr(data_manager, 'registered_modules'):
                    for module_type, module_instance in data_manager.registered_modules.items():
                        if module_type == 'sales' and hasattr(module_instance, 'refresh_meals_data'):
                            module_instance.refresh_meals_data()
        except Exception as e:
            print(f"Failed to notify other modules: {e}")
    
    def show(self):
        """Show meal configuration module"""
        # Register to data manager
        data_manager.register_module('meal', self)
        
        # Reload latest data
        self.meal_data = self.load_meal_data()
        
        self.clear_frames()
        self.update_title()
        self.create_meal_interface()
        
    def clear_frames(self):
        """Clear frames"""
        for widget in self.parent_frame.winfo_children():
            widget.destroy()
        for widget in self.title_frame.winfo_children():
            widget.destroy()
            
    def update_title(self):
        """Update title"""
        # Left title
        title_frame = tk.Frame(self.title_frame, bg=self.colors['surface'])
        title_frame.pack(side="left", fill="y")
        
        icon_label = tk.Label(title_frame, text="🍜", font=('Segoe UI Emoji', 20),
                             bg=self.colors['surface'], fg=self.colors['primary'])
        icon_label.pack(side="left", padx=(30, 10), pady=20)
        
        title_label = tk.Label(title_frame, text="Meal Configuration", font=self.fonts['title'],
                              bg=self.colors['surface'], fg=self.colors['text_primary'])
        title_label.pack(side="left", pady=20)
        
        # Right operation buttons
        action_frame = tk.Frame(self.title_frame, bg=self.colors['surface'])
        action_frame.pack(side="right", padx=30, pady=20)
        
        # Add meal button
        add_btn = tk.Button(action_frame, text="➕ Add Meal", 
                           font=('Segoe UI', 10),
                           bg=self.colors['primary'], fg=self.colors['white'],
                           bd=0, padx=20, pady=8, cursor='hand2',
                           command=self.add_meal)
        add_btn.pack(side='right', padx=5)
        
        # Refresh button
        refresh_btn = tk.Button(action_frame, text="🔄 Refresh", 
                               font=('Segoe UI', 10),
                               bg=self.colors['info'], fg=self.colors['white'],
                               bd=0, padx=20, pady=8, cursor='hand2',
                               command=self.refresh_meals)
        refresh_btn.pack(side='right', padx=5)
        
        # Export button
        export_btn = tk.Button(action_frame, text="📊 Export", 
                              font=('Segoe UI', 10),
                              bg=self.colors['success'], fg=self.colors['white'],
                              bd=0, padx=20, pady=8, cursor='hand2',
                              command=self.export_meals)
        export_btn.pack(side='right', padx=5)
        
    def create_meal_interface(self):
        """Create meal management interface"""
        # Main container
        main_container = tk.Frame(self.parent_frame, bg=self.colors['background'])
        main_container.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Top stats cards
        self.create_stats_cards(main_container)
        
        # Bottom meal grid
        self.create_meals_grid(main_container)
        
    def create_stats_cards(self, parent):
        """Create stats cards"""
        stats_frame = tk.Frame(parent, bg=self.colors['background'])
        stats_frame.pack(fill="x", pady=(0, 20))
        
        # Calculate stats data
        total_meals = len(self.meal_data)
        available_meals = len([meal for meal in self.meal_data if meal['is_available']])
        avg_price = sum(meal['price'] for meal in self.meal_data) / total_meals if total_meals > 0 else 0
        spicy_meals = len([meal for meal in self.meal_data if meal['is_spicy']])
        
        cards_data = [
            {"title": "Total Meals", "value": f"{total_meals}", "icon": "🍽️", "color": self.colors['primary']},
            {"title": "Available Meals", "value": f"{available_meals}", "icon": "✅", "color": self.colors['success']},
            {"title": "Average Price", "value": f"¥{avg_price:.1f}", "icon": "💰", "color": self.colors['accent']},
            {"title": "Spicy Meals", "value": f"{spicy_meals}", "icon": "🌶️", "color": self.colors['error']}
        ]
        
        for i, card_data in enumerate(cards_data):
            self.create_stats_card(stats_frame, card_data, i)
            
    def create_stats_card(self, parent, data, index):
        """Create single stats card"""
        card_frame = tk.Frame(parent, bg=self.colors['surface'], relief="flat", bd=1)
        card_frame.grid(row=0, column=index, padx=10, pady=10, sticky="ew")
        
        # Configure grid weights
        parent.grid_columnconfigure(index, weight=1)
        
        # Card content
        content_frame = tk.Frame(card_frame, bg=self.colors['surface'])
        content_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Icon and title row
        header_frame = tk.Frame(content_frame, bg=self.colors['surface'])
        header_frame.pack(fill="x", pady=(0, 10))
        
        icon_label = tk.Label(header_frame, text=data["icon"], font=('Segoe UI Emoji', 24),
                             bg=self.colors['surface'], fg=data["color"])
        icon_label.pack(side="left")
        
        title_label = tk.Label(header_frame, text=data["title"], font=self.fonts['body'],
                              bg=self.colors['surface'], fg=self.colors['text_secondary'])
        title_label.pack(side="right")
        
        # Value
        value_label = tk.Label(content_frame, text=data["value"], font=self.fonts['title'],
                              bg=self.colors['surface'], fg=self.colors['text_primary'])
        value_label.pack(anchor="w")
        
        # Save reference for update
        self.stats_labels[data["title"]] = value_label
        
    def create_meals_grid(self, parent):
        """Create meal grid"""
        grid_frame = tk.Frame(parent, bg=self.colors['background'])
        grid_frame.pack(fill="both", expand=True)
        
        # Title
        title_frame = tk.Frame(grid_frame, bg=self.colors['background'])
        title_frame.pack(fill="x", pady=(0, 20))
        
        title_label = tk.Label(title_frame, text="🍽️ Meal List", font=self.fonts['heading'],
                              bg=self.colors['background'], fg=self.colors['text_primary'])
        title_label.pack(side="left")
        
        # Scroll area
        canvas = tk.Canvas(grid_frame, bg=self.colors['background'])
        scrollbar = ttk.Scrollbar(grid_frame, orient="vertical", command=canvas.yview)
        
        self.meals_container = tk.Frame(canvas, bg=self.colors['background'])
        
        self.meals_container.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=self.meals_container, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
          # Bind mouse wheel
        def _on_mousewheel(event):
            try:
                if canvas.winfo_exists():
                    canvas.yview_scroll(int(-1*(event.delta/120)), "units")
            except tk.TclError:
                pass  # Widget has been destroyed, ignore error
        
        canvas.bind("<MouseWheel>", _on_mousewheel)
        self.meals_container.bind("<MouseWheel>", _on_mousewheel)
        
        # Show meals
        self.refresh_meals_display()
        
    def refresh_meals_display(self):
        """Refresh meal display"""
        # Clear existing display
        for widget in self.meals_container.winfo_children():
            widget.destroy()
        
        # Get all meal data
        meals_to_show = self.get_filtered_meals()
        
        if not meals_to_show:
            no_data_label = tk.Label(self.meals_container, 
                                   text="No meal data",
                                   font=self.fonts['body'],
                                   bg=self.colors['background'],
                                   fg=self.colors['text_secondary'])
            no_data_label.pack(pady=50)
            return
        
        # Create grid layout
        row = 0
        col = 0
        max_cols = 4
        
        for meal in meals_to_show:
            self.create_meal_card(self.meals_container, meal, row, col)
            
            col += 1
            if col >= max_cols:
                col = 0
                row += 1
        
        # Configure grid weights
        for i in range(max_cols):
            self.meals_container.grid_columnconfigure(i, weight=1)
        
        # Update stats cards
        self.update_stats_cards()
        
    def get_filtered_meals(self):
        """Get filtered meal list"""
        # Return all meals directly, no filtering
        return self.meal_data
    
    def filter_meals(self):
        """Filter meals (simplified, directly refresh display)"""
        self.refresh_meals_display()
        
    def create_meal_card(self, parent, meal, row, col):
        """Create meal card"""
        # Card frame
        card_frame = tk.Frame(parent, bg=self.colors['surface'], relief="flat", bd=1,
                             cursor="hand2")
        card_frame.grid(row=row, column=col, padx=15, pady=15, sticky="ew")
        
        # Configure grid weights
        parent.grid_columnconfigure(col, weight=1)
        
        # Card content
        content_frame = tk.Frame(card_frame, bg=self.colors['surface'])
        content_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Top: Icon, name, and status
        header_frame = tk.Frame(content_frame, bg=self.colors['surface'])
        header_frame.pack(fill="x", pady=(0, 15))
        
        # Meal icon
        icon_label = tk.Label(header_frame, text=meal["image"], font=('Segoe UI Emoji', 36),
                             bg=self.colors['surface'])
        icon_label.pack(side="left")
        
        # Name and status
        name_frame = tk.Frame(header_frame, bg=self.colors['surface'])
        name_frame.pack(side="right", fill="x", expand=True, padx=(15, 0))
        
        name_label = tk.Label(name_frame, text=meal["name"], font=self.fonts['subheading'],
                             bg=self.colors['surface'], fg=self.colors['text_primary'], anchor="w")
        name_label.pack(fill="x")
        
        # Tag row
        tags_frame = tk.Frame(name_frame, bg=self.colors['surface'])
        tags_frame.pack(fill="x", pady=(5, 0))
        
        # Category tag
        category_tag = tk.Label(tags_frame, text=meal["category"], font=self.fonts['small'],
                               bg=self.colors['accent'], fg="white", padx=8, pady=2)
        category_tag.pack(side="left", padx=(0, 5))
        
        # Status tag
        if meal["is_available"]:
            status_tag = tk.Label(tags_frame, text="Available", font=self.fonts['small'],
                                 bg=self.colors['success'], fg="white", padx=8, pady=2)
        else:
            status_tag = tk.Label(tags_frame, text="Out of Stock", font=self.fonts['small'],
                                 bg=self.colors['error'], fg="white", padx=8, pady=2)
        status_tag.pack(side="left", padx=(0, 5))
        
        # Special tag
        if meal["is_spicy"]:
            spicy_tag = tk.Label(tags_frame, text="Spicy", font=self.fonts['small'],
                                bg=self.colors['error'], fg="white", padx=8, pady=2)
            spicy_tag.pack(side="left", padx=(0, 5))
            
        if meal["is_vegetarian"]:
            veg_tag = tk.Label(tags_frame, text="Vegetarian", font=self.fonts['small'],
                              bg=self.colors['success'], fg="white", padx=8, pady=2)
            veg_tag.pack(side="left", padx=(0, 5))
        
        # Description
        desc_label = tk.Label(content_frame, text=meal["description"], font=self.fonts['small'],
                             bg=self.colors['surface'], fg=self.colors['text_secondary'],
                             wraplength=250, justify="left")
        desc_label.pack(fill="x", pady=(0, 15))
        
        # Price and cost
        price_frame = tk.Frame(content_frame, bg=self.colors['surface'])
        price_frame.pack(fill="x", pady=(0, 15))
        
        price_label = tk.Label(price_frame, text=f"¥{meal['price']:.0f}", font=self.fonts['price'],
                              bg=self.colors['surface'], fg=self.colors['primary'])
        price_label.pack(side="left")
        
        cost_label = tk.Label(price_frame, text=f"Cost: ¥{meal['cost']:.0f}", font=self.fonts['small'],
                             bg=self.colors['surface'], fg=self.colors['text_secondary'])
        cost_label.pack(side="right")
        
        # Other information
        info_frame = tk.Frame(content_frame, bg=self.colors['surface'])
        info_frame.pack(fill="x", pady=(0, 15))
        
        time_label = tk.Label(info_frame, text=f"⏱️ {meal['cooking_time']} minutes", font=self.fonts['small'],
                             bg=self.colors['surface'], fg=self.colors['text_secondary'])
        time_label.pack(side="left")
        
        calories_label = tk.Label(info_frame, text=f"🔥 {meal['calories']} calories", font=self.fonts['small'],
                                 bg=self.colors['surface'], fg=self.colors['text_secondary'])
        calories_label.pack(side="right")
        
        # Action buttons
        button_frame = tk.Frame(content_frame, bg=self.colors['surface'])
        button_frame.pack(fill="x")
        
        # Edit button
        edit_btn = tk.Button(button_frame, text="Edit", font=self.fonts['body'],
                            bg=self.colors['background'], fg=self.colors['text_primary'],
                            bd=1, relief="solid", cursor="hand2",
                            command=lambda m=meal: self.edit_meal(m), padx=15, pady=5)
        edit_btn.pack(side="left", padx=(0, 10))
        
        # Toggle status button
        if meal["is_available"]:
            toggle_btn = tk.Button(button_frame, text="Out of Stock", font=self.fonts['body'],
                                  bg=self.colors['warning'], fg="white",
                                  bd=0, relief="flat", cursor="hand2",
                                  command=lambda m=meal: self.toggle_meal_status(m), padx=15, pady=5)
        else:
            toggle_btn = tk.Button(button_frame, text="Available", font=self.fonts['body'],
                                  bg=self.colors['success'], fg="white",
                                  bd=0, relief="flat", cursor="hand2",
                                  command=lambda m=meal: self.toggle_meal_status(m), padx=15, pady=5)
        toggle_btn.pack(side="right")
        
        # Card hover effect
        def on_card_enter(event):
            card_frame.configure(relief="solid", bd=2)
            
        def on_card_leave(event):
            card_frame.configure(relief="flat", bd=1)
              # Bind hover events
        for widget in [card_frame, content_frame, header_frame, icon_label]:
            widget.bind("<Enter>", on_card_enter)
            widget.bind("<Leave>", on_card_leave)
    
    def update_stats_cards(self):
        """Update stats cards"""
        total_meals = len(self.meal_data)
        available_meals = len([meal for meal in self.meal_data if meal['is_available']])
        avg_price = sum(meal['price'] for meal in self.meal_data) / total_meals if total_meals > 0 else 0
        spicy_meals = len([meal for meal in self.meal_data if meal['is_spicy']])
        
        # Update labels
        if "Total Meals" in self.stats_labels:
            self.stats_labels["Total Meals"].config(text=f"{total_meals}")
        if "Available Meals" in self.stats_labels:
            self.stats_labels["Available Meals"].config(text=f"{available_meals}")
        if "Average Price" in self.stats_labels:
            self.stats_labels["Average Price"].config(text=f"¥{avg_price:.1f}")
        if "Spicy Meals" in self.stats_labels:
            self.stats_labels["Spicy Meals"].config(text=f"{spicy_meals}")
            
    def add_meal(self):
        """Add meal"""
        dialog = MealDialog(self.parent_frame, "Add Meal")
        if dialog.result:
            # Generate new ID
            new_id = f"MEAL{len(self.meal_data) + 1:03d}"
            dialog.result['id'] = new_id
            dialog.result['created_date'] = datetime.datetime.now().strftime("%Y-%m-%d")
            
            # Add to data
            self.meal_data.append(dialog.result)
            
            # Save to data manager
            self.save_meal_data()
            
            self.refresh_meals_display()
            messagebox.showinfo("Success", "Meal added successfully!")
            
            # Notify other modules data update
            self.notify_data_update()
            
    def edit_meal(self, meal):
        """Edit meal"""
        dialog = MealDialog(self.parent_frame, "Edit Meal", meal)
        if dialog.result:
            # Update data
            meal.update(dialog.result)
            
            # Save to data manager
            self.save_meal_data()
            
            self.refresh_meals_display()
            messagebox.showinfo("Success", "Meal information updated successfully!")
            
            # Notify other modules data update
            self.notify_data_update()
            
    def toggle_meal_status(self, meal):
        """Toggle meal status"""
        action = "Out of Stock" if not meal["is_available"] else "Available"
        if messagebox.askyesno("Confirm Operation", f"Are you sure you want to {action} meal '{meal['name']}'?"):
            meal["is_available"] = not meal["is_available"]
            
            # Save to data manager
            self.save_meal_data()
            
            self.refresh_meals_display()
            messagebox.showinfo("Success", f"Meal is now {action}!")
            
            # Notify other modules data update
            self.notify_data_update()
            
    def export_meals(self):
        """Export meal data"""
        try:
            from tkinter import filedialog
            import datetime
            
            # Create export selection dialog
            dialog = tk.Toplevel(self.parent_frame)
            dialog.title("Export Meal Data")
            dialog.geometry("400x300")
            dialog.configure(bg=self.colors['background'])
            dialog.transient(self.parent_frame)
            dialog.grab_set()
            
            # Center display
            dialog.update_idletasks()
            x = (dialog.winfo_screenwidth() // 2) - (200)
            y = (dialog.winfo_screenheight() // 2) - (150)
            dialog.geometry(f"400x300+{x}+{y}")
            
            # Title
            tk.Label(dialog, text="Export Meal Data", font=('Segoe UI', 14, 'bold'),
                    bg=self.colors['background'], fg=self.colors['text']).pack(pady=15)
            
            # Export options frame
            options_frame = tk.Frame(dialog, bg=self.colors['background'])
            options_frame.pack(fill="both", expand=True, padx=20, pady=10)
            
            # Export format selection
            tk.Label(options_frame, text="Select export format:", font=('Segoe UI', 12),
                    bg=self.colors['background'], fg=self.colors['text']).pack(anchor="w", pady=(0, 10))
            
            format_var = tk.StringVar(dialog, value="Excel")
            format_options = ["Excel", "CSV", "PDF"]
            
            format_frame = tk.Frame(options_frame, bg=self.colors['background'])
            format_frame.pack(anchor="w")
            
            for i, fmt in enumerate(format_options):
                rb = tk.Radiobutton(format_frame, text=fmt, variable=format_var, value=fmt,
                                  font=('Segoe UI', 10), bg=self.colors['background'], 
                                  fg=self.colors['text'], selectcolor=self.colors['surface'])
                rb.grid(row=0, column=i, sticky="w", padx=(0, 20))
            
            # Meal status filter
            tk.Label(options_frame, text="Meal Status:", font=('Segoe UI', 12),
                    bg=self.colors['background'], fg=self.colors['text']).pack(anchor="w", pady=(20, 10))
            
            status_var = tk.StringVar(dialog, value="All")
            status_options = ["All", "Available", "Out of Stock"]
            
            status_combo = ttk.Combobox(options_frame, textvariable=status_var, 
                                      values=status_options, state="readonly", width=20)
            status_combo.pack(anchor="w")
            
            # Button frame
            btn_frame = tk.Frame(dialog, bg=self.colors['background'])
            btn_frame.pack(fill="x", padx=20, pady=20)
            
            def do_export():
                try:
                    file_format = format_var.get()
                    meal_status = status_var.get()
                    
                    # Get current timestamp
                    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                    filename = f"Meal Data_{meal_status}_{timestamp}"
                    
                    # Select save path
                    if file_format == "Excel":
                        file_path = filedialog.asksaveasfilename(
                            defaultextension=".xlsx",
                            filetypes=[("Excel File", "*.xlsx")],
                            initialname=filename
                        )
                        if file_path:
                            success = self.export_meals_to_excel(file_path, meal_status)
                    elif file_format == "CSV":
                        file_path = filedialog.asksaveasfilename(
                            defaultextension=".csv",
                            filetypes=[("CSV File", "*.csv")],
                            initialname=filename
                        )
                        if file_path:
                            success = self.export_meals_to_csv(file_path, meal_status)
                    elif file_format == "PDF":
                        file_path = filedialog.asksaveasfilename(
                            defaultextension=".pdf",
                            filetypes=[("PDF File", "*.pdf")],
                            initialname=filename
                        )
                        if file_path:
                            success = self.export_meals_to_pdf(file_path, meal_status)
                    
                    if success:
                        messagebox.showinfo("Export Success", f"Meal data exported successfully to {file_format} format", parent=dialog)
                        dialog.destroy()
                    else:
                        messagebox.showerror("Export Failed", "Error occurred during export", parent=dialog)
                        
                except Exception as e:
                    messagebox.showerror("Error", f"Export Failed: {e}", parent=dialog)
            
            tk.Button(btn_frame, text="📊 Start Export", command=do_export,
                     bg=self.colors['primary'], fg='white', bd=0, pady=8, padx=20,
                     font=('Segoe UI', 10)).pack(side="left")
            tk.Button(btn_frame, text="Cancel", command=dialog.destroy,
                     bg=self.colors['text_light'], fg='white', bd=0, pady=8, padx=20,
                     font=('Segoe UI', 10)).pack(side="right")
                     
        except Exception as e:
            messagebox.showerror("Error", f"Failed to open export dialog: {e}")
    
    def export_meals_to_excel(self, file_path: str, meal_status: str) -> bool:
        """Export meal to Excel format"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Meal Data"
            
            # Set title
            title = f"Smart Restaurant Management System - Meal Data ({meal_status})"
            ws['A1'] = title
            ws['A1'].font = Font(size=16, bold=True)
            ws.merge_cells('A1:F1')
            
            # Set header style
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Header
            headers = ["Meal Name", "Price", "Category", "Status", "Description", "Created Time"]
            ws.append(headers)
            
            # Set header style
            for cell in ws[2]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Get meal data
            meals = self.get_filtered_meals_for_export(meal_status)
            
            # Add data
            for meal in meals:
                row = [
                    meal.get('name', ''),
                    f"￥{meal.get('price', 0):.2f}",
                    meal.get('category', ''),
                    meal.get('status', ''),
                    meal.get('description', ''),
                    meal.get('created_at', '')
                ]
                ws.append(row)
            
            # Adjust column width
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(file_path)
            return True
            
        except ImportError:
            messagebox.showerror("Error", "Please install openpyxl library: pip install openpyxl")
            return False
        except Exception as e:
            print(f"Failed to export Excel: {e}")
            return False
    
    def export_meals_to_csv(self, file_path: str, meal_status: str) -> bool:
        """Export meal to CSV format"""
        try:
            import csv
            
            with open(file_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
                fieldnames = ["Meal Name", "Price", "Category", "Status", "Description", "Created Time"]
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                writer.writeheader()
                
                # Get meal data
                meals = self.get_filtered_meals_for_export(meal_status)
                
                for meal in meals:
                    writer.writerow({
                        "Meal Name": meal.get('name', ''),
                        "Price": f"￥{meal.get('price', 0):.2f}",
                        "Category": meal.get('category', ''),
                        "Status": meal.get('status', ''),
                        "Description": meal.get('description', ''),
                        "Created Time": meal.get('created_at', '')
                    })
            
            return True
            
        except Exception as e:
            print(f"Failed to export CSV: {e}")
            return False
    
    def export_meals_to_pdf(self, file_path: str, meal_status: str) -> bool:
        """Export meal to PDF format"""
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            
            doc = SimpleDocTemplate(file_path, pagesize=A4)
            story = []
            
            # Title style
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                spaceAfter=30,
                alignment=1  # Center
            )
            
            # Add title
            title = Paragraph(f"Smart Restaurant Management System - Meal Data ({meal_status})", title_style)
            story.append(title)
            story.append(Spacer(1, 20))
            
            # Get meal data
            meals = self.get_filtered_meals_for_export(meal_status)
            
            # Create table data
            table_data = [["Meal Name", "Price", "Category", "Status", "Description", "Created Time"]]
            
            for meal in meals:
                row = [
                    meal.get('name', ''),
                    f"￥{meal.get('price', 0):.2f}",
                    meal.get('category', ''),
                    meal.get('status', ''),
                    meal.get('description', ''),
                    meal.get('created_at', '')
                ]
                table_data.append(row)
            
            # Create table
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.beige, colors.white])
            ]))
            story.append(table)
            
            doc.build(story)
            return True
            
        except ImportError:
            messagebox.showerror("Error", "Please install reportlab library: pip install reportlab")
            return False
        except Exception as e:
            print(f"Failed to export PDF: {e}")
            return False
    
    def get_filtered_meals_for_export(self, meal_status: str) -> List[Dict]:
        """Get filtered meal data for export"""
        if meal_status == "All":
            return self.meal_data
        else:
            return [meal for meal in self.meal_data if meal.get('status') == meal_status]
    
    def refresh_meals(self):
        """Refresh meal data"""
        try:
            # Reload and reformat meal data
            self.meal_data = self.load_meal_data()
            # Refresh meal list display
            self.refresh_meals_display()
            messagebox.showinfo("Refresh Success", "Meal data refreshed")
        except Exception as e:
            messagebox.showerror("Refresh Failed", f"Error occurred while refreshing meal data: {e}")

class MealDialog:
    """Meal dialog"""
    def __init__(self, parent, title, meal_data=None):
        self.result = None
          # Create dialog window
        self.dialog = tk.Toplevel(parent)
        self.dialog.title(title)
        self.dialog.geometry("600x800")  # Increase height from 700 to 800
        self.dialog.configure(bg="#f8f9fa")
        self.dialog.resizable(False, False)
        self.dialog.grab_set()
        
        # Center display
        self.center_window()
        
        # Color theme
        self.colors = {
            'primary': '#FF6B35',
            'background': '#F8F9FA',
            'surface': '#FFFFFF',
            'text_primary': '#2D3436',
            'text_secondary': '#636E72',
            'border': '#E0E0E0'
        }
        
        # Font
        self.fonts = {
            'heading': ('Segoe UI', 16, 'bold'),
            'body': ('Segoe UI', 12),
            'button': ('Segoe UI', 11, 'bold')
        }
          # Create variables
        self.name_var = tk.StringVar(self.dialog, value=meal_data['name'] if meal_data else "")
        self.category_var = tk.StringVar(self.dialog, value=meal_data['category'] if meal_data else "")
        self.price_var = tk.DoubleVar(self.dialog, value=meal_data['price'] if meal_data else 0.0)
        self.cost_var = tk.DoubleVar(self.dialog, value=meal_data['cost'] if meal_data else 0.0)
        self.description_var = tk.StringVar(self.dialog, value=meal_data['description'] if meal_data else "")
        self.cooking_time_var = tk.IntVar(self.dialog, value=meal_data['cooking_time'] if meal_data else 0)
        self.calories_var = tk.IntVar(self.dialog, value=meal_data['calories'] if meal_data else 0)
        self.is_spicy_var = tk.BooleanVar(self.dialog, value=meal_data['is_spicy'] if meal_data else False)
        self.is_vegetarian_var = tk.BooleanVar(self.dialog, value=meal_data['is_vegetarian'] if meal_data else False)
        self.is_available_var = tk.BooleanVar(self.dialog, value=meal_data['is_available'] if meal_data else True)
        self.image_var = tk.StringVar(self.dialog, value=meal_data['image'] if meal_data else "🍽️")
        
        # Ingredient list
        self.ingredients = meal_data['ingredients'].copy() if meal_data else []
        
        # Create interface
        self.create_dialog_ui()
        
    def center_window(self):
        """Window center"""
        self.dialog.update_idletasks()
        width = self.dialog.winfo_width()
        height = self.dialog.winfo_height()
        x = (self.dialog.winfo_screenwidth() // 2) - (width // 2)
        y = (self.dialog.winfo_screenheight() // 2) - (height // 2)
        self.dialog.geometry(f'{width}x{height}+{x}+{y}')
        
    def create_dialog_ui(self):
        """Create dialog interface"""
        # Main container
        main_frame = tk.Frame(self.dialog, bg=self.colors['surface'])
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Scroll area
        canvas = tk.Canvas(main_frame, bg=self.colors['surface'])
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg=self.colors['surface'])
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Title
        title_label = tk.Label(scrollable_frame, text="🍜 Meal Information", font=self.fonts['heading'],
                              bg=self.colors['surface'], fg=self.colors['text_primary'])
        title_label.pack(pady=(0, 20))
        
        # Basic information
        self.create_basic_info_section(scrollable_frame)
        
        # Detailed information
        self.create_detail_info_section(scrollable_frame)
        
        # Ingredient management
        self.create_ingredients_section(scrollable_frame)
        
        # Options section
        self.create_options_section(scrollable_frame)
        
        # Button area
        button_frame = tk.Frame(scrollable_frame, bg=self.colors['surface'])
        button_frame.pack(fill="x", pady=(20, 0))
        
        # Cancel button
        cancel_btn = tk.Button(button_frame, text="Cancel", font=self.fonts['button'],
                              bg=self.colors['background'], fg=self.colors['text_secondary'],
                              bd=0, relief="flat", cursor="hand2", command=self.cancel,
                              padx=30, pady=10)
        cancel_btn.pack(side="right", padx=(10, 0))
        
        # OK button
        ok_btn = tk.Button(button_frame, text="OK", font=self.fonts['button'],
                          bg=self.colors['primary'], fg="white",
                          bd=0, relief="flat", cursor="hand2", command=self.ok,
                          padx=30, pady=10)
        ok_btn.pack(side="right")
        
    def create_basic_info_section(self, parent):
        """Create basic information section"""
        section_frame = tk.Frame(parent, bg=self.colors['surface'])
        section_frame.pack(fill="x", pady=(0, 20))
        
        section_title = tk.Label(section_frame, text="📝 Basic Information", font=self.fonts['body'],
                                bg=self.colors['surface'], fg=self.colors['text_primary'])
        section_title.pack(anchor="w", pady=(0, 10))
        
        # Meal name
        self.create_form_field(section_frame, "Meal Name *", self.name_var, "entry")
        
        # Category and icon
        row_frame = tk.Frame(section_frame, bg=self.colors['surface'])
        row_frame.pack(fill="x", pady=10)
        
        left_frame = tk.Frame(row_frame, bg=self.colors['surface'])
        left_frame.pack(side="left", fill="x", expand=True, padx=(0, 10))
        self.create_form_field(left_frame, "Category *", self.category_var, "combo", 
                              ["Main Course", "Side Dish", "Appetizer", "Dessert", "Drink", "Snack"])
        
        right_frame = tk.Frame(row_frame, bg=self.colors['surface'])
        right_frame.pack(side="right", fill="x", expand=True, padx=(10, 0))
        self.create_form_field(right_frame, "Icon", self.image_var, "entry")
        
        # Price and cost
        row_frame2 = tk.Frame(section_frame, bg=self.colors['surface'])
        row_frame2.pack(fill="x", pady=10)
        
        left_frame2 = tk.Frame(row_frame2, bg=self.colors['surface'])
        left_frame2.pack(side="left", fill="x", expand=True, padx=(0, 10))
        self.create_form_field(left_frame2, "Price (¥) *", self.price_var, "entry")
        
        right_frame2 = tk.Frame(row_frame2, bg=self.colors['surface'])
        right_frame2.pack(side="right", fill="x", expand=True, padx=(10, 0))
        self.create_form_field(right_frame2, "Cost (¥) *", self.cost_var, "entry")
        
    def create_detail_info_section(self, parent):
        """Create detailed information section"""
        section_frame = tk.Frame(parent, bg=self.colors['surface'])
        section_frame.pack(fill="x", pady=(0, 20))
        
        section_title = tk.Label(section_frame, text="📋 Detailed Information", font=self.fonts['body'],
                                bg=self.colors['surface'], fg=self.colors['text_primary'])
        section_title.pack(anchor="w", pady=(0, 10))
        
        # Description
        self.create_form_field(section_frame, "Meal Description", self.description_var, "text")
        
        # Cooking time and calories
        row_frame = tk.Frame(section_frame, bg=self.colors['surface'])
        row_frame.pack(fill="x", pady=10)
        
        left_frame = tk.Frame(row_frame, bg=self.colors['surface'])
        left_frame.pack(side="left", fill="x", expand=True, padx=(0, 10))
        self.create_form_field(left_frame, "Cooking Time (minutes)", self.cooking_time_var, "entry")
        
        right_frame = tk.Frame(row_frame, bg=self.colors['surface'])
        right_frame.pack(side="right", fill="x", expand=True, padx=(10, 0))
        self.create_form_field(right_frame, "Calories", self.calories_var, "entry")
        
    def create_ingredients_section(self, parent):
        """Create ingredient section"""
        section_frame = tk.Frame(parent, bg=self.colors['surface'])
        section_frame.pack(fill="x", pady=(0, 20))
        
        section_title = tk.Label(section_frame, text="🥗 Ingredient List", font=self.fonts['body'],
                                bg=self.colors['surface'], fg=self.colors['text_primary'])
        section_title.pack(anchor="w", pady=(0, 10))
        
        # Ingredient list
        ingredients_frame = tk.Frame(section_frame, bg=self.colors['background'])
        ingredients_frame.pack(fill="x", pady=(0, 10))
        
        self.ingredients_listbox = tk.Listbox(ingredients_frame, height=4, font=self.fonts['body'])
        self.ingredients_listbox.pack(fill="x")
        
        # Add ingredient
        add_ingredient_frame = tk.Frame(section_frame, bg=self.colors['surface'])
        add_ingredient_frame.pack(fill="x")
        
        self.ingredient_var = tk.StringVar(self.dialog)
        ingredient_entry = tk.Entry(add_ingredient_frame, textvariable=self.ingredient_var,
                                   font=self.fonts['body'], width=30)
        ingredient_entry.pack(side="left", padx=(0, 10), ipady=5)
        
        add_btn = tk.Button(add_ingredient_frame, text="Add", font=self.fonts['body'],
                           bg=self.colors['primary'], fg="white", bd=0,
                           cursor="hand2", command=self.add_ingredient, padx=15)
        add_btn.pack(side="left", padx=(0, 10))
        
        remove_btn = tk.Button(add_ingredient_frame, text="Remove", font=self.fonts['body'],
                              bg=self.colors['background'], fg=self.colors['text_secondary'], bd=1,
                              cursor="hand2", command=self.remove_ingredient, padx=15)
        remove_btn.pack(side="left")
        
        # Refresh ingredient list
        self.refresh_ingredients_list()
        
    def create_options_section(self, parent):
        """Create options section"""
        section_frame = tk.Frame(parent, bg=self.colors['surface'])
        section_frame.pack(fill="x", pady=(0, 20))
        
        section_title = tk.Label(section_frame, text="⚙️ Meal Settings", font=self.fonts['body'],
                                bg=self.colors['surface'], fg=self.colors['text_primary'])
        section_title.pack(anchor="w", pady=(0, 10))
        
        # Options checkboxes
        options_frame = tk.Frame(section_frame, bg=self.colors['surface'])
        options_frame.pack(fill="x")
        
        
        spicy_check = tk.Checkbutton(options_frame, text="🌶️ 辣味菜品", variable=self.is_spicy_var,
                                    bg=self.colors['surface'], font=self.fonts['body'],
                                    activebackground=self.colors['surface'])
        spicy_check.pack(anchor="w", pady=2)
        
        veg_check = tk.Checkbutton(options_frame, text="🥬 素食菜品", variable=self.is_vegetarian_var,
                                  bg=self.colors['surface'], font=self.fonts['body'],
                                  activebackground=self.colors['surface'])
        veg_check.pack(anchor="w", pady=2)
        
        available_check = tk.Checkbutton(options_frame, text="✅ 当前在售", variable=self.is_available_var,
                                        bg=self.colors['surface'], font=self.fonts['body'],
                                        activebackground=self.colors['surface'])
        available_check.pack(anchor="w", pady=2)
        
    def create_form_field(self, parent, label_text, variable, field_type, options=None):
        """创建表单字段"""
        field_frame = tk.Frame(parent, bg=self.colors['surface'])
        field_frame.pack(fill="x", pady=5)
        
        # 标签
        label = tk.Label(field_frame, text=label_text, font=self.fonts['body'],
                        bg=self.colors['surface'], fg=self.colors['text_secondary'], anchor="w")
        label.pack(fill="x", pady=(0, 5))
        
        # 输入控件
        if field_type == "entry":
            entry = tk.Entry(field_frame, textvariable=variable, font=self.fonts['body'],
                            bg=self.colors['background'], bd=1, relief="solid")
            entry.pack(fill="x", ipady=8)
        elif field_type == "combo" and options:
            combo = ttk.Combobox(field_frame, textvariable=variable, values=options,
                                font=self.fonts['body'], state="readonly")
            combo.pack(fill="x", ipady=5)
        elif field_type == "text":
            text_entry = tk.Entry(field_frame, textvariable=variable, font=self.fonts['body'],
                                 bg=self.colors['background'], bd=1, relief="solid")
            text_entry.pack(fill="x", ipady=8)
            
    def add_ingredient(self):
        """添加食材"""
        ingredient = self.ingredient_var.get().strip()
        if ingredient and ingredient not in self.ingredients:
            self.ingredients.append(ingredient)
            self.ingredient_var.set("")
            self.refresh_ingredients_list()
            
    def remove_ingredient(self):
        """删除食材"""
        selection = self.ingredients_listbox.curselection()
        if selection:
            index = selection[0]
            del self.ingredients[index]
            self.refresh_ingredients_list()
            
    def refresh_ingredients_list(self):
        """刷新食材列表"""
        self.ingredients_listbox.delete(0, tk.END)
        for ingredient in self.ingredients:
            self.ingredients_listbox.insert(tk.END, ingredient)
            
    def ok(self):
        """确定按钮处理"""
        # 验证必填字段
        if not self.name_var.get().strip():
            messagebox.showerror("错误", "请输入菜品名称")
            return
        if not self.category_var.get().strip():
            messagebox.showerror("错误", "请选择菜品分类")
            return
            
        # 验证数值
        try:
            price = self.price_var.get()
            cost = self.cost_var.get()
            cooking_time = self.cooking_time_var.get()
            calories = self.calories_var.get()
            
            if price <= 0 or cost <= 0:
                messagebox.showerror("错误", "价格和成本必须大于0")
                return
                
            if cooking_time < 0 or calories < 0:
                messagebox.showerror("错误", "制作时间和热量不能为负数")
                return
                
        except tk.TclError:
            messagebox.showerror("错误", "请输入有效的数值")
            return
        
        # 保存结果
        self.result = {
            'name': self.name_var.get().strip(),
            'category': self.category_var.get(),
            'price': price,
            'cost': cost,
            'description': self.description_var.get().strip(),
            'ingredients': self.ingredients.copy(),
            'cooking_time': cooking_time,
            'calories': calories,
            'is_spicy': self.is_spicy_var.get(),
            'is_vegetarian': self.is_vegetarian_var.get(),
            'is_available': self.is_available_var.get(),
            'image': self.image_var.get().strip() or "🍽️"
        }
        
        self.dialog.destroy()
        
    def cancel(self):
        """取消按钮处理"""
        self.dialog.destroy()

if __name__ == "__main__":
    # 测试代码
    root = tk.Tk()
    root.title("现代化菜品管理模块测试")
    root.geometry("1400x900")
    root.configure(bg="#f8f9fa")
    
    title_frame = tk.Frame(root, bg="#ffffff", height=70)
    title_frame.pack(fill="x")
    title_frame.pack_propagate(False)
    
    main_frame = tk.Frame(root, bg="#f8f9fa")
    main_frame.pack(fill="both", expand=True)
    
    meal_module = ModernMealModule(main_frame, title_frame)
    meal_module.show()
    
    root.mainloop()
