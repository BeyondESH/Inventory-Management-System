#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Modern Inventory Management Module
An inventory management interface with a modern design style
"""

import tkinter as tk
from tkinter import ttk, messagebox, simpledialog
from typing import Dict, List, Any
import datetime
import json
import os

# Import data manager
try:
    from .data_manager import data_manager
except ImportError:
    try:
        from data_manager import data_manager
    except ImportError:
        # Create a mock data manager
        class MockDataManager:
            def get_inventory(self):
                return []
            def save_data(self, data_type, data):
                return True
            def update_inventory(self, item_id, quantity):
                return True
        data_manager = MockDataManager()

class ModernInventoryModule:
    def __init__(self, parent_frame, title_frame):
        self.parent_frame = parent_frame
        self.title_frame = title_frame
        
        # Modern color theme
        self.colors = {
            'primary': '#FF6B35',      # Main color
            'secondary': '#F7931E',    # Secondary color
            'accent': '#FFD23F',       # Accent color
            'background': '#F8F9FA',   # Background color
            'surface': '#FFFFFF',      # Card background
            'text_primary': '#2D3436', # Primary text
            'text_secondary': '#636E72', # Secondary text
            'border': '#E0E0E0',       # Border
            'success': '#00B894',      # Success color
            'warning': '#FDCB6E',      # Warning color
            'error': '#E84393',        # Error color
            'card_shadow': '#F0F0F0',  # Card shadow
            'white': '#FFFFFF',        # White
            'info': '#3498DB'          # Info color
        }
        
        # Font configuration
        self.fonts = {
            'title': ('Segoe UI', 20, 'bold'),
            'heading': ('Segoe UI', 16, 'bold'),
            'subheading': ('Segoe UI', 14, 'bold'),
            'body': ('Segoe UI', 12),
            'small': ('Segoe UI', 10),
            'button': ('Segoe UI', 11, 'bold')
        }
        
        # Inventory data
        self.inventory_data = self.load_inventory_data()        # UI variables (lazy initialization)
        self.search_var = None
        self.category_filter_var = None
        self.stock_filter_var = None
          # UI component references
        self.inventory_tree = None
        self.stats_labels = {}
        self.recipes = []
        self.possible_meals_frame = None
    
    def load_inventory_data(self):
        """Load inventory data from the data management center"""
        try:
            # Get inventory data from the data manager
            inventory_data = data_manager.get_inventory()
            
            # Convert data format to match the existing UI
            formatted_data = []
            for item in inventory_data:
                formatted_item = {
                    "id": item.get('id', ''),
                    "name": item.get('name', ''),
                    "category": item.get('category', ''),
                    # JSON uses 'current_stock'; fallback to 'quantity' if present
                    "current_stock": item.get('current_stock', item.get('quantity', 0)),
                    "min_stock": item.get('min_stock', 0),
                    "max_stock": item.get('max_stock', 100),
                    "unit": item.get('unit', 'unit'),
                    "price": item.get('price', 0.0),
                    "supplier": item.get('supplier', 'Unknown Supplier'),
                    "last_updated": item.get('update_time', datetime.datetime.now().strftime('%Y-%m-%d'))
                }
                formatted_data.append(formatted_item)
            
            return formatted_data
        except Exception as e:
            print(f"Failed to load inventory data: {e}")
            # Return rich default sample data
            return [
                # Vegetables
                {"id": "INV001", "name": "Tomato", "category": "Vegetable", "current_stock": 50, "min_stock": 10, "max_stock": 100, "unit": "kg", "price": 8.0, "supplier": "Quality Veggie Supplier", "last_updated": "2025-06-21"},
                {"id": "INV002", "name": "Onion", "category": "Vegetable", "current_stock": 30, "min_stock": 8, "max_stock": 80, "unit": "kg", "price": 6.0, "supplier": "Quality Veggie Supplier", "last_updated": "2025-06-21"},
                {"id": "INV003", "name": "Green Pepper", "category": "Vegetable", "current_stock": 25, "min_stock": 5, "max_stock": 60, "unit": "kg", "price": 12.0, "supplier": "Quality Veggie Supplier", "last_updated": "2025-06-21"},
                {"id": "INV004", "name": "Lettuce", "category": "Vegetable", "current_stock": 40, "min_stock": 10, "max_stock": 80, "unit": "kg", "price": 10.0, "supplier": "Quality Veggie Supplier", "last_updated": "2025-06-21"},
                {"id": "INV005", "name": "Carrot", "category": "Vegetable", "current_stock": 35, "min_stock": 8, "max_stock": 70, "unit": "kg", "price": 7.0, "supplier": "Quality Veggie Supplier", "last_updated": "2025-06-21"},
                
                # Meats
                {"id": "INV010", "name": "Beef", "category": "Meat", "current_stock": 20, "min_stock": 5, "max_stock": 50, "unit": "kg", "price": 68.0, "supplier": "Quality Meat Supplier", "last_updated": "2025-06-21"},
                {"id": "INV011", "name": "Pork", "category": "Meat", "current_stock": 25, "min_stock": 5, "max_stock": 60, "unit": "kg", "price": 28.0, "supplier": "Quality Meat Supplier", "last_updated": "2025-06-21"},
                {"id": "INV012", "name": "Chicken Breast", "category": "Meat", "current_stock": 15, "min_stock": 3, "max_stock": 40, "unit": "kg", "price": 22.0, "supplier": "Quality Meat Supplier", "last_updated": "2025-06-21"},
                {"id": "INV013", "name": "Eggs", "category": "Meat", "current_stock": 200, "min_stock": 50, "max_stock": 300, "unit": "pcs", "price": 1.2, "supplier": "Quality Meat Supplier", "last_updated": "2025-06-21"},
                
                # Staples
                {"id": "INV020", "name": "Noodles", "category": "Staple", "current_stock": 100, "min_stock": 20, "max_stock": 200, "unit": "pack", "price": 3.5, "supplier": "Quality Grain Supplier", "last_updated": "2025-06-21"},
                {"id": "INV021", "name": "Rice", "category": "Staple", "current_stock": 80, "min_stock": 15, "max_stock": 150, "unit": "kg", "price": 4.5, "supplier": "Quality Grain Supplier", "last_updated": "2025-06-21"},
                {"id": "INV022", "name": "Bread", "category": "Staple", "current_stock": 60, "min_stock": 20, "max_stock": 120, "unit": "pcs", "price": 8.0, "supplier": "Quality Grain Supplier", "last_updated": "2025-06-21"},
                {"id": "INV023", "name": "Potato", "category": "Staple", "current_stock": 45, "min_stock": 10, "max_stock": 90, "unit": "kg", "price": 5.0, "supplier": "Quality Veggie Supplier", "last_updated": "2025-06-21"},
                
                # Beverages
                {"id": "INV030", "name": "Coke", "category": "Beverage", "current_stock": 80, "min_stock": 30, "max_stock": 150, "unit": "bottle", "price": 5.0, "supplier": "Beverage Supplier", "last_updated": "2025-06-21"},
                {"id": "INV031", "name": "Sprite", "category": "Beverage", "current_stock": 75, "min_stock": 25, "max_stock": 120, "unit": "bottle", "price": 5.0, "supplier": "Beverage Supplier", "last_updated": "2025-06-21"},
                {"id": "INV032", "name": "Orange Juice", "category": "Beverage", "current_stock": 50, "min_stock": 20, "max_stock": 100, "unit": "bottle", "price": 8.0, "supplier": "Beverage Supplier", "last_updated": "2025-06-21"},
                {"id": "INV033", "name": "Coffee Beans", "category": "Beverage", "current_stock": 5, "min_stock": 2, "max_stock": 20, "unit": "kg", "price": 180.0, "supplier": "Coffee Supplier", "last_updated": "2025-06-21"},
                {"id": "INV034", "name": "Milk", "category": "Beverage", "current_stock": 40, "min_stock": 15, "max_stock": 80, "unit": "bottle", "price": 6.0, "supplier": "Dairy Supplier", "last_updated": "2025-06-21"},
                
                # Seasonings
                {"id": "INV040", "name": "Cooking Oil", "category": "Seasoning", "current_stock": 10, "min_stock": 3, "max_stock": 25, "unit": "bottle", "price": 25.0, "supplier": "Seasoning Supplier", "last_updated": "2025-06-21"},
                {"id": "INV041", "name": "Soy Sauce", "category": "Seasoning", "current_stock": 8, "min_stock": 2, "max_stock": 20, "unit": "bottle", "price": 12.0, "supplier": "Seasoning Supplier", "last_updated": "2025-06-21"},
                {"id": "INV042", "name": "Dark Soy Sauce", "category": "Seasoning", "current_stock": 6, "min_stock": 2, "max_stock": 15, "unit": "bottle", "price": 15.0, "supplier": "Seasoning Supplier", "last_updated": "2025-06-21"},
                {"id": "INV043", "name": "Salt", "category": "Seasoning", "current_stock": 20, "min_stock": 5, "max_stock": 50, "unit": "pack", "price": 3.0, "supplier": "Seasoning Supplier", "last_updated": "2025-06-21"},
                {"id": "INV044", "name": "Sugar", "category": "Seasoning", "current_stock": 15, "min_stock": 3, "max_stock": 30, "unit": "pack", "price": 8.0, "supplier": "Seasoning Supplier", "last_updated": "2025-06-21"},
                {"id": "INV045", "name": "Chili Powder", "category": "Seasoning", "current_stock": 12, "min_stock": 3, "max_stock": 25, "unit": "pack", "price": 18.0, "supplier": "Seasoning Supplier", "last_updated": "2025-06-21"}
            ]
    
    def show(self):
        """Show the inventory management module"""
        # Register with the data manager
        data_manager.register_module('inventory', self)
        
        # Reload the latest data
        self.inventory_data = self.load_inventory_data()
        # Initialize UI variables (if not already initialized)
        if self.search_var is None:
            self.search_var = tk.StringVar(self.parent_frame)
            self.category_filter_var = tk.StringVar(self.parent_frame, value="All")
            self.stock_filter_var = tk.StringVar(self.parent_frame, value="All")
        
        self.clear_frames()
        self.update_title()
        self.create_inventory_interface()
        
    def clear_frames(self):
        """Clear the frames"""
        for widget in self.parent_frame.winfo_children():
            widget.destroy()
        for widget in self.title_frame.winfo_children():
            widget.destroy()
            
    def update_title(self):
        """Update the title"""
        # Left title
        title_frame = tk.Frame(self.title_frame, bg=self.colors['surface'])
        title_frame.pack(side="left", fill="y")
        
        icon_label = tk.Label(title_frame, text="📦", font=('Segoe UI Emoji', 20),
                             bg=self.colors['surface'], fg=self.colors['primary'])
        icon_label.pack(side="left", padx=(30, 10), pady=20)
        
        title_label = tk.Label(title_frame, text="Inventory Management", font=self.fonts['title'],
                              bg=self.colors['surface'], fg=self.colors['text_primary'])
        title_label.pack(side="left", pady=20)
        
        # Right action buttons
        action_frame = tk.Frame(self.title_frame, bg=self.colors['surface'])
        action_frame.pack(side="right", padx=30, pady=20)
        
        # Refresh button
        refresh_btn = tk.Button(action_frame, text="Refresh", 
                               font=self.fonts['button'],
                               bg=self.colors['primary'], fg=self.colors['white'],
                               bd=0, padx=20, pady=8, cursor='hand2',
                               command=self.refresh_inventory)
        refresh_btn.pack(side='right', padx=5)
        
        # Export button
        export_btn = tk.Button(action_frame, text="📊 Export", 
                              font=self.fonts['button'],
                              bg=self.colors['success'], fg=self.colors['white'],
                              bd=0, padx=20, pady=8, cursor='hand2',
                              command=self.export_inventory)
        export_btn.pack(side='right', padx=5)
        
    def create_inventory_interface(self):
        """Create the inventory management interface"""
        main_container = tk.Frame(self.parent_frame, bg=self.colors['background'])
        main_container.pack(fill="both", expand=True, padx=20, pady=20)
        
        top_container = tk.Frame(main_container, bg=self.colors['background'])
        top_container.pack(fill="x", expand=False)

        self.create_possible_meals_section(top_container)
        
        bottom_container = tk.Frame(main_container, bg=self.colors['background'])
        bottom_container.pack(fill="both", expand=True, pady=(10, 0))

        self.create_inventory_list(bottom_container)
        
    def create_stats_cards(self, parent):
        """Create stats cards"""
        stats_frame = tk.Frame(parent, bg=self.colors['background'])
        stats_frame.pack(fill="x", pady=(0, 20))
        
        # 计算统计数据
        total_items = len(self.inventory_data)
        low_stock_items = len([item for item in self.inventory_data if item['current_stock'] <= item['min_stock']])
        total_value = sum(item['current_stock'] * item['price'] for item in self.inventory_data)
        out_of_stock = len([item for item in self.inventory_data if item['current_stock'] == 0])
        
        cards_data = [
            {"title": "Total Items", "value": f"{total_items}", "icon": "📦", "color": self.colors['primary']},
            {"title": "Low Stock", "value": f"{low_stock_items}", "icon": "⚠️", "color": self.colors['warning']},
            {"title": "Total Value", "value": f"¥{total_value:,.0f}", "icon": "💰", "color": self.colors['success']},
            {"title": "Out of Stock", "value": f"{out_of_stock}", "icon": "🚫", "color": self.colors['error']}
        ]
        
        for i, card_data in enumerate(cards_data):
            self.create_stats_card(stats_frame, card_data, i)
            
    def create_stats_card(self, parent, data, index):
        """Create a single stats card"""
        card_frame = tk.Frame(parent, bg=self.colors['surface'], relief="flat", bd=1)
        card_frame.grid(row=0, column=index, padx=10, pady=10, sticky="ew")
        
        # 配置网格权重
        parent.grid_columnconfigure(index, weight=1)
        
        # 卡片内容
        content_frame = tk.Frame(card_frame, bg=self.colors['surface'])
        content_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # 图标和标题行
        header_frame = tk.Frame(content_frame, bg=self.colors['surface'])
        header_frame.pack(fill="x", pady=(0, 10))
        
        icon_label = tk.Label(header_frame, text=data["icon"], font=('Segoe UI Emoji', 24),
                             bg=self.colors['surface'], fg=data["color"])
        icon_label.pack(side="left")
        
        title_label = tk.Label(header_frame, text=data["title"], font=self.fonts['body'],
                              bg=self.colors['surface'], fg=self.colors['text_secondary'])
        title_label.pack(side="left", padx=10)
        
        # 数值
        value_label = tk.Label(content_frame, text=data["value"], font=self.fonts['heading'],
                              bg=self.colors['surface'], fg=self.colors['text_primary'])
        value_label.pack(anchor="w")
        
        # 保存引用用于更新
        self.stats_labels[data["title"]] = value_label
        
    def create_filter_section(self, parent):
        """Create filter section"""
        filter_frame = tk.Frame(parent, bg=self.colors['surface'], height=80)
        filter_frame.pack(fill="x", pady=(0, 20))
        filter_frame.pack_propagate(False)
        
        content_frame = tk.Frame(filter_frame, bg=self.colors['surface'])
        content_frame.pack(fill="both", expand=True, padx=30, pady=20)
        
        # 搜索框
        search_frame = tk.Frame(content_frame, bg=self.colors['surface'])
        search_frame.pack(side="left", fill="y")
        
        search_label = tk.Label(search_frame, text="🔎", font=self.fonts['body'], bg=self.colors['surface'])
        search_label.pack(side="left")
        
        search_entry = ttk.Entry(search_frame, textvariable=self.search_var, font=self.fonts['body'], width=30)
        search_entry.pack(side="left", padx=5)
        search_entry.bind("<KeyRelease>", self.search_inventory)
        
        # 分类筛选
        category_label = tk.Label(filter_frame, text="Category:", font=self.fonts['body'], bg=self.colors['surface'])
        category_label.pack(side="left", padx=(20, 5))
        
        categories = ["All"] + sorted(list(set(item['category'] for item in self.inventory_data)))
        category_menu = ttk.OptionMenu(filter_frame, self.category_filter_var, "All", *categories, command=self.filter_inventory)
        category_menu.pack(side="left")
        
        # 库存状态筛选
        stock_label = tk.Label(filter_frame, text="Stock Status:", font=self.fonts['body'], bg=self.colors['surface'])
        stock_label.pack(side="left", padx=(20, 5))
        
        stock_options = ["All", "Low Stock", "Out of Stock"]
        stock_menu = ttk.OptionMenu(filter_frame, self.stock_filter_var, "All", *stock_options, command=self.filter_inventory)
        stock_menu.pack(side="left")
        
    def create_inventory_list(self, parent):
        """Create the inventory list"""
        list_container = tk.Frame(parent, bg=self.colors['surface'], relief="flat", bd=1)
        list_container.pack(fill="both", expand=True, pady=(20,0))
        
        # List title
        title_frame = tk.Frame(list_container, bg=self.colors['surface'])
        title_frame.pack(fill="x", padx=20, pady=(10, 5))
        
        title_label = tk.Label(title_frame, text="📦 Ingredient Inventory List", font=self.fonts['subheading'],
                              bg=self.colors['surface'], fg=self.colors['text_primary'])
        title_label.pack(side="left")

        subtitle_label = tk.Label(title_frame, text="(Displays raw ingredients, not finished dishes)", font=self.fonts['small'],
                                 bg=self.colors['surface'], fg=self.colors['text_secondary'])
        subtitle_label.pack(side="left", padx=10)
        
        # Treeview list
        tree_frame = tk.Frame(list_container, bg=self.colors['surface'])
        tree_frame.pack(fill="both", expand=True, padx=20, pady=10)
        
        # Style configuration
        style = ttk.Style()
        style.configure("Modern.Treeview",
                        background=self.colors['surface'],
                        foreground=self.colors['text_primary'],
                        fieldbackground=self.colors['surface'],
                        rowheight=35,
                        font=self.fonts['body'])
        style.configure("Modern.Treeview.Heading",
                        background=self.colors['background'],
                        foreground=self.colors['text_primary'],
                        font=self.fonts['button'],
                        relief="flat")
        style.map("Modern.Treeview.Heading",
                  background=[('active', self.colors['border'])])
        style.layout("Modern.Treeview", [('Treeview.treearea', {'sticky': 'nswe'})]) # Remove borders
        
        # Treeview definition
        columns = ("id", "name", "category", "current_stock", "min_stock", "unit", "price", "status", "supplier", "last_updated")
        self.inventory_tree = ttk.Treeview(tree_frame, columns=columns, show="headings", style="Modern.Treeview")
        
        # Define headings
        headings = {
            "id": "ID", "name": "Product Name", "category": "Category", "current_stock": "Current Stock",
            "min_stock": "Min Stock", "unit": "Unit", "price": "Unit Price", "status": "Status",
            "supplier": "Supplier", "last_updated": "Last Updated"
        }
        for col, text in headings.items():
            self.inventory_tree.heading(col, text=text, anchor='w')
            
        # Define column widths
        widths = {
            "id": 60, "name": 150, "category": 100, "current_stock": 100, "min_stock": 80,
            "unit": 60, "price": 80, "status": 100, "supplier": 150, "last_updated": 120
        }
        for col, width in widths.items():
            self.inventory_tree.column(col, width=width, anchor='w')
        
        # Vertical scrollbar
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.inventory_tree.yview)
        vsb.pack(side='right', fill='y')
        self.inventory_tree.configure(yscrollcommand=vsb.set)
        
        # Horizontal scrollbar
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.inventory_tree.xview)
        hsb.pack(side='bottom', fill='x')
        self.inventory_tree.configure(xscrollcommand=hsb.set)
        
        self.inventory_tree.pack(fill="both", expand=True)
        self.inventory_tree.bind("<Double-1>", self.edit_inventory_item)
        
        # Action buttons below list
        action_bar = tk.Frame(list_container, bg=self.colors['surface'])
        action_bar.pack(fill="x", padx=20, pady=10)
        
        add_btn = tk.Button(action_bar, text="➕ Add Item", font=self.fonts['button'],
                           bg=self.colors['success'], fg='white', bd=0, padx=15, pady=8,
                           command=self.add_inventory_item)
        add_btn.pack(side="left", padx=5)
        
        edit_btn = tk.Button(action_bar, text="✏️ Edit Selected", font=self.fonts['button'],
                            bg=self.colors['info'], fg='white', bd=0, padx=15, pady=8,
                            command=self.edit_selected_item)
        edit_btn.pack(side="left", padx=5)
        
        delete_btn = tk.Button(action_bar, text="❌ Delete Selected", font=self.fonts['button'],
                              bg=self.colors['error'], fg='white', bd=0, padx=15, pady=8,
                              command=self.delete_selected_item)
        delete_btn.pack(side="left", padx=5)
        
        restock_btn = tk.Button(action_bar, text="🚚 Restock", font=self.fonts['button'],
                               bg=self.colors['secondary'], fg='white', bd=0, padx=15, pady=8,
                               command=self.restock_item)
        restock_btn.pack(side="left", padx=5)

        adjust_stock_btn = tk.Button(action_bar, text="🛠️ Adjust Stock", font=self.fonts['button'],
                                     bg=self.colors['warning'], fg=self.colors['text_primary'],
                                     bd=0, padx=15, pady=8, command=self.adjust_stock)
        adjust_stock_btn.pack(side="left", padx=5)

        self.create_context_menu()
        self.refresh_inventory_list()
        self.refresh_possible_meals()
        
    def create_context_menu(self):
        """Create the context menu for the inventory list"""
        menu = tk.Menu(self.inventory_tree, tearoff=0, font=self.fonts['body'],
                       bg=self.colors['surface'], fg=self.colors['text_primary'])
        menu.add_command(label="✏️ Edit", command=self.edit_selected_item)
        menu.add_command(label="❌ Delete", command=self.delete_selected_item)
        menu.add_separator()
        menu.add_command(label="🚚 Restock", command=self.restock_item)
        menu.add_command(label="🛠️ Adjust Stock", command=self.adjust_stock)
        
        def show_context_menu(event):
            # Select row under cursor
            item = self.inventory_tree.identify_row(event.y)
            if item:
                self.inventory_tree.focus(item)
                self.inventory_tree.selection_set(item)
                menu.post(event.x_root, event.y_root)

        self.inventory_tree.bind("<Button-3>", show_context_menu)
    
    def refresh_inventory_list(self):
        """Refresh the inventory list display"""
        # Clear existing items
        for i in self.inventory_tree.get_children():
            self.inventory_tree.delete(i)
            
        filtered_data = self.get_filtered_data()
        
        for item in filtered_data:
            current_stock = item.get('current_stock', 0)
            min_stock = item.get('min_stock', 0)
            
            # Determine stock status
            if current_stock <= 0:
                status = "Out of Stock"
                tags = ('out_of_stock',)
            elif current_stock <= min_stock:
                status = "Low Stock"
                tags = ('low_stock',)
            else:
                status = "In Stock"
                tags = ('in_stock',)
            
            # Insert data into Treeview
            self.inventory_tree.insert(
                "", "end", 
                values=(
                    item.get('id', ''), item.get('name', ''), item.get('category', ''),
                    f"{current_stock:.2f}", f"{min_stock:.2f}",
                    item.get('unit', ''), f"¥{item.get('price', 0.0):.2f}",
                    status, item.get('supplier', ''), item.get('last_updated', '')
                ),
                tags=tags
            )
            
        # Configure row colors
        self.inventory_tree.tag_configure('in_stock', background=self.colors['surface'], foreground=self.colors['text_primary'])
        self.inventory_tree.tag_configure('low_stock', background=self.colors['warning'], foreground=self.colors['text_primary'])
        self.inventory_tree.tag_configure('out_of_stock', background=self.colors['error'], foreground=self.colors['white'])
        
    def get_filtered_data(self):
        """Get filtered data based on search and filter criteria"""
        search_term = self.search_var.get().lower()
        category_filter = self.category_filter_var.get()
        stock_filter = self.stock_filter_var.get()

        filtered_data = self.inventory_data
        
        # Search filter
        if search_term:
            filtered_data = [
                item for item in filtered_data 
                if search_term in item['name'].lower() or \
                   search_term in item['id'].lower() or \
                   search_term in item['supplier'].lower()
            ]
        
        # Category filter
        if category_filter != "All":
            filtered_data = [item for item in filtered_data if item['category'] == category_filter]

        # Stock status filter
        if stock_filter == "Low Stock":
            filtered_data = [item for item in filtered_data if 0 < item['current_stock'] <= item['min_stock']]
        elif stock_filter == "Out of Stock":
            filtered_data = [item for item in filtered_data if item['current_stock'] == 0]
            
        return filtered_data
    
    def filter_ingredients_only(self):
        """
        This method is now part of get_filtered_data and is kept for conceptual reference.
        It's designed to filter out finished products and only show raw materials.
        This logic should be integrated into the data source or initial loading if needed permanently.
        """
        pass # Logic moved to get_filtered_data

    def update_stats_cards(self):
        """Update the values on the statistics cards"""
        total_items = len(self.inventory_data)
        low_stock_items = len([item for item in self.inventory_data if item['current_stock'] <= item['min_stock']])
        total_value = sum(item['current_stock'] * item['price'] for item in self.inventory_data)
        out_of_stock = len([item for item in self.inventory_data if item['current_stock'] == 0])

        if "Total Items" in self.stats_labels:
            self.stats_labels["Total Items"].config(text=f"{total_items}")
        if "Low Stock" in self.stats_labels:
            self.stats_labels["Low Stock"].config(text=f"{low_stock_items}")
        if "Total Value" in self.stats_labels:
            self.stats_labels["Total Value"].config(text=f"¥{total_value:,.0f}")
        if "Out of Stock" in self.stats_labels:
            self.stats_labels["Out of Stock"].config(text=f"{out_of_stock}")

    def search_inventory(self, *args):
        """Search the inventory"""
        self.refresh_inventory_list()

    def filter_inventory(self, *args):
        """Filter the inventory display"""
        self.refresh_inventory_list()
        
    def add_inventory_item(self):
        """Add a new inventory item"""
        dialog = InventoryItemDialog(self.parent_frame, "Add New Item")
        if dialog.result:
            new_item = dialog.result
            
            # Generate new ID
            max_id = 0
            for item in self.inventory_data:
                if item['id'].startswith('INV'):
                    try:
                        num = int(item['id'][3:])
                        if num > max_id:
                            max_id = num
                    except ValueError:
                        continue
            new_item['id'] = f"INV{max_id + 1:03d}"
            
            # Update data source
            self.inventory_data.append(new_item)
            if data_manager.save_data('inventory', self.inventory_data):
                messagebox.showinfo("Success", "Item added successfully.")
                self.refresh_inventory()
            else:
                messagebox.showerror("Error", "Failed to save new item.")
                self.inventory_data.pop() # Revert change
            
    def edit_inventory_item(self, event):
        self.edit_selected_item()

    def edit_selected_item(self):
        """Edit the selected inventory item"""
        selected_item = self.inventory_tree.selection()
        if not selected_item:
            messagebox.showwarning("No Selection", "Please select an item to edit.")
            return
        
        item_id = self.inventory_tree.item(selected_item[0])['values'][0]
        item_data = next((item for item in self.inventory_data if item['id'] == item_id), None)
        
        if item_data:
            dialog = InventoryItemDialog(self.parent_frame, "Edit Item", item_data)
            if dialog.result:
                updated_item = dialog.result
                # Update data source
                for i, item in enumerate(self.inventory_data):
                    if item['id'] == item_id:
                        self.inventory_data[i] = updated_item
                        break
                
                if data_manager.save_data('inventory', self.inventory_data):
                    messagebox.showinfo("Success", "Item updated successfully.")
                    self.refresh_inventory()
                else:
                    messagebox.showerror("Error", "Failed to save updated item.")
                    # Revert change (optional)
    
    def delete_selected_item(self):
        """Delete the selected inventory item"""
        selected_item = self.inventory_tree.selection()
        if not selected_item:
            messagebox.showwarning("No Selection", "Please select an item to delete.")
            return
            
        item_id = self.inventory_tree.item(selected_item[0])['values'][0]
        item_name = self.inventory_tree.item(selected_item[0])['values'][1]

        if messagebox.askyesno("Confirm Deletion", f"Are you sure you want to delete '{item_name}'?"):
            original_data = list(self.inventory_data)
            self.inventory_data = [item for item in self.inventory_data if item['id'] != item_id]
            
            if data_manager.save_data('inventory', self.inventory_data):
                messagebox.showinfo("Success", "Item deleted successfully.")
                self.refresh_inventory()
            else:
                messagebox.showerror("Error", "Failed to delete item.")
                self.inventory_data = original_data # Revert

    def restock_item(self):
        """Restock the selected inventory item"""
        selected_item = self.inventory_tree.selection()
        if not selected_item:
            messagebox.showwarning("No Selection", "Please select an item to restock.")
            return
            
        item_name = self.inventory_tree.item(selected_item[0])['values'][1]
        
        amount = simpledialog.askinteger("Restock", f"Enter restock quantity for '{item_name}':",
                                         parent=self.parent_frame, minvalue=1)
        if amount:
            item_id = self.inventory_tree.item(selected_item[0])['values'][0]
            # Find item and update stock
            for item in self.inventory_data:
                if item['id'] == item_id:
                    item['current_stock'] += amount
                    item['last_updated'] = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    break
            
            if data_manager.save_data('inventory', self.inventory_data):
                messagebox.showinfo("Success", f"{amount} units of '{item_name}' have been restocked.")
                self.refresh_inventory()
            else:
                messagebox.showerror("Error", "Failed to save restock information.")
                # Revert change

    def adjust_stock(self):
        """Manually adjust the stock of an item"""
        selected_item = self.inventory_tree.selection()
        if not selected_item:
            messagebox.showwarning("No Selection", "Please select an item to adjust.")
            return

        item_name = self.inventory_tree.item(selected_item[0])['values'][1]

        new_quantity = simpledialog.askinteger("Adjust Stock", f"Enter new stock quantity for '{item_name}':",
                                               parent=self.parent_frame, minvalue=0)
        
        if new_quantity is not None:
            item_id = self.inventory_tree.item(selected_item[0])['values'][0]
            # Find item and update stock
            for item in self.inventory_data:
                if item['id'] == item_id:
                    item['current_stock'] = new_quantity
                    item['last_updated'] = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    break
            
            if data_manager.save_data('inventory', self.inventory_data):
                messagebox.showinfo("Success", f"Stock for '{item_name}' has been adjusted to {new_quantity}.")
                self.refresh_inventory()
            else:
                messagebox.showerror("Error", "Failed to save stock adjustment.")
                # Revert change

    def export_inventory(self):
        """Export inventory data to a file"""
        
        export_dialog = tk.Toplevel(self.parent_frame)
        export_dialog.title("Export Inventory")
        export_dialog.geometry("400x250")
        export_dialog.resizable(False, False)
        export_dialog.transient(self.parent_frame)
        export_dialog.grab_set()

        # Style
        export_dialog.configure(bg=self.colors['background'])
        
        main_frame = tk.Frame(export_dialog, bg=self.colors['background'], padx=20, pady=20)
        main_frame.pack(fill="both", expand=True)

        title_label = tk.Label(main_frame, text="Export Options", font=self.fonts['heading'],
                               bg=self.colors['background'], fg=self.colors['text_primary'])
        title_label.pack(pady=(0, 20))
        
        # Format selection
        format_frame = tk.Frame(main_frame, bg=self.colors['background'])
        format_frame.pack(fill="x", pady=5)
        format_label = tk.Label(format_frame, text="File Format:", font=self.fonts['body'], bg=self.colors['background'])
        format_label.pack(side="left", padx=5)
        export_format = ttk.Combobox(format_frame, values=["Excel", "CSV", "PDF"], state="readonly", font=self.fonts['body'])
        export_format.set("Excel")
        export_format.pack(side="left", padx=5)

        # Content selection
        content_frame = tk.Frame(main_frame, bg=self.colors['background'])
        content_frame.pack(fill="x", pady=5)
        content_label = tk.Label(content_frame, text="Inventory Type:", font=self.fonts['body'], bg=self.colors['background'])
        content_label.pack(side="left", padx=5)
        inventory_type = ttk.Combobox(content_frame, values=["All", "Low Stock", "Out of Stock"], state="readonly", font=self.fonts['body'])
        inventory_type.set("All")
        inventory_type.pack(side="left", padx=5)

        # Action buttons
        button_frame = tk.Frame(main_frame, bg=self.colors['background'])
        button_frame.pack(fill="x", pady=(20, 0))

        def do_export():
            file_format = export_format.get()
            inv_type = inventory_type.get()
            
            from tkinter import filedialog
            file_extensions = {
                "Excel": ".xlsx",
                "CSV": ".csv",
                "PDF": ".pdf"
            }
            default_filename = f"inventory_export_{datetime.datetime.now().strftime('%Y%m%d')}{file_extensions[file_format]}"
            
            file_path = filedialog.asksaveasfilename(
                initialfile=default_filename,
                defaultextension=file_extensions[file_format],
                filetypes=[(f"{file_format} Files", f"*{file_extensions[file_format]}"), ("All Files", "*.*")]
            )
            
            if not file_path:
                return

            try:
                success = False
                if file_format == "Excel":
                    success = self.export_inventory_to_excel(file_path, inv_type)
                elif file_format == "CSV":
                    success = self.export_inventory_to_csv(file_path, inv_type)
                elif file_format == "PDF":
                    success = self.export_inventory_to_pdf(file_path, inv_type)
                
                if success:
                    messagebox.showinfo("Export Successful", f"Inventory data has been successfully exported to\n{file_path}", parent=export_dialog)
                else:
                    raise Exception("Export function returned False.")
            except Exception as e:
                messagebox.showerror("Export Failed", f"An error occurred during export:\n{e}", parent=export_dialog)
            finally:
                export_dialog.destroy()

        export_button = tk.Button(button_frame, text="Export", command=do_export,
                                  font=self.fonts['button'], bg=self.colors['success'], fg='white',
                                  bd=0, padx=15, pady=8)
        export_button.pack(side="right", padx=5)

        cancel_button = tk.Button(button_frame, text="Cancel", command=export_dialog.destroy,
                                  font=self.fonts['button'], bg=self.colors['error'], fg='white',
                                  bd=0, padx=15, pady=8)
        cancel_button.pack(side="right", padx=5)
        
    def export_inventory_to_excel(self, file_path: str, inventory_type: str) -> bool:
        """Export inventory data to an Excel file"""
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

            inventory_to_export = self.get_filtered_inventory(inventory_type)
            if not inventory_to_export:
                messagebox.showwarning("No Data", "There is no data to export for the selected filter.", parent=self.parent_frame)
                return False

            df = pd.DataFrame(inventory_to_export)
            
            # Create Excel writer
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Inventory')
                workbook = writer.book
                worksheet = writer.sheets['Inventory']

                # --- Formatting ---
                header_font = Font(name='Segoe UI', size=12, bold=True, color='FFFFFF')
                header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
                cell_font = Font(name='Segoe UI', size=11)
                center_alignment = Alignment(horizontal='center', vertical='center')
                border = Border(left=Side(style='thin'), 
                                right=Side(style='thin'), 
                                top=Side(style='thin'), 
                                bottom=Side(style='thin'))

                # Format header
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_alignment
                    cell.border = border
                
                # Format data cells
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, max_col=worksheet.max_column):
                    for cell in row:
                        cell.font = cell_font
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                        cell.border = border

                # Auto-adjust column widths
                for col in worksheet.columns:
                    max_length = 0
                    column = col[0].column_letter # Get the column name
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[column].width = adjusted_width
            
            return True
        except ImportError:
            messagebox.showerror("Dependency Error", "The 'pandas' and 'openpyxl' libraries are required for Excel export. Please install them.", parent=self.parent_frame)
            return False
        except Exception as e:
            print(f"Excel export failed: {e}")
            return False

    def export_inventory_to_csv(self, file_path: str, inventory_type: str) -> bool:
        """Export inventory data to a CSV file"""
        try:
            import csv
            inventory_to_export = self.get_filtered_inventory(inventory_type)
            if not inventory_to_export:
                messagebox.showwarning("No Data", "There is no data to export for the selected filter.", parent=self.parent_frame)
                return False
                
            headers = list(inventory_to_export[0].keys())
            with open(file_path, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.DictWriter(f, fieldnames=headers)
                writer.writeheader()
                writer.writerows(inventory_to_export)
            return True
        except Exception as e:
            print(f"CSV export failed: {e}")
            return False
            
    def export_inventory_to_pdf(self, file_path: str, inventory_type: str) -> bool:
        """Export inventory data to a PDF file"""
        try:
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter, landscape

            inventory_to_export = self.get_filtered_inventory(inventory_type)
            if not inventory_to_export:
                messagebox.showwarning("No Data", "There is no data to export for the selected filter.", parent=self.parent_frame)
                return False

            doc = SimpleDocTemplate(file_path, pagesize=landscape(letter))
            elements = []
            styles = getSampleStyleSheet()

            # Title
            title = Paragraph(f"Inventory Report ({inventory_type})", styles['h1'])
            elements.append(title)
            elements.append(Spacer(1, 12))
            
            # Data
            headers = list(inventory_to_export[0].keys())
            data = [headers] + [[str(item[h]) for h in headers] for item in inventory_to_export]
            
            # Create table
            table = Table(data)
            style = TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.grey),
                ('TEXTCOLOR',(0,0),(-1,0),colors.whitesmoke),
                ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0,0), (-1,0), 12),
                ('BACKGROUND', (0,1), (-1,-1), colors.beige),
                ('GRID', (0,0), (-1,-1), 1, colors.black)
            ])
            table.setStyle(style)
            
            elements.append(table)
            doc.build(elements)
            return True
        except ImportError:
            messagebox.showerror("Dependency Error", "The 'reportlab' library is required for PDF export. Please install it.", parent=self.parent_frame)
            return False
        except Exception as e:
            print(f"PDF export failed: {e}")
            return False
            
    def get_filtered_inventory(self, inventory_type: str) -> List[Dict]:
        """Get inventory data based on the selected filter"""
        if inventory_type == "All":
            return self.inventory_data
        elif inventory_type == "Low Stock":
            return [item for item in self.inventory_data if 0 < item['current_stock'] <= item['min_stock']]
        elif inventory_type == "Out of Stock":
            return [item for item in self.inventory_data if item['current_stock'] == 0]
        return []

    def refresh_inventory(self):
        """Refresh all inventory data and UI components"""
        self.inventory_data = self.load_inventory_data()
        self.refresh_inventory_list()
        self.refresh_possible_meals()
        messagebox.showinfo("Refresh", "Inventory data has been updated.")

    def refresh_data(self):
         """Refresh inventory module data (called by data manager)"""
         # Skip refresh if UI not initialized
         if not getattr(self, 'inventory_tree', None):
             print("⚠️ Inventory module not initialized, skipping refresh")
             return
         try:
             self.inventory_data = self.load_inventory_data()
             self.refresh_inventory_list()
             self.update_stats_cards()
             print("✅ Inventory module data refreshed")
         except tk.TclError:
             print("⚠️ Inventory module UI not active, skipping refresh")
         except Exception as e:
             print(f"❌ Inventory module data refresh failed: {e}")

    # --- Meal Possibility Calculation ---
    def load_recipe_data(self):
        """Load recipe data from data manager"""
        try:
            # Load JSON recipes and merge with meal definitions
            json_recipes = data_manager.load_data('recipes') or []
            meal_defs = {m['name']: m for m in data_manager.load_data('meals')}
            merged = []
            for r in json_recipes:
                ingredients = r.get('ingredients')
                # normalize ingredients list to dict
                if isinstance(ingredients, list):
                    ing_map = {ing.get('ingredient_name','').lower(): ing.get('quantity_per_serving',0) for ing in ingredients}
                else:
                    ing_map = {k.lower(): v for k,v in (ingredients or {}).items()}
                merged.append({
                    'meal_id': r.get('meal_id'),
                    'meal_name': r.get('meal_name'),
                    'ingredients': ing_map,
                    'category': meal_defs.get(r.get('meal_name'),'Other').get('category', 'Other') if isinstance(meal_defs.get(r.get('meal_name')), dict) else 'Other'
                })
            # Add fallback recipes for meals without explicit recipes
            for name, m in meal_defs.items():
                if not any(r['meal_name'] == name for r in merged):
                    merged.append({
                        'meal_id': m['id'],
                        'meal_name': name,
                        'ingredients': {ing.lower():1 for ing in m.get('ingredients',[])},
                        'category': m.get('category','Other')
                    })
            self.recipes = merged
            if not self.recipes:
                self.recipes = self.get_default_recipes()
        except Exception as e:
            print(f"Error loading recipe data: {e}")
            self.recipes = self.get_default_recipes()

    def get_default_recipes(self):
        """Returns a default list of recipes if the data file is unavailable."""
        return [
             {"name": "Tomato Beef Noodles", "ingredients": {"Tomato": 0.2, "Beef": 0.15, "Noodles": 1}},
             {"name": "Egg Fried Rice", "ingredients": {"Egg": 2, "Rice": 0.3}},
             {"name": "Beef Burger", "ingredients": {"Beef": 0.2, "Bread": 1, "Lettuce": 0.05}},
             {"name": "French Fries", "ingredients": {"Potato": 0.3}},
             {"name": "Salmon Set Meal", "ingredients": {"Salmon": 0.2, "Rice": 0.15}},
             {"name": "Homestyle Chicken Rice", "ingredients": {"Chicken": 0.2, "Rice": 0.15}},
             {"name": "Seafood Fried Rice", "ingredients": {"Shrimp": 0.1, "Rice": 0.15, "Egg": 1}},
             {"name": "Classic Beef Rice", "ingredients": {"Beef": 0.2, "Rice": 0.15}},
             {"name": "Vegetable Set Meal", "ingredients": {"Broccoli": 0.1, "Carrot": 0.1, "Rice": 0.15}},
             {"name": "Spicy Tofu", "ingredients": {"Tofu": 0.3, "Chili": 0.05}}
         ]

    def calculate_possible_meals(self):
        """Based on current inventory, calculate how many of each meal can be made."""
        if not self.inventory_data:
            return {}
        
        # Load recipes if not already loaded
        if not self.recipes:
            self.load_recipe_data()
        
        # Create inventory map, support singular and plural names
        inventory_map = {}
        for item in self.inventory_data:
            name = item['name'].lower()
            inventory_map[name] = item['current_stock']
            # 支持去除复数 's' 后的名称匹配
            if name.endswith('s'):
                inventory_map[name[:-1]] = item['current_stock']

        possible_meals = {}
        
        for recipe in self.recipes:
            # Use meal_name key from recipe, fallback to name
            meal_name = recipe.get("meal_name") or recipe.get("name")
            ingredients = recipe.get("ingredients")
            # Normalize ingredients into dict: handle list or dict format
            if isinstance(ingredients, list):
                ingredients_dict = {ing.get('ingredient_name', '').lower(): ing.get('quantity_per_serving', 0) for ing in ingredients}
            else:
                ingredients_dict = {k.lower(): v for k, v in (ingredients or {}).items()}

            if not meal_name or not ingredients_dict:
                continue
            
            can_make = float('inf')
            for ingredient, required_amount in ingredients_dict.items():
                stock = inventory_map.get(ingredient, 0)
                if stock < required_amount:
                    can_make = 0
                    break
                # Determine integer possible servings
                can_make = min(can_make, int(stock // required_amount))
            
            # Only add if at least one serving
            if can_make and can_make != float('inf'):
                possible_meals[meal_name] = {
                    "possible_servings": int(can_make),
                    "recipe": ingredients_dict
                }
        
        return possible_meals

    def create_possible_meals_section(self, parent):
        """Creates the section to display meals that can be made with current inventory."""
        section_frame = tk.Frame(parent, bg=self.colors['surface'])
        section_frame.pack(fill='x', pady=(0,10))

        header_frame = tk.Frame(section_frame, bg=self.colors['surface'])
        header_frame.pack(fill='x', padx=20, pady=(10,5))
        
        # Title for possible meals in English
        tk.Label(header_frame, text="Dishes to Make", font=self.fonts['subheading'], 
                 bg=self.colors['surface'], fg=self.colors['text_primary']).pack(side='left')
        
        # Create container frame for canvas and scrollbar
        container_frame = tk.Frame(section_frame, bg=self.colors['surface'])
        container_frame.pack(fill='x', padx=20, pady=(0, 10))
        
        # Create a canvas with horizontal scrollbar for the meals
        self.meals_canvas = tk.Canvas(container_frame, bg=self.colors['surface'], 
                                     highlightthickness=0, height=200)
        
        # Create horizontal scrollbar
        self.meals_scrollbar = ttk.Scrollbar(container_frame, orient="horizontal", 
                                           command=self.meals_canvas.xview)
        self.meals_canvas.configure(xscrollcommand=self.meals_scrollbar.set)
        
        # Pack canvas and scrollbar
        self.meals_canvas.pack(side="top", fill="both")
        self.meals_scrollbar.pack(side="bottom", fill="x")
        
        # Create the scrollable frame
        self.possible_meals_frame = tk.Frame(self.meals_canvas, bg=self.colors['surface'])
        self.canvas_frame_id = self.meals_canvas.create_window((0, 0), window=self.possible_meals_frame, anchor="nw")
        
        # Bind events
        self.meals_canvas.bind('<Configure>', self.on_canvas_configure)
        self.possible_meals_frame.bind('<Configure>', self.on_frame_configure)
        
        # Bind mouse wheel events for horizontal scrolling
        self.bind_scroll_events()

        self.refresh_possible_meals()
    
    def bind_scroll_events(self):
        """Bind mouse wheel and keyboard events for scrolling."""
        def bind_to_widget(widget):
            widget.bind("<MouseWheel>", self.on_mousewheel)
            widget.bind("<Button-4>", self.on_mousewheel)
            widget.bind("<Button-5>", self.on_mousewheel)
            widget.bind("<Shift-MouseWheel>", self.on_mousewheel)
            widget.bind("<Enter>", lambda e: widget.focus_set())
        
        # Bind to canvas and frame
        bind_to_widget(self.meals_canvas)
        bind_to_widget(self.possible_meals_frame)
    
    def on_canvas_configure(self, event):
        """Handle canvas resize event."""
        # Update the inner frame to fill the canvas width if it's smaller
        canvas_width = event.width
        
        # Get the actual frame width
        self.possible_meals_frame.update_idletasks()
        frame_width = self.possible_meals_frame.winfo_reqwidth()
        
        # Set the frame width appropriately
        if frame_width < canvas_width:
            # If frame is smaller than canvas, make it fill the canvas
            self.meals_canvas.itemconfig(self.canvas_frame_id, width=canvas_width)
        else:
            # If frame is larger, let it be its natural width for scrolling
            self.meals_canvas.itemconfig(self.canvas_frame_id, width=frame_width)
    
    def on_frame_configure(self, event):
        """Handle frame resize event to update scroll region."""
        # Update scroll region to encompass the entire frame
        bbox = self.meals_canvas.bbox("all")
        if bbox:
            self.meals_canvas.configure(scrollregion=bbox)
        
        # Ensure the canvas window is properly sized
        canvas_width = self.meals_canvas.winfo_width()
        frame_width = self.possible_meals_frame.winfo_reqwidth()
        
        if frame_width > canvas_width:
            self.meals_canvas.itemconfig(self.canvas_frame_id, width=frame_width)
    
    def on_mousewheel(self, event):
        """Handle mouse wheel scrolling."""
        # Check if horizontal scrolling is needed
        bbox = self.meals_canvas.bbox("all")
        canvas_width = self.meals_canvas.winfo_width()
        
        if bbox and bbox[2] > canvas_width:
            # Calculate scroll amount
            if hasattr(event, 'delta') and event.delta:
                # Windows - normal mouse wheel
                delta = int(-1 * (event.delta / 120))
            elif hasattr(event, 'num'):
                # Linux
                delta = -1 if event.num == 4 else 1 if event.num == 5 else 0
            else:
                delta = 0
            
            if delta != 0:
                self.meals_canvas.xview_scroll(delta, "units")

    def refresh_possible_meals(self):
        """Refresh the display of meals that can be made."""
        if not hasattr(self, 'possible_meals_frame') or not self.possible_meals_frame:
            return

        # Clear previous cards
        for widget in self.possible_meals_frame.winfo_children():
            widget.destroy()

        possible_meals = self.calculate_possible_meals()
        if not possible_meals:
            tk.Label(
                self.possible_meals_frame,
                text="Not enough ingredients to make any dishes.",
                font=self.fonts['body'], bg=self.colors['surface'], fg=self.colors['text_secondary']
            ).pack(pady=10)
            # Update canvas scroll region
            self.update_scroll_region()
            return

        # Display all cards in a single horizontal row for scrolling
        col = 0
        for meal_name, meal_info in possible_meals.items():
            self.create_meal_card(self.possible_meals_frame, meal_name, meal_info, 0, col)
            col += 1
        
        # Schedule scroll region update after UI renders
        self.possible_meals_frame.after(50, self.update_scroll_region)
    
    def update_scroll_region(self):
        """Update the canvas scroll region and ensure proper scrolling."""
        if not hasattr(self, 'meals_canvas') or not self.meals_canvas:
            return
            
        try:
            # Force update of all widgets
            self.possible_meals_frame.update_idletasks()
            self.meals_canvas.update_idletasks()
            
            # Get bounding box and update scroll region
            bbox = self.meals_canvas.bbox("all")
            if bbox:
                self.meals_canvas.configure(scrollregion=bbox)
                
                # Check if horizontal scrolling is needed
                canvas_width = self.meals_canvas.winfo_width()
                content_width = bbox[2] - bbox[0]
                
                if content_width > canvas_width:
                    # Content is wider than canvas, enable scrolling
                    self.meals_canvas.itemconfig(self.canvas_frame_id, width=content_width)
                    print(f"✓ Horizontal scroll enabled: content {content_width}px > canvas {canvas_width}px")
                else:
                    # Content fits in canvas, center it
                    self.meals_canvas.itemconfig(self.canvas_frame_id, width=canvas_width)
                    print(f"ⓘ No scroll needed: content {content_width}px <= canvas {canvas_width}px")
            
            # Ensure scrollbar visibility
            if hasattr(self, 'meals_scrollbar'):
                self.meals_scrollbar.update()
                
        except Exception as e:
            print(f"Error updating scroll region: {e}")
             
    def create_meal_card(self, parent, meal_name, meal_info, row, col):
        """Creates a small card for a meal that can be made."""
        # Fixed width for horizontal scrolling
        card_width = 180
        card = tk.Frame(parent, bg=self.colors['background'], bd=1, relief='solid', 
                       borderwidth=1, highlightbackground=self.colors['border'], 
                       width=card_width, height=160)
        card.grid(row=row, column=col, padx=8, pady=10, sticky='ns')
        card.grid_propagate(False)  # Prevent frame from shrinking
        
        servings = meal_info.get('possible_servings', 0)
        
        emoji_map = {'noodle': '🍜', 'rice': '🍚', 'burger': '🍔', 'fries': '🍟', 
                    'salmon': '🍣', 'chicken': '🍗', 'seafood': '🦐', 'beef': '🍖', 
                    'vegetable': '🥦', 'tofu': '🌶️'}
        emoji = '🍽️'
        for key, e in emoji_map.items():
            if key in meal_name.lower():
                emoji = e
                break

        # Emoji
        tk.Label(card, text=emoji, font=('Segoe UI Emoji', 24), 
                bg=self.colors['background']).pack(pady=(15,5))
        
        # Meal name with fixed height
        name_label = tk.Label(card, text=meal_name, font=self.fonts['small'], 
                             wraplength=160, justify='center', 
                             bg=self.colors['background'], fg=self.colors['text_primary'])
        name_label.pack(pady=(0,5), padx=5, fill='x')
        
        # Servings info
        tk.Label(card, text=f"Can make {servings} servings", 
                font=('Segoe UI', 11, 'bold'), bg=self.colors['background'], 
                fg=self.colors['success']).pack(pady=(0,15))

    def show_recipe_detail_dialog(self, meal_name, recipe, possible_servings):
        """Show a dialog with recipe details."""
        dialog = tk.Toplevel(self.parent_frame)
        dialog.title(f"Recipe Details - {meal_name}")
        dialog.geometry("450x400")
        dialog.configure(bg=self.colors['background'])
        dialog.resizable(False, False)

        # Header
        header = tk.Frame(dialog, bg=self.colors['primary'])
        header.pack(fill='x')
        tk.Label(header, text=meal_name, font=self.fonts['heading'], bg=self.colors['primary'], fg=self.colors['white']).pack(pady=10)

        # Info
        info_frame = tk.Frame(dialog, bg=self.colors['background'])
        info_frame.pack(padx=20, pady=10)
        tk.Label(info_frame, text=f"You can make: {possible_servings} servings", font=self.fonts['body'], bg=self.colors['background']).pack()

        # Ingredients list
        ingredients_frame = tk.Frame(dialog, bg=self.colors['surface'])
        ingredients_frame.pack(fill='both', expand=True, padx=20, pady=10)

        inventory_map = {item['name'].lower(): (item['current_stock'], item['unit']) for item in self.inventory_data}

        for ingredient, required in recipe.items():
            stock, unit = inventory_map.get(ingredient.lower(), (0, 'unit'))
            status_color = self.colors['success'] if stock >= required else self.colors['error']
            
            row = tk.Frame(ingredients_frame, bg=self.colors['surface'])
            row.pack(fill='x', padx=10, pady=3)
            
            tk.Label(row, text=f"{ingredient}:", font=self.fonts['body'], bg=self.colors['surface']).pack(side='left')
            tk.Label(row, text=f"Required {required}{unit}, Have {stock}{unit}", 
                     font=self.fonts['body'], fg=status_color, bg=self.colors['surface']).pack(side='right')

        # Close button
        tk.Button(dialog, text="Close", command=dialog.destroy, 
                  bg=self.colors['secondary'], fg=self.colors['white'], bd=0,
                  font=self.fonts['button'], padx=15, pady=5).pack(pady=10)

        dialog.transient(self.parent_frame)
        dialog.grab_set()
        self.parent_frame.wait_window(dialog)

class InventoryItemDialog:
    """A dialog for adding or editing an inventory item."""
    def __init__(self, parent, title, item_data=None):
        self.parent = parent
        self.title = title
        self.item_data = item_data
        self.result = None
        
        self.top = tk.Toplevel(parent)
        self.top.title(self.title)
        self.top.transient(parent)
        self.top.grab_set()
        self.top.resizable(False, False)
        self.top.configure(bg='#F0F0F0')

        self.vars = {
            "name": tk.StringVar(value=item_data.get('name', '') if item_data else ''),
            "category": tk.StringVar(value=item_data.get('category', 'Vegetable') if item_data else 'Vegetable'),
            "current_stock": tk.DoubleVar(value=item_data.get('current_stock', 0.0) if item_data else 0.0),
            "min_stock": tk.DoubleVar(value=item_data.get('min_stock', 10.0) if item_data else 10.0),
            "unit": tk.StringVar(value=item_data.get('unit', 'kg') if item_data else 'kg'),
            "price": tk.DoubleVar(value=item_data.get('price', 0.0) if item_data else 0.0),
            "supplier": tk.StringVar(value=item_data.get('supplier', '') if item_data else '')
        }
        
        self.create_dialog_ui()
        self.center_window()
        
        self.top.wait_window(self.top)

    def center_window(self):
        """Center the dialog on the parent window."""
        self.top.update_idletasks()
        width = self.top.winfo_width()
        height = self.top.winfo_height()
        x = (self.top.winfo_screenwidth() // 2) - (width // 2)
        y = (self.top.winfo_screenheight() // 2) - (height // 2)
        self.top.geometry(f"+{x}+{y}")
        
    def create_dialog_ui(self):
        main_frame = tk.Frame(self.top, bg='#FFFFFF', bd=5, relief='groove')
        main_frame.pack(padx=10, pady=10, fill='both', expand=True)

        # Form fields
        self.create_form_field(main_frame, "Name:", self.vars['name'], 'entry')
        self.create_form_field(main_frame, "Category:", self.vars['category'], 'combobox', 
                               options=['Vegetable', 'Meat', 'Staple', 'Beverage', 'Seasoning', 'Other'])
        self.create_form_field(main_frame, "Current Stock:", self.vars['current_stock'], 'spinbox')
        self.create_form_field(main_frame, "Min Stock:", self.vars['min_stock'], 'spinbox')
        self.create_form_field(main_frame, "Unit:", self.vars['unit'], 'entry')
        self.create_form_field(main_frame, "Unit Price:", self.vars['price'], 'spinbox')
        self.create_form_field(main_frame, "Supplier:", self.vars['supplier'], 'entry')

        # Buttons
        button_frame = tk.Frame(main_frame, bg='#FFFFFF')
        button_frame.pack(fill='x', padx=10, pady=(20, 10))
        
        ok_btn = tk.Button(button_frame, text="OK", width=12, command=self.ok)
        ok_btn.pack(side='right', padx=5)
        
        cancel_btn = tk.Button(button_frame, text="Cancel", width=12, command=self.cancel)
        cancel_btn.pack(side='right', padx=5)

    def create_form_field(self, parent, label_text, variable, field_type, options=None):
        """Helper to create a form field."""
        field_frame = tk.Frame(parent, bg='#FFFFFF')
        field_frame.pack(fill='x', padx=10, pady=5)
        
        label = tk.Label(field_frame, text=label_text, width=15, anchor='w', bg='#FFFFFF')
        label.pack(side='left')
        
        if field_type == 'entry':
            widget = tk.Entry(field_frame, textvariable=variable)
        elif field_type == 'spinbox':
            widget = tk.Spinbox(field_frame, from_=0, to=10000, textvariable=variable, width=18)
        elif field_type == 'combobox':
            widget = ttk.Combobox(field_frame, textvariable=variable, values=options, state='readonly')
        
        widget.pack(side='left', fill='x', expand=True)

    def ok(self):
        """Handle OK button click."""
        try:
            self.result = {
                "id": self.item_data.get('id', '') if self.item_data else '',
                "name": self.vars['name'].get(),
                "category": self.vars['category'].get(),
                "current_stock": self.vars['current_stock'].get(),
                "min_stock": self.vars['min_stock'].get(),
                "unit": self.vars['unit'].get(),
                "price": self.vars['price'].get(),
                "supplier": self.vars['supplier'].get(),
                "last_updated": datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }

            if not self.result["name"] or not self.result["unit"]:
                messagebox.showerror("Validation Error", "Name and Unit fields cannot be empty.", parent=self.top)
                self.result = None
                return
                
            self.top.destroy()
        except tk.TclError as e:
            messagebox.showerror("Input Error", f"Invalid input value: {e}", parent=self.top)
            self.result = None

    def cancel(self):
        """Handle Cancel button click."""
        self.result = None
        self.top.destroy()

if __name__ == "__main__":
    # 测试代码
    root = tk.Tk()
    root.title("现代化库存管理模块测试")
    root.geometry("1400x900")
    root.configure(bg="#f8f9fa")
    
    title_frame = tk.Frame(root, bg="#ffffff", height=70)
    title_frame.pack(fill="x")
    title_frame.pack_propagate(False)
    
    main_frame = tk.Frame(root, bg="#f8f9fa")
    main_frame.pack(fill="both", expand=True)
    
    inventory_module = ModernInventoryModule(main_frame, title_frame)
    inventory_module.show()
    
    root.mainloop()
