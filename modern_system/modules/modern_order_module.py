#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
现代化订单管理模块
基于现代化外卖平台风格的订单管理界面
"""

import tkinter as tk
from tkinter import ttk, messagebox, simpledialog
from typing import Dict, List, Any, Optional
import datetime
import json
import os

# 导入数据管理中心
try:
    from .data_manager import data_manager
except ImportError:
    try:
        from data_manager import data_manager
    except ImportError:
        # 如果导入失败，创建一个简单的模拟数据管理器
        class MockDataManager:
            def register_module(self, module_type, instance):
                pass
            def get_orders(self, status_filter=None):
                return []
            def update_order_status(self, order_id, new_status):
                return True
            def add_order(self, order_data):
                return "MOCK_ORDER_ID"
        data_manager = MockDataManager()

class ModernOrderModule:
    def __init__(self, parent_frame, title_frame, inventory_module=None, customer_module=None):
        self.parent_frame = parent_frame
        self.title_frame = title_frame
        self.inventory_module = inventory_module
        self.customer_module = customer_module
        
        # 注册到数据管理中心
        data_manager.register_module('order', self)
        
        # 现代化配色方案
        self.colors = {
            'primary': '#FF6B35',        # 主橙色
            'primary_dark': '#E5522A',   # 深橙色
            'secondary': '#4ECDC4',      # 青绿色
            'success': '#2ECC71',        # 成功绿
            'warning': '#F39C12',        # 警告橙
            'danger': '#E74C3C',         # 危险红
            'info': '#3498DB',           # 信息蓝
            'light': '#ECF0F1',          # 浅灰
            'dark': '#2C3E50',           # 深蓝灰
            'white': '#FFFFFF',          # 白色
            'background': '#F8F9FA',     # 背景色
            'card': '#FFFFFF',           # 卡片背景
            'border': '#E1E8ED',         # 边框色
            'text': '#2C3E50',           # 文本色
            'text_light': '#7F8C8D',     # 浅文本色
            'shadow': '#BDC3C7'          # 阴影色
        }
        
        # 订单状态配色
        self.status_colors = {
            '待接单': '#F39C12',
            '已接单': '#3498DB',
            '制作中': '#9B59B6',
            '已暂停': '#7f8c8d',
            '配送中': '#E67E22',
            '待取餐': '#16a085',
            '已完成': '#2ECC71',
            '已取消': '#E74C3C',
            '已归档': '#bdc3c7'
        }
        
        # 订单数据 - 从数据管理中心获取
        self.order_data = self.load_order_data()
        
        # 界面状态变量
        self.selected_order = None
        self.current_filter = "全部"
        self.search_keyword = ""
        self.stats_frame = None
        self.orders_container = None
    
    def load_order_data(self):
        """从数据管理中心加载订单数据"""
        try:
            orders = data_manager.get_orders()
            # 转换数据格式以适配现有界面
            formatted_orders = []
            for order in orders:
                # 处理菜品数据
                meals = []
                items = order.get('items', [])
                for item in items:
                    meal = {
                        "name": item.get('name', item.get('product_id', '未知菜品')),
                        "price": item.get('price', 0),
                        "quantity": item.get('quantity', 1)
                    }
                    meals.append(meal)
                
                formatted_order = {
                    "id": order.get('id', ''),
                    "customer": order.get('customer_name', order.get('table_number', '未知客户')),
                    "phone": order.get('customer_phone', order.get('phone', '')),
                    "address": order.get('delivery_address', order.get('address', '堂食')),
                    "meals": meals,
                    "total": order.get('total_amount', order.get('total', 0)),
                    "create_time": order.get('create_time', '').replace('T', ' ')[:16] if 'T' in order.get('create_time', '') else order.get('create_time', ''),
                    "status": order.get('status', '待处理'),
                    "type": order.get('order_type', order.get('type', '外卖')),
                    "payment": order.get('payment_method', order.get('payment', '现金')),
                    "note": order.get('note', '')
                }
                formatted_orders.append(formatted_order)
            
            # 如果没有数据或数据太少，使用默认示例数据
            if len(formatted_orders) < 3:
                print("订单数据较少，添加示例数据...")
                formatted_orders.extend(self.get_default_order_data())
            
            return formatted_orders
        except Exception as e:
            print(f"加载订单数据失败: {e}")
            return self.get_default_order_data()
    
    def get_default_order_data(self):
        """获取默认订单数据"""
        return [
            {
                "id": 1001, 
                "customer": "张三", 
                "phone": "138****1234",
                "address": "北京市朝阳区xxx街道1号",
                "meals": [
                    {"name": "番茄牛肉面", "price": 25.0, "quantity": 2},
                    {"name": "可乐", "price": 5.0, "quantity": 1}
                ],
                "total": 55.0, 
                "create_time": "2024-06-15 12:30", 
                "status": "已完成", 
                "type": "外卖",
                "payment": "微信支付",
                "note": "少放辣椒"
            },
            {
                "id": 1002, 
                "customer": "李四", 
                "phone": "139****5678",
                "address": "北京市海淀区xxx路88号",
                "meals": [
                    {"name": "鸡蛋炒饭", "price": 18.0, "quantity": 1}
                ],
                "total": 18.0, 
                "create_time": "2024-06-15 12:45", 
                "status": "制作中", 
                "type": "外卖",
                "payment": "支付宝",
                "note": ""
            },
            {
                "id": 1003, 
                "customer": "王五", 
                "phone": "136****9012",
                "address": "北京市西城区xxx胡同66号",
                "meals": [
                    {"name": "牛肉汉堡", "price": 32.0, "quantity": 3},
                    {"name": "薯条", "price": 12.0, "quantity": 2}
                ],
                "total": 120.0, 
                "create_time": "2024-06-15 11:20", 
                "status": "待接单", 
                "type": "外卖",
                "payment": "现金",
                "note": "汉堡不要洋葱"
            },
            {
                "id": 1004, 
                "customer": "赵六", 
                "phone": "137****3456",
                "address": "堂食",
                "meals": [
                    {"name": "红烧肉", "price": 35.0, "quantity": 1},
                    {"name": "米饭", "price": 3.0, "quantity": 2}
                ],
                "total": 41.0, 
                "create_time": "2024-06-15 13:15", 
                "status": "配送中", 
                "type": "堂食",
                "payment": "微信支付",
                "note": ""            }
        ]
        
        self.selected_order = None
        self.current_filter = "全部"
        self.search_keyword = ""
    
    def create_status_card(self, parent, status, count, color):
        """创建状态统计卡片"""
        card_frame = tk.Frame(parent, bg=self.colors['card'], relief='flat', bd=1,
                             highlightbackground=self.colors['border'], highlightthickness=1)
        card_frame.pack(side='left', padx=10, pady=5, fill='both', expand=True)
        
        # 设置最小尺寸
        card_frame.configure(width=200, height=100)
        
        # 状态图标和数字
        icon_frame = tk.Frame(card_frame, bg=color, width=80, height=80)
        icon_frame.pack(side='left', padx=15, pady=10)
        icon_frame.pack_propagate(False)
        
        count_label = tk.Label(icon_frame, text=str(count), font=('Microsoft YaHei UI', 18, 'bold'),
                              bg=color, fg=self.colors['white'])
        count_label.pack(expand=True)
        
        # 状态信息
        info_frame = tk.Frame(card_frame, bg=self.colors['card'])
        info_frame.pack(side='left', padx=(0, 15), pady=15, fill='both', expand=True)
        
        status_label = tk.Label(info_frame, text=status, font=('Microsoft YaHei UI', 12, 'bold'),
                               bg=self.colors['card'], fg=self.colors['text'])
        status_label.pack(anchor='w', pady=(5, 0))
        
        desc_label = tk.Label(info_frame, text='订单数量', font=('Microsoft YaHei UI', 10),
                             bg=self.colors['card'], fg=self.colors['text_light'])
        desc_label.pack(anchor='w', pady=(0, 5))
        
        return card_frame
    
    def create_order_card(self, parent, order):
        """创建订单卡片"""
        card_frame = tk.Frame(parent, bg=self.colors['card'], relief='flat', bd=1,
                             highlightbackground=self.colors['border'], highlightthickness=1)
        card_frame.pack(fill='x', padx=5, pady=5)
        
        # 卡片头部
        header_frame = tk.Frame(card_frame, bg=self.colors['card'], height=50)
        header_frame.pack(fill='x', padx=15, pady=(15, 10))
        header_frame.pack_propagate(False)
        
        # 订单号和状态
        order_info_frame = tk.Frame(header_frame, bg=self.colors['card'])
        order_info_frame.pack(side='left', fill='y')
        
        order_id_label = tk.Label(order_info_frame, text=f"#{order['id']}", 
                                 font=('Microsoft YaHei UI', 14, 'bold'),
                                 bg=self.colors['card'], fg=self.colors['primary'])
        order_id_label.pack(anchor='w')
        
        time_label = tk.Label(order_info_frame, text=order['create_time'], 
                             font=('Microsoft YaHei UI', 10),
                             bg=self.colors['card'], fg=self.colors['text_light'])
        time_label.pack(anchor='w')
        
        # 状态标签
        status_color = self.status_colors.get(order['status'], self.colors['info'])
        status_frame = tk.Frame(header_frame, bg=status_color, padx=10, pady=5)
        status_frame.pack(side='right', pady=5)
        
        status_label = tk.Label(status_frame, text=order['status'], 
                               font=('Microsoft YaHei UI', 10, 'bold'),
                               bg=status_color, fg=self.colors['white'])
        status_label.pack()
        
        # 客户信息
        customer_frame = tk.Frame(card_frame, bg=self.colors['card'])
        customer_frame.pack(fill='x', padx=15, pady=5)
        
        customer_label = tk.Label(customer_frame, text=f"👤 {order['customer']} | 📞 {order['phone']}", 
                                 font=('Microsoft YaHei UI', 11),
                                 bg=self.colors['card'], fg=self.colors['text'])
        customer_label.pack(anchor='w')
        
        address_label = tk.Label(customer_frame, text=f"📍 {order['address']}", 
                                font=('Microsoft YaHei UI', 10),
                                bg=self.colors['card'], fg=self.colors['text_light'])
        address_label.pack(anchor='w')
        
        # 菜品信息
        meals_frame = tk.Frame(card_frame, bg=self.colors['background'], padx=10, pady=8)
        meals_frame.pack(fill='x', padx=15, pady=5)
        
        for meal in order['meals']:
            meal_item = tk.Label(meals_frame, 
                               text=f"🍽️ {meal['name']} × {meal['quantity']} = ¥{meal['price'] * meal['quantity']:.2f}", 
                               font=('Microsoft YaHei UI', 10),
                               bg=self.colors['background'], fg=self.colors['text'],
                               anchor='w')
            meal_item.pack(fill='x', pady=2)
        
        # 订单总额和操作按钮
        bottom_frame = tk.Frame(card_frame, bg=self.colors['card'])
        bottom_frame.pack(fill='x', padx=15, pady=(5, 15))
        
        # 总额
        total_label = tk.Label(bottom_frame, text=f"总计：¥{order['total']:.2f}", 
                              font=('Microsoft YaHei UI', 12, 'bold'),
                              bg=self.colors['card'], fg=self.colors['primary'])
        total_label.pack(side='left')
        
        # 支付方式和类型
        payment_label = tk.Label(bottom_frame, text=f"{order['payment']} | {order['type']}", 
                               font=('Microsoft YaHei UI', 9),
                               bg=self.colors['card'], fg=self.colors['text_light'])
        payment_label.pack(side='left', padx=(20, 0))
        
        # 操作按钮
        actions_frame = tk.Frame(bottom_frame, bg=self.colors['card'])
        actions_frame.pack(side='right')

        # 查看详情按钮
        detail_btn = tk.Button(actions_frame, text="查看详情",
                              font=('Microsoft YaHei UI', 9),
                              bg=self.colors['info'], fg=self.colors['white'],
                              bd=0, padx=15, pady=5, cursor='hand2',
                              command=lambda o=order: self.show_order_detail(o))
        detail_btn.pack(side='right', padx=5)

        # 动态添加状态操作按钮
        order_status = order.get('status', '未知')
        order_type = order.get('type', '外卖')

        if order_status == '待接单':
            self.add_action_button(actions_frame, "接单", self.colors['success'],
                                   lambda o=order: self.update_order_status(o['id'], '已接单'))
            self.add_action_button(actions_frame, "取消", self.colors['danger'],
                                   lambda o=order: self.update_order_status(o['id'], '已取消'))

        elif order_status == '已接单':
            self.add_action_button(actions_frame, "开始制作", self.colors['warning'],
                                   lambda o=order: self.update_order_status(o['id'], '制作中'))
            self.add_action_button(actions_frame, "取消", self.colors['danger'],
                                   lambda o=order: self.update_order_status(o['id'], '已取消'))

        elif order_status == '制作中':
            next_status = '配送中' if order_type == '外卖' else '待取餐'
            self.add_action_button(actions_frame, "制作完成", self.colors['primary'],
                                   lambda o=order, s=next_status: self.update_order_status(o['id'], s))
            self.add_action_button(actions_frame, "暂停", '#7f8c8d',
                                   lambda o=order: self.update_order_status(o['id'], '已暂停'))

        elif order_status == '已暂停':
             self.add_action_button(actions_frame, "继续制作", self.colors['success'],
                                   lambda o=order: self.update_order_status(o['id'], '制作中'))

        elif order_status == '配送中' or order_status == '待取餐':
            self.add_action_button(actions_frame, "已送达", self.colors['success'],
                                   lambda o=order: self.update_order_status(o['id'], '已完成'))

        elif order_status == '已完成':
            self.add_action_button(actions_frame, "归档", self.colors['info'],
                                   lambda o=order: self.update_order_status(o['id'], '已归档'))
        
        # 备注信息
        if order.get('note'):
            note_frame = tk.Frame(card_frame, bg=self.colors['light'], padx=10, pady=5)
            note_frame.pack(fill='x', padx=15, pady=(0, 15))
            
            note_label = tk.Label(note_frame, text=f"📝 备注：{order['note']}", 
                                 font=('Microsoft YaHei UI', 9),
                                 bg=self.colors['light'], fg=self.colors['text'])
            note_label.pack(anchor='w')
        
        return card_frame

    def add_action_button(self, parent, text, color, command):
        """辅助函数，用于创建标准化的操作按钮"""
        btn = tk.Button(parent, text=text,
                        font=('Microsoft YaHei UI', 9),
                        bg=color, fg=self.colors['white'],
                        bd=0, padx=15, pady=5, cursor='hand2',
                        command=command)
        btn.pack(side='right', padx=5)

    def update_order_status(self, order_id, new_status):
        """更新订单状态并刷新UI"""
        success = data_manager.update_order_status(order_id, new_status)
        if success:
            messagebox.showinfo("成功", f"订单 #{order_id} 状态已更新为: {new_status}")
            # 从数据库重新加载数据以确保一致性
            self.refresh_data()
        else:
            messagebox.showerror("失败", "更新订单状态失败，请重试")
    
    def show_order_detail(self, order):
        """显示订单详情"""
        detail_window = tk.Toplevel()
        detail_window.title(f"订单详情 - #{order['id']}")
        detail_window.geometry("500x600")
        detail_window.configure(bg=self.colors['background'])
        detail_window.resizable(False, False)
        
        # 标题
        title_frame = tk.Frame(detail_window, bg=self.colors['primary'], height=60)
        title_frame.pack(fill='x')
        title_frame.pack_propagate(False)
        
        title_label = tk.Label(title_frame, text="【测试修改V2】订单详情 #" + str(order['id']),
                              font=('Microsoft YaHei UI', 16, 'bold'),
                              bg=self.colors['primary'], fg=self.colors['white'])
        title_label.pack(expand=True)
        
        # 详情内容
        content_frame = tk.Frame(detail_window, bg=self.colors['background'])
        content_frame.pack(fill='both', expand=True, padx=20, pady=20)
        
        # 基本信息
        info_frame = tk.Frame(content_frame, bg=self.colors['card'], padx=20, pady=15)
        info_frame.pack(fill='x', pady=(0, 10))
        
        tk.Label(info_frame, text="基本信息", font=('Microsoft YaHei UI', 12, 'bold'),
                bg=self.colors['card'], fg=self.colors['text']).pack(anchor='w')
        
        info_text = f"""订单号：#{order['id']}
客户姓名：{order['customer']}
联系电话：{order['phone']}
配送地址：{order['address']}
订单类型：{order['type']}
支付方式：{order['payment']}
下单时间：{order['create_time']}
订单状态：{order['status']}"""
        
        tk.Label(info_frame, text=info_text, font=('Microsoft YaHei UI', 10),
                bg=self.colors['card'], fg=self.colors['text'], justify='left').pack(anchor='w', pady=(10, 0))
        
        # 菜品信息
        meals_frame = tk.Frame(content_frame, bg=self.colors['card'], padx=20, pady=15)
        meals_frame.pack(fill='x', pady=(0, 10))
        
        tk.Label(meals_frame, text="菜品信息", font=('Microsoft YaHei UI', 12, 'bold'),
                bg=self.colors['card'], fg=self.colors['text']).pack(anchor='w')
        
        for meal in order['meals']:
            meal_frame = tk.Frame(meals_frame, bg=self.colors['background'], padx=10, pady=5)
            meal_frame.pack(fill='x', pady=5)
            
            tk.Label(meal_frame, text=meal['name'], font=('Microsoft YaHei UI', 10, 'bold'),
                    bg=self.colors['background'], fg=self.colors['text']).pack(side='left')
            
            tk.Label(meal_frame, text=f"¥{meal['price']:.2f} × {meal['quantity']} = ¥{meal['price'] * meal['quantity']:.2f}", 
                    font=('Microsoft YaHei UI', 10),
                    bg=self.colors['background'], fg=self.colors['text']).pack(side='right')
        
        # 总计
        total_frame = tk.Frame(meals_frame, bg=self.colors['primary'], padx=10, pady=8)
        total_frame.pack(fill='x', pady=(10, 0))
        
        tk.Label(total_frame, text=f"订单总计：¥{order['total']:.2f}", 
                font=('Microsoft YaHei UI', 12, 'bold'),
                bg=self.colors['primary'], fg=self.colors['white']).pack()
        
        # 备注信息
        if order['note']:
            note_frame = tk.Frame(content_frame, bg=self.colors['card'], padx=20, pady=15)
            note_frame.pack(fill='x', pady=(0, 10))
            
            tk.Label(note_frame, text="备注信息", font=('Microsoft YaHei UI', 12, 'bold'),
                    bg=self.colors['card'], fg=self.colors['text']).pack(anchor='w')
            
            tk.Label(note_frame, text=order['note'], font=('Microsoft YaHei UI', 10),
                    bg=self.colors['card'], fg=self.colors['text'], wraplength=400).pack(anchor='w', pady=(10, 0))
          # 关闭按钮
        tk.Button(content_frame, text="关闭", font=('Microsoft YaHei UI', 10),
                 bg=self.colors['text_light'], fg=self.colors['white'],
                 bd=0, padx=30, pady=8, cursor='hand2',
                 command=detail_window.destroy).pack(pady=20)
    
    def filter_orders(self, status):
        """筛选订单"""
        self.current_filter = status
        self.refresh_order_list()
    
    def refresh_data(self):
        """从数据库重新加载数据并刷新整个UI"""
        # 1. 从数据库加载最新数据
        self.order_data = self.load_order_data()
        
        # 2. 更新统计卡片
        self.update_statistics()
        
        # 3. 刷新订单列表
        if hasattr(self, 'orders_container'):
            self.refresh_order_list()
    
    def update_statistics(self):
        """更新统计信息"""
        if not self.stats_frame:
            return
            
        # 清空现有统计卡片
        for widget in self.stats_frame.winfo_children():
            widget.destroy()
        
        # 统计各状态订单数量
        status_counts = {status: 0 for status in self.status_colors}
        for order in self.order_data:
            status = order.get('status', '未知')
            if status in status_counts:
                status_counts[status] += 1
        
        # 重新创建统计卡片
        # 我们只显示有订单的状态，或者一些关键状态，避免UI过于拥挤
        key_statuses = ['待接单', '制作中', '配送中', '待取餐', '已完成', '已取消']
        
        # 添加当前存在订单的其他状态
        for status, count in status_counts.items():
            if count > 0 and status not in key_statuses:
                key_statuses.append(status)

        for status in key_statuses:
             if status in status_counts:
                count = status_counts[status]
                color = self.status_colors.get(status, '#bdc3c7')
                self.create_status_card(self.stats_frame, status, count, color)
    
    def add_new_order(self):
        """添加新订单"""
        # 创建新订单窗口
        order_window = tk.Toplevel()
        order_window.title("新建订单")
        order_window.geometry("600x800")  # 增加高度从700到800
        order_window.configure(bg=self.colors['background'])
        order_window.resizable(False, False)
        
        # 标题
        title_frame = tk.Frame(order_window, bg=self.colors['primary'], height=60)
        title_frame.pack(fill='x')
        title_frame.pack_propagate(False)
        
        title_label = tk.Label(title_frame, text="新建订单", 
                              font=('Microsoft YaHei UI', 16, 'bold'),
                              bg=self.colors['primary'], fg=self.colors['white'])
        title_label.pack(expand=True)
        
        # 表单内容
        form_frame = tk.Frame(order_window, bg=self.colors['background'])
        form_frame.pack(fill='both', expand=True, padx=20, pady=20)
        
        # 客户信息
        customer_frame = tk.Frame(form_frame, bg=self.colors['card'], padx=20, pady=15)
        customer_frame.pack(fill='x', pady=(0, 10))
        
        tk.Label(customer_frame, text="客户信息", font=('Microsoft YaHei UI', 12, 'bold'),
                bg=self.colors['card'], fg=self.colors['text']).pack(anchor='w')
        
        # 客户姓名
        name_frame = tk.Frame(customer_frame, bg=self.colors['card'])
        name_frame.pack(fill='x', pady=5)
        tk.Label(name_frame, text="客户姓名:", font=('Microsoft YaHei UI', 10),
                bg=self.colors['card'], fg=self.colors['text']).pack(side='left')
        customer_name_var = tk.StringVar(order_window)
        name_entry = tk.Entry(name_frame, textvariable=customer_name_var, font=('Microsoft YaHei UI', 10))
        name_entry.pack(side='right', fill='x', expand=True, padx=(10, 0))
        
        # 联系电话
        phone_frame = tk.Frame(customer_frame, bg=self.colors['card'])
        phone_frame.pack(fill='x', pady=5)
        tk.Label(phone_frame, text="联系电话:", font=('Microsoft YaHei UI', 10),
                bg=self.colors['card'], fg=self.colors['text']).pack(side='left')
        customer_phone_var = tk.StringVar(order_window)
        phone_entry = tk.Entry(phone_frame, textvariable=customer_phone_var, font=('Microsoft YaHei UI', 10))
        phone_entry.pack(side='right', fill='x', expand=True, padx=(10, 0))
        
        # 配送地址
        address_frame = tk.Frame(customer_frame, bg=self.colors['card'])
        address_frame.pack(fill='x', pady=5)
        tk.Label(address_frame, text="配送地址:", font=('Microsoft YaHei UI', 10),
                bg=self.colors['card'], fg=self.colors['text']).pack(side='left')
        customer_address_var = tk.StringVar(order_window)
        address_entry = tk.Entry(address_frame, textvariable=customer_address_var, font=('Microsoft YaHei UI', 10))
        address_entry.pack(side='right', fill='x', expand=True, padx=(10, 0))
        
        # 订单类型
        type_frame = tk.Frame(customer_frame, bg=self.colors['card'])
        type_frame.pack(fill='x', pady=5)
        tk.Label(type_frame, text="订单类型:", font=('Microsoft YaHei UI', 10),
                bg=self.colors['card'], fg=self.colors['text']).pack(side='left')
        order_type_var = tk.StringVar(order_window, value="外卖")
        type_combo = ttk.Combobox(type_frame, textvariable=order_type_var, 
                                 values=["外卖", "堂食"], state="readonly")
        type_combo.pack(side='right', padx=(10, 0))
        
        # 支付方式
        payment_frame = tk.Frame(customer_frame, bg=self.colors['card'])
        payment_frame.pack(fill='x', pady=5)
        tk.Label(payment_frame, text="支付方式:", font=('Microsoft YaHei UI', 10),
                bg=self.colors['card'], fg=self.colors['text']).pack(side='left')
        payment_var = tk.StringVar(order_window, value="微信支付")
        payment_combo = ttk.Combobox(payment_frame, textvariable=payment_var, 
                                    values=["微信支付", "支付宝", "现金", "银行卡"], state="readonly")
        payment_combo.pack(side='right', padx=(10, 0))
        
        # 备注
        note_frame = tk.Frame(customer_frame, bg=self.colors['card'])
        note_frame.pack(fill='x', pady=5)
        tk.Label(note_frame, text="订单备注:", font=('Microsoft YaHei UI', 10),
                bg=self.colors['card'], fg=self.colors['text']).pack(anchor='w')
        note_var = tk.StringVar(order_window)
        note_entry = tk.Entry(note_frame, textvariable=note_var, font=('Microsoft YaHei UI', 10))
        note_entry.pack(fill='x', pady=(5, 0))
        
        # 菜品选择
        meals_frame = tk.Frame(form_frame, bg=self.colors['card'], padx=20, pady=15)
        meals_frame.pack(fill='both', expand=True, pady=(0, 10))
        
        tk.Label(meals_frame, text="菜品选择", font=('Microsoft YaHei UI', 12, 'bold'),
                bg=self.colors['card'], fg=self.colors['text']).pack(anchor='w')
        
        # 简化的菜品选择（实际应该从菜品模块获取）
        sample_meals = [
            {"name": "番茄牛肉面", "price": 25.0},
            {"name": "鸡蛋炒饭", "price": 18.0},
            {"name": "牛肉汉堡", "price": 32.0},
            {"name": "红烧肉", "price": 35.0},
            {"name": "可乐", "price": 5.0},
            {"name": "米饭", "price": 3.0}
        ]
        
        selected_meals = []
        meal_vars = {}
        
        for meal in sample_meals:
            meal_frame = tk.Frame(meals_frame, bg=self.colors['background'], padx=10, pady=5)
            meal_frame.pack(fill='x', pady=2)
            
            var = tk.IntVar()
            meal_vars[meal['name']] = var
            
            cb = tk.Checkbutton(meal_frame, text=f"{meal['name']} - ¥{meal['price']:.2f}",
                               variable=var, font=('Microsoft YaHei UI', 10),
                               bg=self.colors['background'], fg=self.colors['text'])
            cb.pack(side='left')
            
            # 数量选择
            qty_var = tk.IntVar(value=1)
            meal_vars[f"{meal['name']}_qty"] = qty_var
            
            tk.Label(meal_frame, text="数量:", font=('Microsoft YaHei UI', 9),
                    bg=self.colors['background'], fg=self.colors['text']).pack(side='right', padx=(0, 5))
            
            qty_spinbox = tk.Spinbox(meal_frame, from_=1, to=10, width=5, textvariable=qty_var)
            qty_spinbox.pack(side='right')
        
        # 按钮
        button_frame = tk.Frame(form_frame, bg=self.colors['background'])
        button_frame.pack(fill='x', pady=10)
        
        def save_order():
            # 验证输入
            if not customer_name_var.get() or not customer_phone_var.get():
                messagebox.showerror("错误", "请填写客户姓名和联系电话")
                return
            
            # 收集选中的菜品
            order_meals = []
            total_amount = 0
            
            for meal in sample_meals:
                if meal_vars[meal['name']].get():
                    quantity = meal_vars[f"{meal['name']}_qty"].get()
                    order_meals.append({
                        "name": meal['name'],
                        "price": meal['price'],
                        "quantity": quantity
                    })
                    total_amount += meal['price'] * quantity
            
            if not order_meals:
                messagebox.showerror("错误", "请至少选择一个菜品")
                return
            
            try:
                # 准备订单数据用于库存检查
                order_items = []
                for meal in order_meals:
                    order_items.append({
                        'product_id': meal['name'],  # 使用菜品名称作为产品ID
                        'quantity': meal['quantity']
                    })
                
                # 创建订单数据
                order_data = {
                    "customer_name": customer_name_var.get(),
                    "phone": customer_phone_var.get(),
                    "address": customer_address_var.get() if customer_address_var.get() else "堂食",
                    "items": order_items,
                    "meals": order_meals,  # 保留原有的meals格式用于显示
                    "total_amount": total_amount,
                    "type": order_type_var.get(),
                    "payment": payment_var.get(),
                    "note": note_var.get(),
                    "status": "待接单"
                }
                
                # 使用数据管理器创建订单（包含库存检查）
                try:
                    order_id = data_manager.create_order(order_data)
                    messagebox.showinfo("成功", f"订单 #{order_id} 创建成功！")
                    order_window.destroy()
                    self.refresh_order_list()
                except ValueError as e:
                    if "库存不足" in str(e):
                        messagebox.showerror("库存不足", "当前库存不足，无法创建订单。\n请检查菜品库存后重试。")
                    else:
                        messagebox.showerror("错误", f"创建订单失败：{e}")
                    return
                except Exception as e:
                    messagebox.showerror("错误", f"创建订单失败：{e}")
                    return
                    
            except Exception as e:
                messagebox.showerror("错误", f"订单处理失败：{e}")
                return
        
        save_btn = tk.Button(button_frame, text="保存订单", 
                           font=('Microsoft YaHei UI', 10, 'bold'),
                           bg=self.colors['primary'], fg=self.colors['white'],
                           bd=0, padx=30, pady=8, cursor='hand2',
                           command=save_order)
        save_btn.pack(side='right', padx=5)
        
        cancel_btn = tk.Button(button_frame, text="取消", 
                             font=('Microsoft YaHei UI', 10),
                             bg=self.colors['text_light'], fg=self.colors['white'],
                             bd=0, padx=30, pady=8, cursor='hand2',
                             command=order_window.destroy)
        cancel_btn.pack(side='right', padx=5)
    
    def on_data_changed(self, event_type, data):
        """处理数据变更通知"""
        if event_type in ['order_added', 'order_updated']:
            # 刷新订单数据
            self.order_data = self.load_order_data()
            # 如果当前正在显示订单界面，刷新显示
            if hasattr(self, 'order_list_frame'):
                self.refresh_order_list()
    
    def refresh_order_list(self):
        """刷新订单列表"""
        # 重新加载订单数据
        self.order_data = self.load_order_data()
        
        # 清空容器（如果存在）
        if hasattr(self, 'orders_container') and self.orders_container:
            for widget in self.orders_container.winfo_children():
                widget.destroy()
        
        # 筛选和搜索订单
        filtered_orders = self.order_data
        
        # 应用状态筛选
        if self.current_filter != "全部":
            filtered_orders = [order for order in self.order_data if order['status'] == self.current_filter]
        
        # 应用搜索
        if hasattr(self, 'search_keyword') and self.search_keyword:
            filtered_orders = [order for order in filtered_orders 
                              if self.search_keyword.lower() in order.get('customer', '').lower() 
                              or self.search_keyword in str(order['id'])
                              or self.search_keyword.lower() in order.get('phone', '').lower()]
        
        # 创建订单卡片
        if hasattr(self, 'orders_container') and self.orders_container:
            for order in filtered_orders:
                self.create_order_card(self.orders_container, order)
        
        # 更新统计信息
        self.update_statistics()
    
    def update_title_frame(self):
        """更新标题框架，但保留面包屑导航"""
        # 不清空整个title_frame，而是查找并更新特定元素
        # 如果title_frame为空或没有找到合适的元素，创建新的标题
        found_title = False
        
        try:
            for widget in self.title_frame.winfo_children():
                if isinstance(widget, tk.Frame):
                    for child in widget.winfo_children():
                        if isinstance(child, tk.Label):
                            text = child.cget("text")
                            # 如果是模块标题（不是面包屑），则更新
                            if "订单管理" in text or ("管理" in text and "首页" not in text):
                                child.configure(text="📋 订单管理")
                                found_title = True
                                break
                    if found_title:
                        break
        except tk.TclError:
            # Widget可能已被销毁
            pass
        
        # 如果没有找到现有标题，创建新的标题区域
        if not found_title:
            # 标题栏
            title_container = tk.Frame(self.title_frame, bg=self.colors['white'])
            title_container.pack(fill='x', side='bottom')  # 放在底部，不影响面包屑
            
            # 标题
            title_label = tk.Label(title_container, text="📋 订单管理", 
                                  font=('Microsoft YaHei UI', 18, 'bold'),
                                  bg=self.colors['white'], fg=self.colors['text'])
            title_label.pack(side='left', padx=20, pady=15)
            
            # 操作按钮
            actions_frame = tk.Frame(title_container, bg=self.colors['white'])
            actions_frame.pack(side='right', padx=20, pady=15)
            
            # 新建订单按钮
            add_btn = tk.Button(actions_frame, text="➕ 新建订单", 
                               font=('Microsoft YaHei UI', 10, 'bold'),
                               bg=self.colors['primary'], fg=self.colors['white'],
                               bd=0, padx=20, pady=8, cursor='hand2',
                               command=self.add_new_order)
            add_btn.pack(side='right', padx=5)
            
            # 刷新按钮
            refresh_btn = tk.Button(actions_frame, text="🔄 刷新", 
                                   font=('Microsoft YaHei UI', 10),
                                   bg=self.colors['info'], fg=self.colors['white'],
                                   bd=0, padx=20, pady=8, cursor='hand2',
                                   command=self.refresh_order_list)
            refresh_btn.pack(side='right', padx=5)
            
            # 导出按钮
            export_btn = tk.Button(actions_frame, text="📊 导出", 
                                  font=('Microsoft YaHei UI', 10),
                                  bg=self.colors['success'], fg=self.colors['white'],
                                  bd=0, padx=20, pady=8, cursor='hand2',
                                  command=self.export_orders)
            export_btn.pack(side='right', padx=5)
    
    def show(self):
        """显示订单管理界面"""
        self.clear_frames()
        self.update_title_frame()
        
        # 创建主容器
        main_frame = tk.Frame(self.parent_frame, bg=self.colors['background'])
        main_frame.pack(fill='both', expand=True, padx=20, pady=(0, 20))
        
        # 顶部统计信息
        self.stats_frame = tk.Frame(main_frame, bg=self.colors['background'])
        self.stats_frame.pack(fill='x', pady=(10, 5))
        
        # 筛选和搜索
        self.create_filter_bar(main_frame)
        
        # 订单列表
        self.create_order_list(main_frame)
        
        # 首次加载时刷新数据
        self.refresh_data()

    def clear_frames(self):
        """清空所有子框架"""
        for widget in self.parent_frame.winfo_children():
            widget.destroy()

    def create_filter_bar(self, parent):
        """创建筛选栏"""
        filter_frame = tk.Frame(parent, bg=self.colors['background'])
        filter_frame.pack(fill='x', pady=(0, 20))
        
        tk.Label(filter_frame, text="筛选订单：", font=('Microsoft YaHei UI', 12, 'bold'),
                bg=self.colors['background'], fg=self.colors['text']).pack(side='left')
        
        filter_buttons = ["全部", "待接单", "已接单", "制作中", "配送中", "已完成", "已取消"]
        for filter_name in filter_buttons:
            btn_color = self.colors['primary'] if filter_name == self.current_filter else self.colors['light']
            text_color = self.colors['white'] if filter_name == self.current_filter else self.colors['text']
            
            filter_btn = tk.Button(filter_frame, text=filter_name, 
                                  font=('Microsoft YaHei UI', 9),
                                  bg=btn_color, fg=text_color,
                                  bd=0, padx=15, pady=5, cursor='hand2',
                                  command=lambda f=filter_name: self.filter_orders(f))
            filter_btn.pack(side='left', padx=5)
    
    def create_order_list(self, parent):
        """创建订单列表"""
        # 订单列表容器
        list_frame = tk.Frame(parent, bg=self.colors['background'])
        list_frame.pack(fill='both', expand=True)
        
        # 滚动区域
        canvas = tk.Canvas(list_frame, bg=self.colors['background'], highlightthickness=0)
        scrollbar = ttk.Scrollbar(list_frame, orient="vertical", command=canvas.yview)
        self.orders_container = tk.Frame(canvas, bg=self.colors['background'])
        
        self.orders_container.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=self.orders_container, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # 初始化显示
        self.refresh_order_list()
        
        # 绑定鼠标滚轮事件
        def on_mousewheel(event):
            try:
                if canvas.winfo_exists():
                    canvas.yview_scroll(int(-1*(event.delta/120)), "units")
            except tk.TclError:
                pass  # Widget已被销毁，忽略错误
        
        canvas.bind("<MouseWheel>", on_mousewheel)
        self.orders_container.bind("<MouseWheel>", on_mousewheel)

    def export_orders(self):
        """导出订单数据"""
        try:
            from tkinter import filedialog
            import datetime
            
            # 创建导出选择对话框
            dialog = tk.Toplevel(self.parent_frame)
            dialog.title("导出订单数据")
            dialog.geometry("400x300")
            dialog.configure(bg=self.colors['background'])
            dialog.transient(self.parent_frame)
            dialog.grab_set()
            
            # 居中显示
            dialog.update_idletasks()
            x = (dialog.winfo_screenwidth() // 2) - (200)
            y = (dialog.winfo_screenheight() // 2) - (150)
            dialog.geometry(f"400x300+{x}+{y}")
            
            # 标题
            tk.Label(dialog, text="导出订单数据", font=('Microsoft YaHei UI', 14, 'bold'),
                    bg=self.colors['background'], fg=self.colors['text']).pack(pady=15)
            
            # 导出选项框架
            options_frame = tk.Frame(dialog, bg=self.colors['background'])
            options_frame.pack(fill="both", expand=True, padx=20, pady=10)
            
            # 导出格式选择
            tk.Label(options_frame, text="选择导出格式:", font=('Microsoft YaHei UI', 12),
                    bg=self.colors['background'], fg=self.colors['text']).pack(anchor="w", pady=(0, 10))
            
            format_var = tk.StringVar(dialog, value="Excel")
            format_options = ["Excel", "CSV", "PDF"]
            
            format_frame = tk.Frame(options_frame, bg=self.colors['background'])
            format_frame.pack(anchor="w")
            
            for i, fmt in enumerate(format_options):
                rb = tk.Radiobutton(format_frame, text=fmt, variable=format_var, value=fmt,
                                  font=('Microsoft YaHei UI', 10), bg=self.colors['background'], 
                                  fg=self.colors['text'], selectcolor=self.colors['surface'])
                rb.grid(row=0, column=i, sticky="w", padx=(0, 20))
            
            # 状态筛选
            tk.Label(options_frame, text="状态筛选:", font=('Microsoft YaHei UI', 12),
                    bg=self.colors['background'], fg=self.colors['text']).pack(anchor="w", pady=(20, 10))
            
            status_var = tk.StringVar(dialog, value="全部")
            status_options = ["全部", "待接单", "已接单", "制作中", "配送中", "已完成", "已取消"]
            
            status_combo = ttk.Combobox(options_frame, textvariable=status_var, 
                                      values=status_options, state="readonly", width=20)
            status_combo.pack(anchor="w")
            
            # 按钮框架
            btn_frame = tk.Frame(dialog, bg=self.colors['background'])
            btn_frame.pack(fill="x", padx=20, pady=20)
            
            def do_export():
                try:
                    file_format = format_var.get()
                    status_filter = status_var.get()
                    
                    # 获取当前时间戳
                    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                    filename = f"订单数据_{status_filter}_{timestamp}"
                    
                    # 选择保存路径
                    if file_format == "Excel":
                        file_path = filedialog.asksaveasfilename(
                            defaultextension=".xlsx",
                            filetypes=[("Excel文件", "*.xlsx")],
                            initialname=filename
                        )
                        if file_path:
                            success = self.export_orders_to_excel(file_path, status_filter)
                    elif file_format == "CSV":
                        file_path = filedialog.asksaveasfilename(
                            defaultextension=".csv",
                            filetypes=[("CSV文件", "*.csv")],
                            initialname=filename
                        )
                        if file_path:
                            success = self.export_orders_to_csv(file_path, status_filter)
                    elif file_format == "PDF":
                        file_path = filedialog.asksaveasfilename(
                            defaultextension=".pdf",
                            filetypes=[("PDF文件", "*.pdf")],
                            initialname=filename
                        )
                        if file_path:
                            success = self.export_orders_to_pdf(file_path, status_filter)
                    
                    if success:
                        messagebox.showinfo("导出成功", f"订单数据已成功导出为 {file_format} 格式", parent=dialog)
                        dialog.destroy()
                    else:
                        messagebox.showerror("导出失败", "导出过程中发生错误", parent=dialog)
                        
                except Exception as e:
                    messagebox.showerror("错误", f"导出失败：{e}", parent=dialog)
            
            tk.Button(btn_frame, text="📊 开始导出", command=do_export,
                     bg=self.colors['primary'], fg='white', bd=0, pady=8, padx=20,
                     font=('Microsoft YaHei UI', 10)).pack(side="left")
            tk.Button(btn_frame, text="取消", command=dialog.destroy,
                     bg=self.colors['text_light'], fg='white', bd=0, pady=8, padx=20,
                     font=('Microsoft YaHei UI', 10)).pack(side="right")
                     
        except Exception as e:
            messagebox.showerror("错误", f"打开导出对话框失败：{e}")
    
    def export_orders_to_excel(self, file_path: str, status_filter: str) -> bool:
        """导出订单为Excel格式"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "订单数据"
            
            # 设置标题
            title = f"智慧餐饮管理系统 - 订单数据 ({status_filter})"
            ws['A1'] = title
            ws['A1'].font = Font(size=16, bold=True)
            ws.merge_cells('A1:H1')
            
            # 设置表头样式
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # 表头
            headers = ["订单号", "客户姓名", "联系电话", "配送地址", "菜品", "总金额", "订单状态", "下单时间"]
            ws.append(headers)
            
            # 设置表头样式
            for cell in ws[2]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # 获取订单数据
            orders = self.get_filtered_orders(status_filter)
            
            # 添加数据
            for order in orders:
                # 处理菜品信息
                meals_text = ""
                for meal in order.get('meals', []):
                    meals_text += f"{meal.get('name', '')}x{meal.get('quantity', 1)} "
                
                row = [
                    f"#{order.get('id', '')}",
                    order.get('customer', ''),
                    order.get('phone', ''),
                    order.get('address', ''),
                    meals_text.strip(),
                    f"￥{order.get('total', 0):.2f}",
                    order.get('status', ''),
                    order.get('create_time', '')
                ]
                ws.append(row)
            
            # 调整列宽
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(file_path)
            return True
            
        except ImportError:
            messagebox.showerror("错误", "请安装openpyxl库：pip install openpyxl")
            return False
        except Exception as e:
            print(f"导出Excel失败: {e}")
            return False
    
    def export_orders_to_csv(self, file_path: str, status_filter: str) -> bool:
        """导出订单为CSV格式"""
        try:
            import csv
            
            with open(file_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
                fieldnames = ["订单号", "客户姓名", "联系电话", "配送地址", "菜品", "总金额", "订单状态", "下单时间"]
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                writer.writeheader()
                
                # 获取订单数据
                orders = self.get_filtered_orders(status_filter)
                
                for order in orders:
                    # 处理菜品信息
                    meals_text = ""
                    for meal in order.get('meals', []):
                        meals_text += f"{meal.get('name', '')}x{meal.get('quantity', 1)} "
                    
                    writer.writerow({
                        "订单号": f"#{order.get('id', '')}",
                        "客户姓名": order.get('customer', ''),
                        "联系电话": order.get('phone', ''),
                        "配送地址": order.get('address', ''),
                        "菜品": meals_text.strip(),
                        "总金额": f"￥{order.get('total', 0):.2f}",
                        "订单状态": order.get('status', ''),
                        "下单时间": order.get('create_time', '')
                    })
            
            return True
            
        except Exception as e:
            print(f"导出CSV失败: {e}")
            return False
    
    def export_orders_to_pdf(self, file_path: str, status_filter: str) -> bool:
        """导出订单为PDF格式"""
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            
            doc = SimpleDocTemplate(file_path, pagesize=A4)
            story = []
            
            # 标题样式
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                spaceAfter=30,
                alignment=1  # 居中
            )
            
            # 添加标题
            title = Paragraph(f"智慧餐饮管理系统 - 订单数据 ({status_filter})", title_style)
            story.append(title)
            story.append(Spacer(1, 20))
            
            # 获取订单数据
            orders = self.get_filtered_orders(status_filter)
            
            # 创建表格数据
            table_data = [["订单号", "客户姓名", "联系电话", "配送地址", "菜品", "总金额", "订单状态", "下单时间"]]
            
            for order in orders:
                # 处理菜品信息
                meals_text = ""
                for meal in order.get('meals', []):
                    meals_text += f"{meal.get('name', '')}x{meal.get('quantity', 1)} "
                
                row = [
                    f"#{order.get('id', '')}",
                    order.get('customer', ''),
                    order.get('phone', ''),
                    order.get('address', ''),
                    meals_text.strip(),
                    f"￥{order.get('total', 0):.2f}",
                    order.get('status', ''),
                    order.get('create_time', '')
                ]
                table_data.append(row)
            
            # 创建表格
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.beige, colors.white])
            ]))
            story.append(table)
            
            doc.build(story)
            return True
            
        except ImportError:
            messagebox.showerror("错误", "请安装reportlab库：pip install reportlab")
            return False
        except Exception as e:
            print(f"导出PDF失败: {e}")
            return False
    
    def get_filtered_orders(self, status_filter: str) -> List[Dict]:
        """获取筛选后的订单数据"""
        if status_filter == "全部":
            return self.order_data
        else:
            return [order for order in self.order_data if order.get('status') == status_filter]
