#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
现代化财务管理模块
"""

import tkinter as tk
from tkinter import ttk, messagebox
from typing import Dict, List, Any, Optional
import datetime
import json
import os

# 导入数据管理中心
try:
    from .data_manager import data_manager
except ImportError:
    try:
        from data_manager import data_manager
    except ImportError:
        # 模拟数据管理器
        class MockDataManager:
            def get_finance_records(self):
                return []
            def register_module(self, module_type, instance):
                pass
        data_manager = MockDataManager()

class ModernFinanceModule:
    def __init__(self, parent_frame, title_frame, order_module=None, employee_module=None):
        self.parent_frame = parent_frame
        self.title_frame = title_frame
        self.order_module = order_module
        self.employee_module = employee_module
        
        # 注册到数据管理中心
        data_manager.register_module('finance', self)
        
        # 固定成本数据文件路径
        import os
        self.fixed_costs_file = os.path.join(
            os.path.dirname(os.path.dirname(__file__)), 
            'data', 
            'fixed_costs.json'
        )
        
        # 现代化配色方案
        self.colors = {
            'primary': '#FF6B35',
            'secondary': '#F7931E', 
            'success': '#00B894',
            'warning': '#FDCB6E',
            'danger': '#E74C3C',
            'info': '#3498DB',
            'background': '#F8F9FA',
            'surface': '#FFFFFF',
            'text_primary': '#2D3436',
            'text_secondary': '#636E72',
            'border': '#E1E8ED',
            'white': '#FFFFFF'
        }
        
        # 字体配置
        self.fonts = {
            'title': ('Microsoft YaHei UI', 16, 'bold'),
            'heading': ('Microsoft YaHei UI', 14, 'bold'),
            'body': ('Microsoft YaHei UI', 12),
            'small': ('Microsoft YaHei UI', 10)
        }
        
        self.main_frame = None
          # 初始化Tkinter变量（指定父窗口）
        self.search_var = None
        self.date_filter_var = None
        self.type_filter_var = None
        
    def show(self):
        """显示财务管理界面"""
        if self.main_frame:
            self.main_frame.destroy()
        
        self.main_frame = tk.Frame(self.parent_frame, bg=self.colors['background'])
        self.main_frame.pack(fill="both", expand=True)
        
        # 初始化Tkinter变量（指定父窗口）
        self.search_var = tk.StringVar(self.main_frame)
        self.date_filter_var = tk.StringVar(self.main_frame, value="全部")
        self.type_filter_var = tk.StringVar(self.main_frame, value="全部")
        
        # 标题
        title_label = tk.Label(self.main_frame, text="💼 财务管理", 
                              font=self.fonts['title'],
                              bg=self.colors['background'], 
                              fg=self.colors['text_primary'])
        title_label.pack(pady=(0, 20))
          # 财务概览
        self.create_finance_overview()
        
        # 创建选项卡
        self.create_finance_tabs()
        
    def create_finance_overview(self):
        """创建财务概览"""
        overview_frame = tk.Frame(self.main_frame, bg=self.colors['background'])
        overview_frame.pack(fill="x", pady=(0, 20))
        
        # 财务统计
        stats = [
            {"title": "今日收入", "value": "￥2,580", "icon": "💰", "color": self.colors['success']},
            {"title": "今日支出", "value": "￥680", "icon": "💸", "color": self.colors['danger']},
            {"title": "净利润", "value": "￥1,900", "icon": "📈", "color": self.colors['primary']},
            {"title": "本月收入", "value": "￥58,960", "icon": "💳", "color": self.colors['info']}
        ]
        
        for stat in stats:
            card = tk.Frame(overview_frame, bg=self.colors['surface'], relief="flat", bd=1)
            card.pack(side="left", fill="both", expand=True, padx=(0, 10))
            
            # 图标
            icon_label = tk.Label(card, text=stat['icon'], font=('Segoe UI Emoji', 24),
                                bg=self.colors['surface'], fg=stat['color'])
            icon_label.pack(pady=(10, 5))
            
            # 数值
            value_label = tk.Label(card, text=stat['value'], font=self.fonts['heading'],
                                 bg=self.colors['surface'], fg=self.colors['text_primary'])
            value_label.pack()
            
            # 标题
            title_label = tk.Label(card, text=stat['title'], font=self.fonts['body'],
                                 bg=self.colors['surface'], fg=self.colors['text_secondary'])
            title_label.pack(pady=(5, 10))
            
    def create_finance_tabs(self):
        """创建财务管理选项卡"""
        tabs_frame = tk.Frame(self.main_frame, bg=self.colors['background'])
        tabs_frame.pack(fill="both", expand=True)
        
        # 创建选项卡控件
        self.notebook = ttk.Notebook(tabs_frame)
        self.notebook.pack(fill="both", expand=True, padx=10, pady=10)
        
        # 收支记录选项卡
        self.records_frame = tk.Frame(self.notebook, bg=self.colors['surface'])
        self.notebook.add(self.records_frame, text="📊 收支记录")
        self.create_finance_records()
        
        # 固定成本管理选项卡
        self.fixed_costs_frame = tk.Frame(self.notebook, bg=self.colors['surface'])
        self.notebook.add(self.fixed_costs_frame, text="🏢 固定成本")
        self.create_fixed_costs_management()
        
    def create_fixed_costs_management(self):
        """创建固定成本管理界面"""
        # 标题
        title_label = tk.Label(self.fixed_costs_frame, text="🏢 固定成本管理", 
                              font=self.fonts['heading'],
                              bg=self.colors['surface'], 
                              fg=self.colors['text_primary'])
        title_label.pack(pady=(10, 20))
          # 成本概览卡片
        overview_frame = tk.Frame(self.fixed_costs_frame, bg=self.colors['surface'])
        overview_frame.pack(fill="x", padx=20, pady=(0, 20))
        
        # 计算固定成本统计
        cost_stats = self.calculate_fixed_cost_stats()
        
        for stat in cost_stats:
            card = tk.Frame(overview_frame, bg=self.colors['background'], relief="solid", bd=1)
            card.pack(side="left", fill="both", expand=True, padx=(0, 10))
            
            # 图标
            icon_label = tk.Label(card, text=stat['icon'], font=('Segoe UI Emoji', 20),
                                bg=self.colors['background'], fg=stat['color'])
            icon_label.pack(pady=(10, 5))
            
            # 数值
            value_label = tk.Label(card, text=stat['value'], font=self.fonts['heading'],
                                 bg=self.colors['background'], fg=self.colors['text_primary'])
            value_label.pack()
            
            # 标题
            title_label = tk.Label(card, text=stat['title'], font=self.fonts['body'],
                                 bg=self.colors['background'], fg=self.colors['text_secondary'])
            title_label.pack(pady=(5, 10))
        
        # 固定成本列表
        list_frame = tk.Frame(self.fixed_costs_frame, bg=self.colors['surface'])
        list_frame.pack(fill="both", expand=True, padx=20, pady=(0, 20))
        
        # 操作按钮
        btn_frame = tk.Frame(list_frame, bg=self.colors['surface'])
        btn_frame.pack(fill="x", pady=(0, 10))
        
        add_cost_btn = tk.Button(btn_frame, text="➕ 添加固定成本", 
                               font=self.fonts['body'],
                               bg=self.colors['primary'], fg='white',
                               bd=0, pady=8, padx=15, cursor="hand2",
                               command=self.add_fixed_cost)
        add_cost_btn.pack(side="left", padx=(0, 10))
        
        edit_cost_btn = tk.Button(btn_frame, text="✏️ 编辑成本", 
                                font=self.fonts['body'],
                                bg=self.colors['info'], fg='white',
                                bd=0, pady=8, padx=15, cursor="hand2",
                                command=self.edit_fixed_cost)
        edit_cost_btn.pack(side="left", padx=(0, 10))
        
        delete_cost_btn = tk.Button(btn_frame, text="🗑️ 删除成本", 
                                  font=self.fonts['body'],
                                  bg=self.colors['danger'], fg='white',
                                  bd=0, pady=8, padx=15, cursor="hand2",
                                  command=self.delete_fixed_cost)
        delete_cost_btn.pack(side="left")
          # 固定成本表格
        columns = ("成本类型", "成本项目", "金额", "周期", "下次缴费", "状态", "备注")
        self.costs_tree = ttk.Treeview(list_frame, columns=columns, show='headings', height=12)
        
        # 设置列标题和宽度
        column_widths = [100, 150, 100, 80, 120, 80, 150]
        for i, col in enumerate(columns):
            self.costs_tree.heading(col, text=col)
            self.costs_tree.column(col, width=column_widths[i], anchor="center")
        
        # 从数据文件加载固定成本数据
        self.load_and_display_fixed_costs()
        
        # 设置标签样式
        self.costs_tree.tag_configure("unpaid", background="#FFE6E6", foreground="#D63031")
        self.costs_tree.tag_configure("paid", background="#E8F5E8", foreground="#00B894")
        
        # 添加滚动条
        costs_scrollbar = ttk.Scrollbar(list_frame, orient="vertical", command=self.costs_tree.yview)
        self.costs_tree.configure(yscrollcommand=costs_scrollbar.set)
        
        # 布局
        self.costs_tree.pack(side="left", fill="both", expand=True)
        costs_scrollbar.pack(side="right", fill="y")
    
    def load_and_display_fixed_costs(self):
        """加载并显示固定成本数据"""
        try:
            # 清空现有数据
            for item in self.costs_tree.get_children():
                self.costs_tree.delete(item)
            
            # 加载数据
            costs_data = self.load_fixed_costs()
            
            for cost in costs_data:
                # 根据状态设置不同颜色
                if cost.get('status') == "未付":
                    tags = ("unpaid",)
                else:
                    tags = ("paid",)
                
                # 插入数据到表格
                self.costs_tree.insert("", "end", values=(
                    cost.get('cost_type', ''),
                    cost.get('item', ''),
                    f"￥{cost.get('amount', 0):,.0f}",
                    cost.get('period', ''),
                    cost.get('next_date', ''),
                    cost.get('status', ''),
                    cost.get('note', '')
                ), tags=tags)
                
        except Exception as e:
            print(f"加载固定成本数据失败: {e}")            # 如果加载失败，显示默认数据
            self.display_default_costs()
    
    def display_default_costs(self):
        """显示默认的固定成本数据"""
        sample_costs = [
            ("租金", "店铺租金", "￥8,000", "月付", "2024-07-01", "已付", "主店面租金"),
            ("人力", "厨师工资", "￥5,000", "月付", "2024-07-01", "已付", "主厨月薪"),
            ("水电", "电费", "￥800", "月付", "2024-07-05", "未付", "店铺用电"),
        ]
        
        for cost in sample_costs:
            if cost[5] == "未付":
                tags = ("unpaid",)
            else:
                tags = ("paid",)
            
            self.costs_tree.insert("", "end", values=cost, tags=tags)
    
    def add_fixed_cost(self):
        """添加固定成本"""
        try:
            root = self.main_frame.winfo_toplevel()
            
            # 创建添加对话框
            dialog = tk.Toplevel(root)
            dialog.title("添加固定成本")
            dialog.geometry("450x600")  # 增加高度从500到600
            dialog.configure(bg=self.colors['background'])
            dialog.transient(root)
            dialog.grab_set()
            
            # 居中显示
            dialog.update_idletasks()
            x = (dialog.winfo_screenwidth() // 2) - (225)
            y = (dialog.winfo_screenheight() // 2) - (250)
            dialog.geometry(f"450x500+{x}+{y}")
            
            # 标题
            tk.Label(dialog, text="添加固定成本", font=self.fonts['heading'],
                    bg=self.colors['background'], fg=self.colors['text_primary']).pack(pady=15)
            
            # 输入框架
            form_frame = tk.Frame(dialog, bg=self.colors['background'])
            form_frame.pack(fill="both", expand=True, padx=20, pady=10)
            
            # 成本类型
            tk.Label(form_frame, text="成本类型:", bg=self.colors['background']).pack(anchor="w")
            type_var = tk.StringVar(dialog, value="租金")
            type_combo = ttk.Combobox(form_frame, textvariable=type_var,
                                    values=["租金", "人力", "水电", "通讯", "保险", "许可", "设备", "其他"])
            type_combo.pack(fill="x", pady=(5, 15))
            
            # 成本项目
            tk.Label(form_frame, text="成本项目:", bg=self.colors['background']).pack(anchor="w")
            item_var = tk.StringVar(dialog)
            item_entry = tk.Entry(form_frame, textvariable=item_var, font=self.fonts['body'])
            item_entry.pack(fill="x", pady=(5, 15))
            
            # 金额
            tk.Label(form_frame, text="金额:", bg=self.colors['background']).pack(anchor="w")
            amount_var = tk.StringVar(dialog)
            amount_entry = tk.Entry(form_frame, textvariable=amount_var, font=self.fonts['body'])
            amount_entry.pack(fill="x", pady=(5, 15))
            
            # 缴费周期
            tk.Label(form_frame, text="缴费周期:", bg=self.colors['background']).pack(anchor="w")
            period_var = tk.StringVar(dialog, value="月付")
            period_combo = ttk.Combobox(form_frame, textvariable=period_var,
                                      values=["日付", "周付", "月付", "季付", "年付", "一次性"])
            period_combo.pack(fill="x", pady=(5, 15))
            
            # 下次缴费日期
            tk.Label(form_frame, text="下次缴费日期:", bg=self.colors['background']).pack(anchor="w")
            next_date_var = tk.StringVar(dialog)
            next_date_entry = tk.Entry(form_frame, textvariable=next_date_var, font=self.fonts['body'])
            next_date_entry.pack(fill="x", pady=(5, 5))
            tk.Label(form_frame, text="格式: YYYY-MM-DD", font=self.fonts['small'],
                    bg=self.colors['background'], fg=self.colors['text_secondary']).pack(anchor="w", pady=(0, 15))
            
            # 状态
            tk.Label(form_frame, text="状态:", bg=self.colors['background']).pack(anchor="w")
            status_var = tk.StringVar(dialog, value="未付")
            status_combo = ttk.Combobox(form_frame, textvariable=status_var,
                                      values=["已付", "未付", "逾期"])
            status_combo.pack(fill="x", pady=(5, 15))
            
            # 备注
            tk.Label(form_frame, text="备注:", bg=self.colors['background']).pack(anchor="w")
            note_var = tk.StringVar(dialog)
            note_entry = tk.Entry(form_frame, textvariable=note_var, font=self.fonts['body'])
            note_entry.pack(fill="x", pady=(5, 15))
            
            # 按钮
            btn_frame = tk.Frame(dialog, bg=self.colors['background'])
            btn_frame.pack(fill="x", padx=20, pady=20)
            
            def save_cost():
                try:
                    cost_type = type_var.get().strip()
                    item = item_var.get().strip()
                    amount_str = amount_var.get().strip()
                    period = period_var.get().strip()
                    next_date = next_date_var.get().strip()
                    status = status_var.get().strip()
                    note = note_var.get().strip()
                    
                    if not all([cost_type, item, amount_str, period]):
                        messagebox.showerror("错误", "请填写所有必填字段", parent=dialog)
                        return
                    
                    try:
                        amount = float(amount_str)
                        if amount <= 0:
                            raise ValueError
                    except ValueError:
                        messagebox.showerror("错误", "请输入有效的金额", parent=dialog)
                        return
                    
                    # 验证日期格式
                    if next_date:
                        try:
                            datetime.datetime.strptime(next_date, "%Y-%m-%d")
                        except ValueError:
                            messagebox.showerror("错误", "日期格式不正确，请使用 YYYY-MM-DD", parent=dialog)
                            return
                    
                    # 添加到表格
                    tags = ("unpaid",) if status == "未付" else ("paid",)
                    self.costs_tree.insert("", "end", values=(
                        cost_type, item, f"￥{amount:,.0f}", period, next_date, status, note
                    ), tags=tags)
                    
                    # 保存固定成本数据
                    self.save_fixed_costs(self.get_costs_from_tree())
                    
                    messagebox.showinfo("成功", "固定成本添加成功", parent=dialog)
                    dialog.destroy()
                    
                except Exception as e:
                    messagebox.showerror("错误", f"添加失败：{e}", parent=dialog)
            
            tk.Button(btn_frame, text="保存", command=save_cost,
                     bg=self.colors['primary'], fg='white', bd=0, pady=8, padx=20).pack(side="left")
            tk.Button(btn_frame, text="取消", command=dialog.destroy,
                     bg=self.colors['text_secondary'], fg='white', bd=0, pady=8, padx=20).pack(side="right")                     
        except Exception as e:
            root = self.main_frame.winfo_toplevel()
            messagebox.showerror("错误", f"打开添加对话框失败：{e}", parent=root)
    
    def edit_fixed_cost(self):
        """编辑固定成本"""
        try:
            selected = self.costs_tree.selection()
            if not selected:
                messagebox.showwarning("提示", "请先选择要编辑的成本项目")
                return
            
            item = self.costs_tree.item(selected[0])
            values = item['values']
            
            root = self.main_frame.winfo_toplevel()
            
            # 创建编辑对话框
            dialog = tk.Toplevel(root)
            dialog.title("编辑固定成本")
            dialog.geometry("450x600")
            dialog.configure(bg=self.colors['background'])
            dialog.transient(root)
            dialog.grab_set()
            
            # 居中显示
            dialog.update_idletasks()
            x = (dialog.winfo_screenwidth() // 2) - (225)
            y = (dialog.winfo_screenheight() // 2) - (300)
            dialog.geometry(f"450x600+{x}+{y}")
            
            # 标题
            tk.Label(dialog, text="编辑固定成本", font=self.fonts['heading'],
                    bg=self.colors['background'], fg=self.colors['text_primary']).pack(pady=15)
            
            # 输入框架
            form_frame = tk.Frame(dialog, bg=self.colors['background'])
            form_frame.pack(fill="both", expand=True, padx=20, pady=10)
            
            # 成本类型
            tk.Label(form_frame, text="成本类型:", bg=self.colors['background']).pack(anchor="w")
            type_var = tk.StringVar(dialog, value=values[0])
            type_combo = ttk.Combobox(form_frame, textvariable=type_var,
                                    values=["租金", "人力", "水电", "通讯", "保险", "许可", "设备", "其他"])
            type_combo.pack(fill="x", pady=(5, 15))
            
            # 成本项目
            tk.Label(form_frame, text="成本项目:", bg=self.colors['background']).pack(anchor="w")
            item_var = tk.StringVar(dialog, value=values[1])
            item_entry = tk.Entry(form_frame, textvariable=item_var, font=self.fonts['body'])
            item_entry.pack(fill="x", pady=(5, 15))
            
            # 金额 - 去掉￥符号和逗号
            amount_value = str(values[2]).replace("￥", "").replace(",", "")
            tk.Label(form_frame, text="金额:", bg=self.colors['background']).pack(anchor="w")
            amount_var = tk.StringVar(dialog, value=amount_value)
            amount_entry = tk.Entry(form_frame, textvariable=amount_var, font=self.fonts['body'])
            amount_entry.pack(fill="x", pady=(5, 15))
            
            # 缴费周期
            tk.Label(form_frame, text="缴费周期:", bg=self.colors['background']).pack(anchor="w")
            period_var = tk.StringVar(dialog, value=values[3])
            period_combo = ttk.Combobox(form_frame, textvariable=period_var,
                                      values=["日付", "周付", "月付", "季付", "年付", "一次性"])
            period_combo.pack(fill="x", pady=(5, 15))
            
            # 下次缴费日期
            tk.Label(form_frame, text="下次缴费日期:", bg=self.colors['background']).pack(anchor="w")
            next_date_var = tk.StringVar(dialog, value=values[4])
            next_date_entry = tk.Entry(form_frame, textvariable=next_date_var, font=self.fonts['body'])
            next_date_entry.pack(fill="x", pady=(5, 5))
            tk.Label(form_frame, text="格式: YYYY-MM-DD", font=self.fonts['small'],
                    bg=self.colors['background'], fg=self.colors['text_secondary']).pack(anchor="w", pady=(0, 15))
            
            # 状态
            tk.Label(form_frame, text="状态:", bg=self.colors['background']).pack(anchor="w")
            status_var = tk.StringVar(dialog, value=values[5])
            status_combo = ttk.Combobox(form_frame, textvariable=status_var,
                                      values=["已付", "未付", "逾期"])
            status_combo.pack(fill="x", pady=(5, 15))
            
            # 备注
            tk.Label(form_frame, text="备注:", bg=self.colors['background']).pack(anchor="w")
            note_var = tk.StringVar(dialog, value=values[6] if len(values) > 6 else "")
            note_entry = tk.Entry(form_frame, textvariable=note_var, font=self.fonts['body'])
            note_entry.pack(fill="x", pady=(5, 15))
            
            # 按钮
            btn_frame = tk.Frame(dialog, bg=self.colors['background'])
            btn_frame.pack(fill="x", padx=20, pady=20)
            
            def update_cost():
                try:
                    cost_type = type_var.get().strip()
                    item_name = item_var.get().strip()
                    amount_str = amount_var.get().strip()
                    period = period_var.get().strip()
                    next_date = next_date_var.get().strip()
                    status = status_var.get().strip()
                    note = note_var.get().strip()
                    
                    if not all([cost_type, item_name, amount_str, period]):
                        messagebox.showerror("错误", "请填写所有必填字段", parent=dialog)
                        return
                    
                    try:
                        amount = float(amount_str)
                        if amount <= 0:
                            raise ValueError
                    except ValueError:
                        messagebox.showerror("错误", "请输入有效的金额", parent=dialog)
                        return
                    
                    # 验证日期格式
                    if next_date:
                        try:
                            datetime.datetime.strptime(next_date, "%Y-%m-%d")
                        except ValueError:
                            messagebox.showerror("错误", "日期格式不正确，请使用 YYYY-MM-DD", parent=dialog)
                            return
                    
                    # 更新表格中的选中项
                    tags = ("unpaid",) if status == "未付" else ("paid",)
                    self.costs_tree.item(selected[0], values=(
                        cost_type, item_name, f"￥{amount:,.0f}", period, next_date, status, note
                    ), tags=tags)
                    
                    # 保存固定成本数据
                    self.save_fixed_costs(self.get_costs_from_tree())
                    
                    messagebox.showinfo("成功", "固定成本更新成功", parent=dialog)
                    dialog.destroy()
                    
                except Exception as e:
                    messagebox.showerror("错误", f"更新失败：{e}", parent=dialog)
            
            tk.Button(btn_frame, text="更新", command=update_cost,
                     bg=self.colors['primary'], fg='white', bd=0, pady=8, padx=20).pack(side="left")
            tk.Button(btn_frame, text="取消", command=dialog.destroy,
                     bg=self.colors['text_secondary'], fg='white', bd=0, pady=8, padx=20).pack(side="right")
                     
        except Exception as e:
            root = self.main_frame.winfo_toplevel()
            messagebox.showerror("错误", f"编辑失败：{e}", parent=root)
    
    def delete_fixed_cost(self):
        """删除固定成本"""
        try:
            selected = self.costs_tree.selection()
            if not selected:
                messagebox.showwarning("提示", "请先选择要删除的成本项目")
                return
            
            item = self.costs_tree.item(selected[0])
            values = item['values']
            
            result = messagebox.askyesno("确认删除", f"确定要删除成本项目 '{values[1]}' 吗？")
            if result:
                self.costs_tree.delete(selected[0])
                
                # 保存固定成本数据
                self.save_fixed_costs(self.get_costs_from_tree())
                
                messagebox.showinfo("成功", "成本项目删除成功")
        except Exception as e:
            root = self.main_frame.winfo_toplevel()
            messagebox.showerror("错误", f"删除失败：{e}", parent=root)
            
    def create_finance_records(self):
        """创建收支记录表格"""
        # 表格标题
        records_title = tk.Label(self.records_frame, text="📊 收支记录", 
                                font=self.fonts['heading'],
                                bg=self.colors['surface'], 
                                fg=self.colors['text_primary'])
        records_title.pack(pady=10)
        
        # 创建Treeview表格
        columns = ("时间", "类型", "描述", "金额", "支付方式", "备注")
        self.finance_tree = ttk.Treeview(self.records_frame, columns=columns, show='headings', height=15)
        
        # 设置列标题
        for col in columns:
            self.finance_tree.heading(col, text=col)
            self.finance_tree.column(col, width=120, anchor="center")
        
        # 从数据库加载真实财务数据
        self.load_finance_records()
        
        # 添加滚动条
        scrollbar = ttk.Scrollbar(self.records_frame, orient="vertical", command=self.finance_tree.yview)
        self.finance_tree.configure(yscrollcommand=scrollbar.set)
        
        # 布局
        self.finance_tree.pack(side="left", fill="both", expand=True, padx=10, pady=10)
        scrollbar.pack(side="right", fill="y", pady=10)
        
        # 添加操作按钮
        self.create_finance_buttons(self.records_frame)
    
    def load_finance_records(self):
        """从数据库加载财务记录"""
        try:
            # 清空现有数据
            for item in self.finance_tree.get_children():
                self.finance_tree.delete(item)
            
            # 从数据管理器获取财务记录
            finance_records = data_manager.get_financial_records()
            
            if not finance_records:
                # 如果没有数据，显示提示
                self.finance_tree.insert("", "end", values=("暂无数据", "", "", "", "", ""))
                return
            
            # 插入真实数据
            for record in finance_records:
                # 处理时间格式
                time_str = ""
                if record.get('income_date'):
                    time_str = str(record['income_date'])
                elif record.get('created_at'):
                    time_str = str(record['created_at'])[:10]  # 取日期部分
                
                # 处理金额
                amount = record.get('amount', 0)
                amount_str = f"￥{amount:.2f}" if amount else "￥0.00"
                
                # 处理类型
                record_type = "收入" if record.get('income_type') == 'revenue' else "支出"
                
                # 处理描述
                description = record.get('description', '')
                
                # 处理支付方式（从订单关联获取）
                payment_method = "现金"  # 默认值
                if record.get('order_price_id'):
                    # 可以进一步查询订单表获取支付方式
                    payment_method = "订单收入"
                
                # 处理备注
                note = record.get('description', '')[:20] + "..." if len(record.get('description', '')) > 20 else record.get('description', '')
                
                self.finance_tree.insert("", "end", values=(
                    time_str, record_type, description, amount_str, payment_method, note
                ))
                
        except Exception as e:
            print(f"加载财务记录失败: {e}")
            # 如果加载失败，显示默认数据
            self.finance_tree.insert("", "end", values=("加载失败", "", "", "", "", ""))
    
    def refresh_finance_records(self):
        """刷新财务记录"""
        if hasattr(self, 'finance_tree'):
            self.load_finance_records()
    
    def create_finance_buttons(self, parent):
        """创建财务操作按钮"""
        button_frame = tk.Frame(parent, bg=self.colors['surface'])
        button_frame.pack(fill="x", padx=10, pady=(0, 10))
        
        # 添加收入按钮
        add_income_btn = tk.Button(button_frame, text="📈 添加收入", 
                                  font=self.fonts['body'],
                                  bg=self.colors['success'], fg='white',
                                  bd=0, pady=8, padx=15, cursor="hand2",
                                  command=self.add_income_record)
        add_income_btn.pack(side="left", padx=(0, 10))
        
        # 添加支出按钮
        add_expense_btn = tk.Button(button_frame, text="📉 添加支出", 
                                   font=self.fonts['body'],
                                   bg=self.colors['danger'], fg='white',
                                   bd=0, pady=8, padx=15, cursor="hand2",
                                   command=self.add_expense_record)
        add_expense_btn.pack(side="left", padx=(0, 10))
        
        # 导出报表按钮
        export_btn = tk.Button(button_frame, text="📊 导出报表", 
                              font=self.fonts['body'],
                              bg=self.colors['info'], fg='white',
                              bd=0, pady=8, padx=15, cursor="hand2",
                              command=self.export_finance_report)
        export_btn.pack(side="right")
        
    def add_income_record(self):
        """添加收入记录"""
        try:
            # 获取根窗口以避免Tkinter错误
            root = self.main_frame.winfo_toplevel()
            
            # 创建输入对话框
            dialog = tk.Toplevel(root)
            dialog.title("添加收入记录")
            dialog.geometry("400x400")  # 增加高度从300到400
            dialog.configure(bg=self.colors['background'])
            dialog.transient(root)
            dialog.grab_set()
            
            # 居中显示
            dialog.update_idletasks()
            x = (dialog.winfo_screenwidth() // 2) - (200)
            y = (dialog.winfo_screenheight() // 2) - (150)
            dialog.geometry(f"400x300+{x}+{y}")
            
            # 输入字段
            tk.Label(dialog, text="添加收入记录", font=self.fonts['heading'],
                    bg=self.colors['background'], fg=self.colors['text_primary']).pack(pady=10)
            
            # 描述
            tk.Label(dialog, text="描述:", bg=self.colors['background'], 
                    fg=self.colors['text_primary']).pack(anchor="w", padx=20)
            desc_var = tk.StringVar(dialog)
            desc_entry = tk.Entry(dialog, textvariable=desc_var, font=self.fonts['body'])
            desc_entry.pack(fill="x", padx=20, pady=5)
            
            # 金额
            tk.Label(dialog, text="金额:", bg=self.colors['background'], 
                    fg=self.colors['text_primary']).pack(anchor="w", padx=20)
            amount_var = tk.StringVar(dialog)
            amount_entry = tk.Entry(dialog, textvariable=amount_var, font=self.fonts['body'])
            amount_entry.pack(fill="x", padx=20, pady=5)
            
            # 支付方式
            tk.Label(dialog, text="支付方式:", bg=self.colors['background'], 
                    fg=self.colors['text_primary']).pack(anchor="w", padx=20)
            payment_var = tk.StringVar(dialog, value="现金")
            payment_combo = ttk.Combobox(dialog, textvariable=payment_var, 
                                        values=["现金", "银行卡", "微信支付", "支付宝"])
            payment_combo.pack(fill="x", padx=20, pady=5)
            
            # 备注
            tk.Label(dialog, text="备注:", bg=self.colors['background'], 
                    fg=self.colors['text_primary']).pack(anchor="w", padx=20)
            note_var = tk.StringVar(dialog)
            note_entry = tk.Entry(dialog, textvariable=note_var, font=self.fonts['body'])
            note_entry.pack(fill="x", padx=20, pady=5)
            
            # 按钮
            btn_frame = tk.Frame(dialog, bg=self.colors['background'])
            btn_frame.pack(fill="x", padx=20, pady=20)
            
            def save_income():
                try:
                    desc = desc_var.get().strip()
                    amount_str = amount_var.get().strip()
                    payment = payment_var.get().strip()
                    note = note_var.get().strip()
                    
                    if not desc or not amount_str:
                        messagebox.showerror("错误", "请填写描述和金额", parent=dialog)
                        return
                    
                    amount = float(amount_str)
                    if amount <= 0:
                        messagebox.showerror("错误", "金额必须大于0", parent=dialog)
                        return
                    
                    # 添加到表格
                    current_time = datetime.datetime.now().strftime("%H:%M")
                    self.finance_tree.insert("", 0, values=(
                        current_time, "收入", desc, f"￥{amount:.2f}", payment, note
                    ))
                    
                    messagebox.showinfo("成功", "收入记录添加成功", parent=dialog)
                    dialog.destroy()
                    
                except ValueError:
                    messagebox.showerror("错误", "请输入有效的金额", parent=dialog)
                except Exception as e:
                    messagebox.showerror("错误", f"添加记录失败：{e}", parent=dialog)
            
            tk.Button(btn_frame, text="保存", command=save_income,
                     bg=self.colors['success'], fg='white', bd=0, pady=8, padx=20).pack(side="left")
            tk.Button(btn_frame, text="取消", command=dialog.destroy,
                     bg=self.colors['text_secondary'], fg='white', bd=0, pady=8, padx=20).pack(side="right")                     
        except Exception as e:
            root = self.main_frame.winfo_toplevel()
            messagebox.showerror("错误", f"打开添加收入对话框失败：{e}", parent=root)
            
    def add_expense_record(self):
        """添加支出记录"""
        try:
            # 获取根窗口以避免Tkinter错误
            root = self.main_frame.winfo_toplevel()
            
            # 创建输入对话框
            dialog = tk.Toplevel(root)
            dialog.title("添加支出记录")
            dialog.geometry("400x450")
            dialog.configure(bg=self.colors['background'])
            dialog.transient(root)
            dialog.grab_set()
            
            # 居中显示
            dialog.update_idletasks()
            x = (dialog.winfo_screenwidth() // 2) - (200)
            y = (dialog.winfo_screenheight() // 2) - (225)
            dialog.geometry(f"400x450+{x}+{y}")
            
            # 输入字段
            tk.Label(dialog, text="添加支出记录", font=self.fonts['heading'],
                    bg=self.colors['background'], fg=self.colors['text_primary']).pack(pady=15)
            
            form_frame = tk.Frame(dialog, bg=self.colors['background'])
            form_frame.pack(fill="both", expand=True, padx=20, pady=10)
            
            # 描述
            tk.Label(form_frame, text="描述:", bg=self.colors['background'], 
                    fg=self.colors['text_primary']).pack(anchor="w")
            desc_var = tk.StringVar(dialog)
            desc_entry = tk.Entry(form_frame, textvariable=desc_var, font=self.fonts['body'])
            desc_entry.pack(fill="x", pady=(5, 15))
            
            # 金额
            tk.Label(form_frame, text="金额:", bg=self.colors['background'], 
                    fg=self.colors['text_primary']).pack(anchor="w")
            amount_var = tk.StringVar(dialog)
            amount_entry = tk.Entry(form_frame, textvariable=amount_var, font=self.fonts['body'])
            amount_entry.pack(fill="x", pady=(5, 15))
            
            # 支付方式
            tk.Label(form_frame, text="支付方式:", bg=self.colors['background'], 
                    fg=self.colors['text_primary']).pack(anchor="w")
            payment_var = tk.StringVar(dialog, value="现金")
            payment_combo = ttk.Combobox(form_frame, textvariable=payment_var, 
                                        values=["现金", "银行卡", "微信支付", "支付宝", "银行转账"])
            payment_combo.pack(fill="x", pady=(5, 15))
            
            # 支出类型
            tk.Label(form_frame, text="支出类型:", bg=self.colors['background'], 
                    fg=self.colors['text_primary']).pack(anchor="w")
            expense_type_var = tk.StringVar(dialog, value="原料采购")
            expense_type_combo = ttk.Combobox(form_frame, textvariable=expense_type_var, 
                                           values=["原料采购", "员工工资", "设备维护", "租金水电", "营销费用", "其他"])
            expense_type_combo.pack(fill="x", pady=(5, 15))
            
            # 备注
            tk.Label(form_frame, text="备注:", bg=self.colors['background'], 
                    fg=self.colors['text_primary']).pack(anchor="w")
            note_var = tk.StringVar(dialog)
            note_entry = tk.Entry(form_frame, textvariable=note_var, font=self.fonts['body'])
            note_entry.pack(fill="x", pady=(5, 15))
            
            # 按钮
            btn_frame = tk.Frame(dialog, bg=self.colors['background'])
            btn_frame.pack(fill="x", padx=20, pady=20)
            
            def save_expense():
                try:
                    desc = desc_var.get().strip()
                    amount_str = amount_var.get().strip()
                    payment = payment_var.get().strip()
                    expense_type = expense_type_var.get().strip()
                    note = note_var.get().strip()
                    
                    if not desc or not amount_str:
                        messagebox.showerror("错误", "请填写描述和金额", parent=dialog)
                        return
                    
                    amount = float(amount_str)
                    if amount <= 0:
                        messagebox.showerror("错误", "金额必须大于0", parent=dialog)
                        return
                    
                    # 添加到表格
                    current_time = datetime.datetime.now().strftime("%H:%M")
                    self.finance_tree.insert("", 0, values=(
                        current_time, "支出", f"{expense_type}-{desc}", f"￥{amount:.2f}", payment, note
                    ))                    
                    messagebox.showinfo("成功", "支出记录添加成功", parent=dialog)
                    dialog.destroy()
                    
                except ValueError:
                    messagebox.showerror("错误", "请输入有效的金额", parent=dialog)
                except Exception as e:
                    messagebox.showerror("错误", f"添加记录失败：{e}", parent=dialog)
            
            tk.Button(btn_frame, text="保存", command=save_expense,
                     bg=self.colors['danger'], fg='white', bd=0, pady=8, padx=20).pack(side="left")
            tk.Button(btn_frame, text="取消", command=dialog.destroy,
                     bg=self.colors['text_secondary'], fg='white', bd=0, pady=8, padx=20).pack(side="right")
                     
        except Exception as e:
            root = self.main_frame.winfo_toplevel()
            messagebox.showerror("错误", f"打开添加支出对话框失败：{e}", parent=root)
            
    def export_finance_report(self):
        """导出财务报表"""
        try:
            root = self.main_frame.winfo_toplevel()
            
            # 创建导出选择对话框
            dialog = tk.Toplevel(root)
            dialog.title("导出财务报表")
            dialog.geometry("500x400")
            dialog.configure(bg=self.colors['background'])
            dialog.transient(root)
            dialog.grab_set()
            
            # 居中显示
            dialog.update_idletasks()
            x = (dialog.winfo_screenwidth() // 2) - (250)
            y = (dialog.winfo_screenheight() // 2) - (200)
            dialog.geometry(f"500x400+{x}+{y}")
            
            # 标题
            tk.Label(dialog, text="导出财务报表", font=self.fonts['heading'],
                    bg=self.colors['background'], fg=self.colors['text_primary']).pack(pady=15)
            
            # 导出选项框架
            options_frame = tk.Frame(dialog, bg=self.colors['background'])
            options_frame.pack(fill="both", expand=True, padx=20, pady=10)
            
            # 导出类型选择
            tk.Label(options_frame, text="选择导出内容:", font=self.fonts['body'],
                    bg=self.colors['background'], fg=self.colors['text_primary']).pack(anchor="w", pady=(0, 10))
            
            export_type_var = tk.StringVar(dialog, value="收支记录")
            export_options = ["收支记录", "固定成本", "财务概览", "完整报表"]
            
            for option in export_options:
                rb = tk.Radiobutton(options_frame, text=option, variable=export_type_var, value=option,
                                  font=self.fonts['body'], bg=self.colors['background'], 
                                  fg=self.colors['text_primary'], selectcolor=self.colors['surface'])
                rb.pack(anchor="w", pady=2)
            
            # 导出格式选择
            tk.Label(options_frame, text="选择导出格式:", font=self.fonts['body'],
                    bg=self.colors['background'], fg=self.colors['text_primary']).pack(anchor="w", pady=(20, 10))
            
            format_var = tk.StringVar(dialog, value="Excel")
            format_options = ["Excel", "CSV", "PDF"]
            
            format_frame = tk.Frame(options_frame, bg=self.colors['background'])
            format_frame.pack(anchor="w")
            
            for i, fmt in enumerate(format_options):
                rb = tk.Radiobutton(format_frame, text=fmt, variable=format_var, value=fmt,
                                  font=self.fonts['body'], bg=self.colors['background'], 
                                  fg=self.colors['text_primary'], selectcolor=self.colors['surface'])
                rb.grid(row=0, column=i, sticky="w", padx=(0, 20))
            
            # 时间范围选择
            tk.Label(options_frame, text="选择时间范围:", font=self.fonts['body'],
                    bg=self.colors['background'], fg=self.colors['text_primary']).pack(anchor="w", pady=(20, 10))
            
            time_range_var = tk.StringVar(dialog, value="本月")
            time_options = ["今日", "本周", "本月", "本季度", "本年", "全部"]
            
            time_combo = ttk.Combobox(options_frame, textvariable=time_range_var, 
                                    values=time_options, state="readonly", width=20)
            time_combo.pack(anchor="w")
            
            # 按钮框架
            btn_frame = tk.Frame(dialog, bg=self.colors['background'])
            btn_frame.pack(fill="x", padx=20, pady=20)
            
            def do_export():
                try:
                    export_type = export_type_var.get()
                    file_format = format_var.get()
                    time_range = time_range_var.get()
                    
                    # 执行导出
                    success = self.perform_export(export_type, file_format, time_range)
                    
                    if success:
                        messagebox.showinfo("导出成功", f"财务报表已成功导出为 {file_format} 格式", parent=dialog)
                        dialog.destroy()
                    else:
                        messagebox.showerror("导出失败", "导出过程中发生错误", parent=dialog)
                        
                except Exception as e:
                    messagebox.showerror("错误", f"导出失败：{e}", parent=dialog)
            
            tk.Button(btn_frame, text="📊 开始导出", command=do_export,
                     bg=self.colors['primary'], fg='white', bd=0, pady=8, padx=20,
                     font=self.fonts['body']).pack(side="left")
            tk.Button(btn_frame, text="取消", command=dialog.destroy,
                     bg=self.colors['text_secondary'], fg='white', bd=0, pady=8, padx=20,
                     font=self.fonts['body']).pack(side="right")
                     
        except Exception as e:
            root = self.main_frame.winfo_toplevel()
            messagebox.showerror("错误", f"打开导出对话框失败：{e}", parent=root)
            
    def load_fixed_costs(self):
        """加载固定成本数据"""
        try:
            if os.path.exists(self.fixed_costs_file):
                with open(self.fixed_costs_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
            else:
                return []
        except Exception as e:
            print(f"加载固定成本数据失败: {e}")
            return []
    
    def save_fixed_costs(self, costs_data):
        """保存固定成本数据"""
        try:
            # 确保目录存在
            os.makedirs(os.path.dirname(self.fixed_costs_file), exist_ok=True)
            
            with open(self.fixed_costs_file, 'w', encoding='utf-8') as f:
                json.dump(costs_data, f, ensure_ascii=False, indent=4)
            return True
        except Exception as e:
            print(f"保存固定成本数据失败: {e}")
            return False
    
    def get_costs_from_tree(self):
        """从表格中获取所有成本数据"""
        costs = []
        for item in self.costs_tree.get_children():
            values = self.costs_tree.item(item)['values']
            cost = {
                "id": f"cost_{len(costs) + 1:03d}",
                "cost_type": values[0],
                "item": values[1],
                "amount": float(str(values[2]).replace("￥", "").replace(",", "")),
                "period": values[3],
                "next_date": values[4],
                "status": values[5],
                "note": values[6] if len(values) > 6 else "",
                "created_at": datetime.datetime.now().strftime("%Y-%m-%d"),
                "updated_at": datetime.datetime.now().strftime("%Y-%m-%d")
            }
            costs.append(cost)
        return costs
    
    def save_costs_to_file(self):
        """保存当前表格中的成本数据到文件"""
        try:
            costs_data = self.get_costs_from_tree()
            self.save_fixed_costs(costs_data)
        except Exception as e:
            print(f"保存固定成本数据失败: {e}")
    
    def calculate_fixed_cost_stats(self):
        """计算固定成本统计"""
        try:
            costs_data = self.load_fixed_costs()
            
            # 初始化统计
            monthly_rent = 0
            monthly_salary = 0
            monthly_utilities = 0
            total_monthly = 0
            
            for cost in costs_data:
                amount = cost.get('amount', 0)
                period = cost.get('period', '月付')
                cost_type = cost.get('cost_type', '')
                
                # 将所有成本转换为月成本
                monthly_amount = self.convert_to_monthly(amount, period)
                total_monthly += monthly_amount
                
                # 按类型分类
                if cost_type == "租金":
                    monthly_rent += monthly_amount
                elif cost_type == "人力":
                    monthly_salary += monthly_amount
                elif cost_type in ["水电", "通讯"]:
                    monthly_utilities += monthly_amount
            
            return [
                {"title": "月租金", "value": f"￥{monthly_rent:,.0f}", "icon": "🏠", "color": self.colors['primary']},
                {"title": "员工工资", "value": f"￥{monthly_salary:,.0f}", "icon": "👥", "color": self.colors['info']},
                {"title": "水电通讯", "value": f"￥{monthly_utilities:,.0f}", "icon": "⚡", "color": self.colors['warning']},
                {"title": "总固定成本", "value": f"￥{total_monthly:,.0f}", "icon": "💼", "color": self.colors['danger']}
            ]
            
        except Exception as e:
            print(f"计算固定成本统计失败: {e}")
            # 返回默认统计
            return [
                {"title": "月租金", "value": "￥8,000", "icon": "🏠", "color": self.colors['primary']},
                {"title": "员工工资", "value": "￥15,000", "icon": "👥", "color": self.colors['info']},
                {"title": "水电通讯", "value": "￥1,200", "icon": "⚡", "color": self.colors['warning']},
                {"title": "总固定成本", "value": "￥24,200", "icon": "💼", "color": self.colors['danger']}
            ]
    
    def convert_to_monthly(self, amount, period):
        """将不同周期的成本转换为月成本"""
        try:
            # 确保 amount 不为 None 且是数值
            if amount is None:
                amount = 0
            amount = float(amount)
            
            if period == "日付":
                return amount * 30
            elif period == "周付":
                return amount * 4.33  # 一个月约4.33周
            elif period == "月付":
                return amount
            elif period == "季付":
                return amount / 3 if amount != 0 else 0
            elif period == "年付":
                return amount / 12 if amount != 0 else 0
            elif period == "一次性":
                return 0  # 一次性成本不计入月成本
            else:
                return amount  # 默认按月计算
        except (ValueError, TypeError, ZeroDivisionError):
            return 0
    
    def perform_export(self, export_type: str, file_format: str, time_range: str) -> bool:
        """执行导出操作"""
        try:
            from tkinter import filedialog
            import csv
            import datetime
            
            # 获取当前时间戳
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            
            # 根据导出类型获取数据
            if export_type == "收支记录":
                data = self.get_finance_records_for_export(time_range)
                filename = f"财务收支记录_{timestamp}"
            elif export_type == "固定成本":
                data = self.get_fixed_costs_for_export()
                filename = f"固定成本报表_{timestamp}"
            elif export_type == "财务概览":
                data = self.get_finance_overview_for_export(time_range)
                filename = f"财务概览报表_{timestamp}"
            elif export_type == "完整报表":
                data = self.get_complete_finance_report(time_range)
                filename = f"完整财务报表_{timestamp}"
            else:
                return False
            
            # 选择保存路径
            root = self.main_frame.winfo_toplevel()
            if file_format == "Excel":
                file_path = filedialog.asksaveasfilename(
                    defaultextension=".xlsx",
                    filetypes=[("Excel文件", "*.xlsx")],
                    initialname=filename
                )
                if file_path:
                    return self.export_to_excel(data, file_path, export_type)
            elif file_format == "CSV":
                file_path = filedialog.asksaveasfilename(
                    defaultextension=".csv",
                    filetypes=[("CSV文件", "*.csv")],
                    initialname=filename
                )
                if file_path:
                    return self.export_to_csv(data, file_path, export_type)
            elif file_format == "PDF":
                file_path = filedialog.asksaveasfilename(
                    defaultextension=".pdf",
                    filetypes=[("PDF文件", "*.pdf")],
                    initialname=filename
                )
                if file_path:
                    return self.export_to_pdf(data, file_path, export_type)
            
            return False
            
        except Exception as e:
            print(f"导出失败: {e}")
            return False
    
    def get_finance_records_for_export(self, time_range: str) -> List[Dict]:
        """获取收支记录数据用于导出"""
        try:
            # 从数据库获取财务记录
            records = data_manager.get_financial_records()
            
            # 根据时间范围过滤
            filtered_records = []
            current_date = datetime.datetime.now().date()
            
            for record in records:
                record_date = None
                if record.get('income_date'):
                    record_date = datetime.datetime.strptime(str(record['income_date']), '%Y-%m-%d').date()
                elif record.get('created_at'):
                    record_date = datetime.datetime.strptime(str(record['created_at'])[:10], '%Y-%m-%d').date()
                
                if record_date:
                    if time_range == "今日" and record_date == current_date:
                        filtered_records.append(record)
                    elif time_range == "本周" and (current_date - record_date).days <= 7:
                        filtered_records.append(record)
                    elif time_range == "本月" and record_date.month == current_date.month and record_date.year == current_date.year:
                        filtered_records.append(record)
                    elif time_range == "本季度":
                        quarter_start = datetime.date(current_date.year, ((current_date.month - 1) // 3) * 3 + 1, 1)
                        if record_date >= quarter_start:
                            filtered_records.append(record)
                    elif time_range == "本年" and record_date.year == current_date.year:
                        filtered_records.append(record)
                    elif time_range == "全部":
                        filtered_records.append(record)
            
            return filtered_records
        except Exception as e:
            print(f"获取财务记录失败: {e}")
            return []
    
    def get_fixed_costs_for_export(self) -> List[Dict]:
        """获取固定成本数据用于导出"""
        try:
            costs = []
            for item in self.costs_tree.get_children():
                values = self.costs_tree.item(item)['values']
                cost = {
                    "成本类型": values[0],
                    "项目": values[1],
                    "金额": values[2],
                    "周期": values[3],
                    "下次日期": values[4],
                    "状态": values[5],
                    "备注": values[6] if len(values) > 6 else ""
                }
                costs.append(cost)
            return costs
        except Exception as e:
            print(f"获取固定成本失败: {e}")
            return []
    
    def get_finance_overview_for_export(self, time_range: str) -> Dict:
        """获取财务概览数据用于导出"""
        try:
            # 获取统计数据
            stats = data_manager.get_dashboard_stats()
            
            # 获取收支记录
            records = self.get_finance_records_for_export(time_range)
            
            # 计算收入支出
            total_income = sum(r.get('amount', 0) for r in records if r.get('income_type') == 'revenue')
            total_expense = sum(r.get('amount', 0) for r in records if r.get('income_type') == 'cost')
            
            return {
                "统计信息": {
                    "今日销售额": f"￥{stats.get('today_sales', 0):.2f}",
                    "订单数量": stats.get('order_count', 0),
                    "库存预警": stats.get('low_stock_count', 0),
                    "客户总数": stats.get('customer_count', 0)
                },
                "财务概览": {
                    "总收入": f"￥{total_income:.2f}",
                    "总支出": f"￥{total_expense:.2f}",
                    "净收入": f"￥{total_income - total_expense:.2f}",
                    "时间范围": time_range
                }
            }
        except Exception as e:
            print(f"获取财务概览失败: {e}")
            return {}
    
    def get_complete_finance_report(self, time_range: str) -> Dict:
        """获取完整财务报表数据"""
        try:
            return {
                "财务概览": self.get_finance_overview_for_export(time_range),
                "收支记录": self.get_finance_records_for_export(time_range),
                "固定成本": self.get_fixed_costs_for_export()
            }
        except Exception as e:
            print(f"获取完整报表失败: {e}")
            return {}
    
    def export_to_excel(self, data: Any, file_path: str, export_type: str) -> bool:
        """导出为Excel格式"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "财务报表"
            
            # 设置标题
            title = f"智慧餐饮管理系统 - {export_type}"
            ws['A1'] = title
            ws['A1'].font = Font(size=16, bold=True)
            ws.merge_cells('A1:F1')
            
            # 设置表头样式
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            if export_type == "收支记录":
                headers = ["时间", "类型", "描述", "金额", "支付方式", "备注"]
                ws.append(headers)
                
                # 设置表头样式
                for cell in ws[2]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # 添加数据
                for record in data:
                    row = [
                        str(record.get('income_date', record.get('created_at', '')))[:10],
                        "收入" if record.get('income_type') == 'revenue' else "支出",
                        record.get('description', ''),
                        f"￥{record.get('amount', 0):.2f}",
                        "订单收入" if record.get('order_price_id') else "现金",
                        record.get('description', '')[:20]
                    ]
                    ws.append(row)
                    
            elif export_type == "固定成本":
                headers = ["成本类型", "项目", "金额", "周期", "下次日期", "状态", "备注"]
                ws.append(headers)
                
                # 设置表头样式
                for cell in ws[2]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # 添加数据
                for cost in data:
                    row = [
                        cost.get("成本类型", ""),
                        cost.get("项目", ""),
                        cost.get("金额", ""),
                        cost.get("周期", ""),
                        cost.get("下次日期", ""),
                        cost.get("状态", ""),
                        cost.get("备注", "")
                    ]
                    ws.append(row)
            
            # 调整列宽
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(file_path)
            return True
            
        except ImportError:
            messagebox.showerror("错误", "请安装openpyxl库：pip install openpyxl")
            return False
        except Exception as e:
            print(f"导出Excel失败: {e}")
            return False
    
    def export_to_csv(self, data: Any, file_path: str, export_type: str) -> bool:
        """导出为CSV格式"""
        try:
            import csv
            
            with open(file_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
                if export_type == "收支记录":
                    fieldnames = ["时间", "类型", "描述", "金额", "支付方式", "备注"]
                    writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                    writer.writeheader()
                    
                    for record in data:
                        writer.writerow({
                            "时间": str(record.get('income_date', record.get('created_at', '')))[:10],
                            "类型": "收入" if record.get('income_type') == 'revenue' else "支出",
                            "描述": record.get('description', ''),
                            "金额": f"￥{record.get('amount', 0):.2f}",
                            "支付方式": "订单收入" if record.get('order_price_id') else "现金",
                            "备注": record.get('description', '')[:20]
                        })
                        
                elif export_type == "固定成本":
                    fieldnames = ["成本类型", "项目", "金额", "周期", "下次日期", "状态", "备注"]
                    writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                    writer.writeheader()
                    
                    for cost in data:
                        writer.writerow(cost)
            
            return True
            
        except Exception as e:
            print(f"导出CSV失败: {e}")
            return False
    
    def export_to_pdf(self, data: Any, file_path: str, export_type: str) -> bool:
        """导出为PDF格式"""
        try:
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.lib.units import inch
            
            doc = SimpleDocTemplate(file_path, pagesize=A4)
            story = []
            
            # 标题样式
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                spaceAfter=30,
                alignment=1  # 居中
            )
            
            # 添加标题
            title = Paragraph(f"智慧餐饮管理系统 - {export_type}", title_style)
            story.append(title)
            story.append(Spacer(1, 20))
            
            if export_type == "收支记录":
                # 创建表格数据
                table_data = [["时间", "类型", "描述", "金额", "支付方式", "备注"]]
                
                for record in data:
                    row = [
                        str(record.get('income_date', record.get('created_at', '')))[:10],
                        "收入" if record.get('income_type') == 'revenue' else "支出",
                        record.get('description', ''),
                        f"￥{record.get('amount', 0):.2f}",
                        "订单收入" if record.get('order_price_id') else "现金",
                        record.get('description', '')[:20]
                    ]
                    table_data.append(row)
                
                # 创建表格
                table = Table(table_data)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                story.append(table)
                
            elif export_type == "固定成本":
                # 创建表格数据
                table_data = [["成本类型", "项目", "金额", "周期", "下次日期", "状态", "备注"]]
                
                for cost in data:
                    row = [
                        cost.get("成本类型", ""),
                        cost.get("项目", ""),
                        cost.get("金额", ""),
                        cost.get("周期", ""),
                        cost.get("下次日期", ""),
                        cost.get("状态", ""),
                        cost.get("备注", "")
                    ]
                    table_data.append(row)
                
                # 创建表格
                table = Table(table_data)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                story.append(table)
            
            doc.build(story)
            return True
            
        except ImportError:
            messagebox.showerror("错误", "请安装reportlab库：pip install reportlab")
            return False
        except Exception as e:
            print(f"导出PDF失败: {e}")
            return False
